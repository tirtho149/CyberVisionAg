"""
Export disease_label_subset_report.json to a formatted Excel workbook.

Sheet 1 — Summary     : one row per crop, verdict counts, progress bar chart
Sheet 2 — Diseases    : one row per disease verdict, color-coded
Sheet 3 — Similar Groups : one row per duplicate/overlapping group
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

REPORT_PATH = Path(__file__).parent / "disease_label_subset_report.json"
OUT_PATH    = Path(__file__).parent / "disease_label_subset_report.xlsx"

# ── colour palette ────────────────────────────────────────────────────────────
GREEN      = PatternFill("solid", fgColor="C6EFCE")
ORANGE     = PatternFill("solid", fgColor="FFEB9C")
RED        = PatternFill("solid", fgColor="FFC7CE")
BLUE_HDR   = PatternFill("solid", fgColor="1F4E79")
GREY_ROW   = PatternFill("solid", fgColor="F2F2F2")
WHITE_ROW  = PatternFill("solid", fgColor="FFFFFF")
PURPLE     = PatternFill("solid", fgColor="E2CFEA")

HDR_FONT   = Font(bold=True, color="FFFFFF", size=11)
BOLD       = Font(bold=True)
WRAP       = Alignment(wrap_text=True, vertical="top")
CENTER     = Alignment(horizontal="center", vertical="center")
TOP_LEFT   = Alignment(vertical="top")

VERDICT_FILL = {
    "CORRECT":      GREEN,
    "QUESTIONABLE": ORANGE,
    "INCORRECT":    RED,
}

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def header_row(ws, cols: list[str], row: int = 1):
    for c, title in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=c, value=title)
        cell.fill   = BLUE_HDR
        cell.font   = HDR_FONT
        cell.alignment = CENTER
        cell.border = thin_border()

def set_col_widths(ws, widths: dict[int, int]):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

# ── load data ─────────────────────────────────────────────────────────────────
with open(REPORT_PATH, encoding="utf-8") as f:
    records = json.load(f)

wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Summary"
ws1.freeze_panes = "A2"

cols = ["Crop", "Category", "Crop Verdict", "Total Diseases",
        "Correct", "Questionable", "Incorrect", "Similar Groups",
        "Quality %", "Summary"]
header_row(ws1, cols)
set_col_widths(ws1, {1:15, 2:14, 3:14, 4:14, 5:10, 6:14, 7:10, 8:14, 9:10, 10:60})

for i, rec in enumerate(records, start=2):
    fill = WHITE_ROW if i % 2 == 0 else GREY_ROW

    cv      = rec.get("crop_validation", {})
    dc      = rec.get("disease_check", {})
    verdicts = dc.get("disease_verdicts", [])

    total   = len(verdicts)
    correct = sum(1 for v in verdicts if v["verdict"] == "CORRECT")
    quest   = sum(1 for v in verdicts if v["verdict"] == "QUESTIONABLE")
    incorr  = sum(1 for v in verdicts if v["verdict"] == "INCORRECT")
    similar = len(dc.get("similar_groups", []))
    quality = round(100 * correct / total, 1) if total else 0

    crop_verdict = cv.get("verdict", "")
    verdict_fill = GREEN if crop_verdict == "VALID" else RED

    row_data = [
        rec.get("crop", ""),
        cv.get("category", ""),
        crop_verdict,
        total, correct, quest, incorr, similar,
        quality,
        dc.get("summary", dc.get("reason", "")),
    ]
    for c, val in enumerate(row_data, start=1):
        cell = ws1.cell(row=i, column=c, value=val)
        cell.fill      = verdict_fill if c == 3 else fill
        cell.alignment = WRAP if c == 10 else CENTER
        cell.border    = thin_border()
        # color-code count columns
        if c == 5:
            cell.fill = GREEN
        elif c == 6:
            cell.fill = ORANGE
        elif c == 7:
            cell.fill = RED
        elif c == 9:
            # gradient: red < 60, orange 60-80, green >= 80
            pct = quality
            cell.fill = GREEN if pct >= 80 else (ORANGE if pct >= 60 else RED)

ws1.row_dimensions[1].height = 22
for i in range(2, len(records) + 2):
    ws1.row_dimensions[i].height = 60

# ── stacked bar chart ─────────────────────────────────────────────────────────
chart = BarChart()
chart.type     = "bar"
chart.grouping = "stacked"
chart.title    = "Disease Label Quality by Crop"
chart.y_axis.title = "Number of Diseases"
chart.x_axis.title = "Crop"
chart.width  = 28
chart.height = 14

n = len(records)
cats = Reference(ws1, min_col=1, min_row=2, max_row=n + 1)

for col_idx, label, color in [
    (5, "Correct",      "00B050"),
    (6, "Questionable", "FFC000"),
    (7, "Incorrect",    "FF0000"),
]:
    data = Reference(ws1, min_col=col_idx, min_row=1, max_row=n + 1)
    chart.add_data(data, titles_from_data=True)
    chart.series[-1].graphicalProperties.solidFill = color
    chart.series[-1].graphicalProperties.line.solidFill = color

chart.set_categories(cats)
ws1.add_chart(chart, f"A{n + 4}")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — DISEASE VERDICTS
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Disease Verdicts")
ws2.freeze_panes = "A2"

cols2 = ["Crop", "Disease", "Verdict", "Reason"]
header_row(ws2, cols2)
set_col_widths(ws2, {1:14, 2:36, 3:14, 4:70})

row = 2
for rec in records:
    dc       = rec.get("disease_check", {})
    verdicts = dc.get("disease_verdicts", [])
    crop     = rec.get("crop", "")

    for v in verdicts:
        verdict = v.get("verdict", "")
        vfill   = VERDICT_FILL.get(verdict, WHITE_ROW)

        for c, val in enumerate([crop, v["disease"], verdict, v.get("reason", "")], start=1):
            cell = ws2.cell(row=row, column=c, value=val)
            cell.fill      = vfill
            cell.alignment = WRAP if c == 4 else TOP_LEFT
            cell.border    = thin_border()
            if c == 3:
                cell.font = BOLD

        ws2.row_dimensions[row].height = 40
        row += 1

ws2.row_dimensions[1].height = 22

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — CORRECT LABELS ONLY
# ═══════════════════════════════════════════════════════════════════════════════
ws_correct = wb.create_sheet("Correct Labels")
ws_correct.freeze_panes = "A2"

header_row(ws_correct, ["Crop", "Disease", "Reason"])
set_col_widths(ws_correct, {1:14, 2:36, 3:70})

row = 2
for rec in records:
    crop     = rec.get("crop", "")
    verdicts = rec.get("disease_check", {}).get("disease_verdicts", [])
    for v in verdicts:
        if v.get("verdict") == "CORRECT":
            for c, val in enumerate([crop, v["disease"], v.get("reason", "")], start=1):
                cell = ws_correct.cell(row=row, column=c, value=val)
                cell.fill      = GREEN
                cell.alignment = WRAP if c == 3 else TOP_LEFT
                cell.border    = thin_border()
            ws_correct.row_dimensions[row].height = 38
            row += 1

ws_correct.row_dimensions[1].height = 22

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — INCORRECT LABELS ONLY
# ═══════════════════════════════════════════════════════════════════════════════
ws_incorrect = wb.create_sheet("Incorrect Labels")
ws_incorrect.freeze_panes = "A2"

header_row(ws_incorrect, ["Crop", "Disease", "Reason"])
set_col_widths(ws_incorrect, {1:14, 2:36, 3:70})

row = 2
for rec in records:
    crop     = rec.get("crop", "")
    verdicts = rec.get("disease_check", {}).get("disease_verdicts", [])
    for v in verdicts:
        if v.get("verdict") == "INCORRECT":
            for c, val in enumerate([crop, v["disease"], v.get("reason", "")], start=1):
                cell = ws_incorrect.cell(row=row, column=c, value=val)
                cell.fill      = RED
                cell.alignment = WRAP if c == 3 else TOP_LEFT
                cell.border    = thin_border()
            ws_incorrect.row_dimensions[row].height = 38
            row += 1

ws_incorrect.row_dimensions[1].height = 22

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — QUESTIONABLE LABELS ONLY
# ═══════════════════════════════════════════════════════════════════════════════
ws_quest = wb.create_sheet("Questionable Labels")
ws_quest.freeze_panes = "A2"

header_row(ws_quest, ["Crop", "Disease", "Reason"])
set_col_widths(ws_quest, {1:14, 2:36, 3:70})

row = 2
for rec in records:
    crop     = rec.get("crop", "")
    verdicts = rec.get("disease_check", {}).get("disease_verdicts", [])
    for v in verdicts:
        if v.get("verdict") == "QUESTIONABLE":
            for c, val in enumerate([crop, v["disease"], v.get("reason", "")], start=1):
                cell = ws_quest.cell(row=row, column=c, value=val)
                cell.fill      = ORANGE
                cell.alignment = WRAP if c == 3 else TOP_LEFT
                cell.border    = thin_border()
            ws_quest.row_dimensions[row].height = 38
            row += 1

ws_quest.row_dimensions[1].height = 22

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — SIMILAR / DUPLICATE GROUPS
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Similar Groups")  # sheet 5
ws3.freeze_panes = "A2"

cols3 = ["Crop", "Group #", "Overlapping Disease Labels", "Reason"]
header_row(ws3, cols3)
set_col_widths(ws3, {1:14, 2:9, 3:50, 4:60})

row = 2
for rec in records:
    crop   = rec.get("crop", "")
    groups = rec.get("disease_check", {}).get("similar_groups", [])

    for g_idx, g in enumerate(groups, start=1):
        labels = " | ".join(g.get("diseases", []))
        reason = g.get("reason", "")

        fill = WHITE_ROW if row % 2 == 0 else GREY_ROW
        for c, val in enumerate([crop, g_idx, labels, reason], start=1):
            cell = ws3.cell(row=row, column=c, value=val)
            cell.fill      = PURPLE if c in (2, 3) else fill
            cell.alignment = WRAP
            cell.border    = thin_border()

        ws3.row_dimensions[row].height = 45
        row += 1

ws3.row_dimensions[1].height = 22

# ── save ──────────────────────────────────────────────────────────────────────
wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
