#!/usr/bin/env python3
"""
generate_symptoms.py
====================
Reads crop_disease_registry.xlsx, runs ALL THREE symptom modes for every
disease independently, and writes results back into the xlsx with one
DISCRETE column-set per mode so results are never mixed or overwritten.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
THREE MODES — run independently, stored in separate xlsx columns
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

  Mode A  ·  PDF / document extraction
    Trigger : a matching file exists in knowledge_docs/
    Method  : upload to OpenAI Files API once; call Responses API with
              input_file attached.
    Columns : Symptoms_A  |  Citations_A
    Citation: "Extracted from: <filename>"
    Skipped : leaves Symptoms_A / Citations_A as "(mode not run)" when
              no matching doc is found for the crop.

  Mode B  ·  Live web search  (web_search tool)
    Trigger : always runs (unless --no-web-search passed)
    Method  : Responses API with tool=[{"type":"web_search"}].
              Multiple independent searches are performed per disease.
              url_citation annotations are extracted and stored.
    Columns : Symptoms_B  |  Citations_B
    Citation: "[1] Title — URL\n[2] …"  (mandatory; warned if empty)

  Mode C  ·  GPT parametric knowledge
    Trigger : always runs (unless --no-gpt-generate passed)
    Method  : plain Responses API call, no tools, no files.
    Columns : Symptoms_C  |  Citations_C
    Citation: "(GPT parametric knowledge — no external sources)"

  All three modes are INDEPENDENT.  Every disease gets up to three
  separate symptom paragraphs stored in their own discrete columns.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
XLSX COLUMNS WRITTEN (Summary sheet)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  Ref Image Paths   — pipe-separated relative paths to reference images
  Symptoms_A        — Mode A: extracted from document          [blue]
  Citations_A       — Mode A: document filename                [blue]
  Symptoms_B        — Mode B: web-search-grounded paragraph   [purple]
  Citations_B       — Mode B: [n] Title — URL (one per line)  [purple]
  Symptoms_C        — Mode C: GPT parametric paragraph         [green]
  Citations_C       — Mode C: "(GPT parametric knowledge)"     [green]

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
DIRECTORY LAYOUT
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
project/
  Curated_Dataset/
    Reference_Image/
      <Crop>/
        <Disease>/      ← reference images resolved here
    Benchmark/          ← not used by this script
    manifest.json

  knowledge_docs/       ← one reference doc per crop (any format)
  results/

  crop_disease_registry.xlsx     ← INPUT  (Summary sheet read)
                                 ← OUTPUT (*_updated.xlsx written)
  generate_symptoms.py           ← THIS FILE


━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
USAGE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  python generate_symptoms.py [OPTIONS]

  --registry       PATH   crop_disease_registry.xlsx
  --curated        PATH   Curated_Dataset dir
  --docs           PATH   knowledge_docs dir
  --output         PATH   combined markdown output
  --model          MODEL  OpenAI model  (default: gpt-4o-mini)
  --no-web-search         Skip Mode B entirely
  --no-gpt-generate       Skip Mode C entirely
  --no-delete             Keep uploaded files on OpenAI after the run
  --dry-run               Print plan; no API calls

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
REQUIREMENTS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  pip install openai python-dotenv openpyxl
"""

from __future__ import annotations

import os
import re
import sys
import shutil
import argparse
import textwrap
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from dotenv import dotenv_values
from openai import OpenAI
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── load .env ──────────────────────────────────────────────────────────────────
_env_file = Path(__file__).parent / ".env"
if _env_file.exists():
    for k, v in dotenv_values(_env_file).items():
        if v:
            os.environ[k] = v

# ── constants ──────────────────────────────────────────────────────────────────
DEFAULT_MODEL       = "gpt-4o-mini"
MAX_TOKENS          = 600    # symptom paragraph calls
MAX_TOKENS_REG_A    = 8000   # Registry_A: full crop JSON (many diseases)
MAX_TOKENS_REG_B    = 2000   # Registry_B: per-disease enriched JSON
MAX_REF_IMAGES      = 3
IMAGE_EXT           = (".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".webp")
SUMMARY_SHEET       = "Summary"
COL_CROP            = "Crop"
COL_DISEASE         = "Disease"
COL_REF_PATHS       = "Ref Image Paths"

# Summary sheet only stores reference image paths now
_NEW_COLS = ["Ref Image Paths"]

# Per-column header style: fill colour, font colour
_COL_STYLES: dict[str, tuple[str, str]] = {
    "Ref Image Paths": ("4E342E", "FFFFFF"),   # brown
}

_COL_WIDTHS: dict[str, int] = {
    "Ref Image Paths": 55,
}

# ── Registry Method A columns (local doc extraction) ──────────────────────────
_REG_A_SHEET = "Local_Registry"
_REG_A_COLS  = [
    "Disease", "Pathogen", "Type", "Affected Parts",
    "Visual Description", "Confidence",
]
_REG_A_STYLES: dict[str, tuple[str, str]] = {
    "Disease":            ("1565C0", "FFFFFF"),
    "Pathogen":           ("1976D2", "FFFFFF"),
    "Type":               ("1E88E5", "FFFFFF"),
    "Affected Parts":     ("42A5F5", "000000"),
    "Visual Description": ("90CAF9", "000000"),
    "Confidence":         ("BBDEFB", "000000"),
}
_REG_A_WIDTHS: dict[str, int] = {
    "Disease": 30, "Pathogen": 30, "Type": 14,
    "Affected Parts": 25, "Visual Description": 80, "Confidence": 14,
}

# ── Registry Method B columns (web-search enriched) ───────────────────────────
_REG_B_SHEET = "Registry"
_REG_B_COLS  = [
    "Disease", "Pathogen", "Type", "Affected Parts",
    "Visual Description", "Confidence",
    "Transmission Mode", "Look-alikes",
    "Economic Importance", "Favorable Conditions",
    "Conflicts / Alternate Sources", "Citations_B",
]
_REG_B_STYLES: dict[str, tuple[str, str]] = {
    "Disease":                      ("6A1B9A", "FFFFFF"),
    "Pathogen":                     ("7B1FA2", "FFFFFF"),
    "Type":                         ("8E24AA", "FFFFFF"),
    "Affected Parts":               ("AB47BC", "FFFFFF"),
    "Visual Description":           ("CE93D8", "000000"),
    "Confidence":                   ("E1BEE7", "000000"),
    "Transmission Mode":            ("2E7D32", "FFFFFF"),
    "Look-alikes":                  ("388E3C", "FFFFFF"),
    "Economic Importance":          ("43A047", "FFFFFF"),
    "Favorable Conditions":         ("66BB6A", "000000"),
    "Conflicts / Alternate Sources":("A5D6A7", "000000"),
    "Citations_B":                  ("4A148C", "FFFFFF"),
}
_REG_B_WIDTHS: dict[str, int] = {
    "Disease": 30, "Pathogen": 30, "Type": 14,
    "Affected Parts": 25, "Visual Description": 80, "Confidence": 14,
    "Transmission Mode": 35, "Look-alikes": 40,
    "Economic Importance": 35, "Favorable Conditions": 45,
    "Conflicts / Alternate Sources": 60, "Citations_B": 65,
}

ACCEPTED_DOC_TYPES = {
    ".pdf":  "application/pdf",
    ".doc":  "application/msword",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".odt":  "application/vnd.oasis.opendocument.text",
    ".rtf":  "application/rtf",
    ".ppt":  "application/vnd.ms-powerpoint",
    ".pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    ".xls":  "application/vnd.ms-excel",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".csv":  "text/csv",   ".tsv":  "text/tsv",
    ".txt":  "text/plain", ".md":   "text/markdown",
    ".rst":  "text/x-rst", ".html": "text/html",
    ".htm":  "text/html",  ".json": "application/json",
    ".xml":  "text/xml",
}


# ══════════════════════════════════════════════════════════════════════════════
# UTILITY — safe coercion of any model-returned value to a plain string
# ══════════════════════════════════════════════════════════════════════════════

def _to_str(value) -> str:
    """
    Coerce any value that OpenAI may return for a JSON field to a plain string
    safe for openpyxl.  The model occasionally returns lists instead of strings
    (e.g. citations, look_alikes, conflicts) — join them with newlines.
    """
    if value is None:
        return ""
    if isinstance(value, list):
        return "\n".join(_to_str(item) for item in value)
    if isinstance(value, dict):
        # e.g. {"text": "..."} — just stringify it
        return str(value)
    return str(value)


# ── Result dataclasses ─────────────────────────────────────────────────────────

@dataclass
class ModeResult:
    """Output of ONE mode for ONE (crop, disease) pair."""
    symptoms:  str  = ""
    citations: str  = ""
    ran:       bool = False   # False = mode was explicitly skipped


@dataclass
class DiseaseResult:
    """All mode results + image paths for one disease."""
    A:          ModeResult = field(default_factory=ModeResult)
    B:          ModeResult = field(default_factory=ModeResult)
    C:          ModeResult = field(default_factory=ModeResult)
    ref_images: list       = field(default_factory=list)


@dataclass
class RegistryRow:
    """One disease row for Registry Method A (local doc extraction)."""
    crop:               str = ""
    disease:            str = ""
    pathogen:           str = ""
    type_of_disease:    str = ""
    affected_parts:     str = ""
    visual_description: str = ""
    confidence:         str = ""


@dataclass
class RegistryRowB:
    """One disease row for Registry Method B (web-search enriched)."""
    crop:                str = ""
    disease:             str = ""
    pathogen:            str = ""
    type_of_disease:     str = ""
    affected_parts:      str = ""
    visual_description:  str = ""
    confidence:          str = ""
    transmission_mode:   str = ""
    look_alikes:         str = ""
    economic_importance: str = ""
    favorable_conditions:str = ""
    conflicts:           str = ""
    citations:           str = ""


# ══════════════════════════════════════════════════════════════════════════════
# PROMPTS
# ══════════════════════════════════════════════════════════════════════════════

_SYSTEM_A = textwrap.dedent("""
    You are a plant pathology expert and a retriever-first diagnostic writer.
    A reference document is attached. Extract ONLY the visual symptom
    information for the specified plant disease from that document.

    STRICT RULES:
    - Use ONLY information present in the attached document. Do NOT invent.
    - Focus on observable visual features: color, shape, texture, size,
      elevation (raised/sunken/flat), halos, margins, plant-part distribution,
      and progression over time.
    - Ignore treatment, pathogen biology, management, and epidemiology.
    - End with one KEY DIFFERENCE sentence vs the most visually similar disease
      on the same crop (only if mentioned in the doc).
    - Write as ONE coherent paragraph (5-8 sentences). No bullet points.
    - If the document has NO visual symptom info for this disease:  NO_INFO
""").strip()

_SYSTEM_B = textwrap.dedent("""
    You are a plant pathology expert writing visual symptom descriptions for
    machine learning and field scouting.  You have access to a web search tool.

    MULTI-SOURCE SEARCH STRATEGY — MANDATORY:
    For each disease you MUST perform AT LEAST 3 separate searches targeting
    different authoritative source types before writing your answer:
      Search 1: university extension pages (e.g. site:extension.umn.edu,
                 extension.purdue.edu, extension.iastate.edu)
      Search 2: CABI / USDA / Crop Protection Network
      Search 3: peer-reviewed plant pathology journals or disease compendiums

    Synthesise all results. If sources disagree on any detail, note the
    conflict in your response.

    RULES:
    - Base your description ONLY on information verified via web search.
    - Focus on observable visual features: color, shape, texture, size,
      elevation, halos, margins, affected plant parts, disease progression.
    - Write as ONE coherent paragraph (5-8 sentences). No bullet points.
    - End with one KEY DIFFERENCE sentence vs the most similar disease on
      the same crop.
    - ALWAYS include inline citations for the web sources you used.
""").strip()

_PROMPT_B = textwrap.dedent("""
    Search the web using MULTIPLE queries (at least 3 from different source
    types — university extension, CABI/USDA, and peer-reviewed/compendium)
    and write a 5-8 sentence visual symptom description for
    "{disease}" disease on {crop}.

    Cover ALL of the following:
    - Distinctive visual patterns (spots, lesions, galls, coatings, cankers,
      wilting, discolouration)
    - Colour and texture of affected tissue
    - Shape, size, and distribution across the plant
    - Raised, sunken, or superficial symptoms
    - Halos, margins, spores, exudate, mycelium, sclerotia, gummosis
    - Affected plant parts (leaf, stem, root, pod, seed, etc.)
    - KEY DIFFERENCE vs the most visually similar disease on the same crop

    One coherent paragraph. Include inline citations.
    Prioritise: university extension, CABI, USDA, peer-reviewed sources.
    Cross-check at least 3 sources before finalising each detail.
""").strip()


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — Discover structure from xlsx registry
# ══════════════════════════════════════════════════════════════════════════════

def discover_from_xlsx(registry_path: Path) -> tuple[dict, dict]:
    """
    Read the Summary sheet.
    Returns:
        structure  : { "Crop": { "Disease": [] }, ... }
        row_index  : { "Crop|Disease": xlsx_row_number (1-based) }
    """
    if not registry_path.exists():
        raise FileNotFoundError(f"Registry not found: {registry_path}")

    wb = openpyxl.load_workbook(registry_path)
    if SUMMARY_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{SUMMARY_SHEET}' not found in {registry_path}.\n"
            f"Available: {wb.sheetnames}"
        )

    ws   = wb[SUMMARY_SHEET]
    rows = list(ws.iter_rows(values_only=True))

    header_row_idx, header = None, None
    for i, row in enumerate(rows):
        if row and any(str(c).strip().lower() == "crop" for c in row if c):
            header_row_idx = i
            header = [str(c).strip() if c else "" for c in row]
            break

    if header is None:
        raise ValueError("Could not find header row with 'Crop' column.")

    crop_col    = header.index(COL_CROP)
    disease_col = header.index(COL_DISEASE)

    _SKIP_CROPS    = {"NONE", "TOTAL", ""}
    _SKIP_DISEASES = {"TOTAL", "=COUNTA(C2:C231)"}

    structure, row_index = {}, {}
    for i, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
        if not row or all(c is None for c in row):
            continue
        crop    = str(row[crop_col]).strip()    if row[crop_col]    else ""
        disease = str(row[disease_col]).strip() if row[disease_col] else ""
        if not disease or disease.upper() in _SKIP_DISEASES or disease.startswith("="):
            continue
        if not crop or crop.upper() in _SKIP_CROPS or crop.startswith("="):
            continue
        structure.setdefault(crop, {})[disease] = []
        row_index[f"{crop}|{disease}"] = i

    return structure, row_index


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — Resolve reference image paths from filesystem
# ══════════════════════════════════════════════════════════════════════════════

def _folder_candidates(name: str) -> list[str]:
    return list(dict.fromkeys([
        name, name.replace(" ", "_"),
        name.lower(), name.lower().replace(" ", "_"),
    ]))


def resolve_ref_images(curated_dir: Path, crop: str, disease: str) -> list[str]:
    ref_root = curated_dir / "Reference_Image"
    if not ref_root.exists():
        return []
    crop_dir = next(
        (ref_root / v for v in _folder_candidates(crop) if (ref_root / v).is_dir()), None
    )
    if not crop_dir:
        return []
    disease_dir = next(
        (crop_dir / v for v in _folder_candidates(disease) if (crop_dir / v).is_dir()), None
    )
    if not disease_dir:
        return []
    return sorted(
        str(f.relative_to(curated_dir))
        for f in disease_dir.iterdir()
        if f.is_file() and f.suffix.lower() in IMAGE_EXT
    )[:MAX_REF_IMAGES]


def fill_ref_images(structure: dict, curated_dir: Path) -> None:
    for crop, diseases in structure.items():
        for disease in diseases:
            diseases[disease] = resolve_ref_images(curated_dir, crop, disease)


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — Upload knowledge documents  (Mode A only)
# ══════════════════════════════════════════════════════════════════════════════

def _normalise(name: str) -> str:
    return re.sub(r"[\s_\-]+", "", name.lower())


def upload_knowledge_docs(client: OpenAI, docs_dir: Path) -> tuple[list, dict]:
    file_registry, file_paths = [], {}
    if not docs_dir.exists():
        print(f"  [INFO] knowledge_docs not found: {docs_dir}")
        return file_registry, file_paths
    docs = sorted(f for f in docs_dir.iterdir()
                  if f.is_file() and f.suffix.lower() in ACCEPTED_DOC_TYPES)
    if not docs:
        print(f"  [INFO] No supported docs in {docs_dir}")
        return file_registry, file_paths
    for doc in docs:
        ext  = doc.suffix.lower()
        mime = ACCEPTED_DOC_TYPES[ext]
        print(f"  [UPLOAD] {doc.name:<50} → ", end="", flush=True)
        try:
            with open(doc, "rb") as fh:
                resp = client.files.create(file=(doc.name, fh, mime), purpose="user_data")
            fid = resp.id
            file_registry.append({"file_id": fid, "path": doc, "stem_norm": _normalise(doc.stem)})
            file_paths[fid] = doc
            print(f"{fid}  [{ext}]")
        except Exception as exc:
            print(f"FAILED ({exc})")
    return file_registry, file_paths


def delete_uploaded_files(client: OpenAI, file_paths: dict) -> None:
    for fid, path in file_paths.items():
        try:
            client.files.delete(fid)
            print(f"  [DELETE] {fid}  ({path.name})")
        except Exception as exc:
            print(f"  [WARN] Could not delete {fid}: {exc}")


def find_file_ids_for_category(category: str, file_registry: list) -> list[str]:
    crop_words = [w for w in re.split(r"[_\s]+", category.lower()) if len(w) > 3]
    return [e["file_id"] for e in file_registry
            if any(w in e["stem_norm"] for w in crop_words)]


def doc_names_for_ids(file_ids: list[str], file_registry: list) -> list[str]:
    id_set = set(file_ids)
    return [e["path"].name for e in file_registry if e["file_id"] in id_set]


# ══════════════════════════════════════════════════════════════════════════════
# LOW-LEVEL Responses API wrapper
# ══════════════════════════════════════════════════════════════════════════════

def _call_responses(
    client: OpenAI,
    model: str,
    system: str,
    user_content: list,
    tools: Optional[list] = None,
    max_tokens: int = MAX_TOKENS,
) -> tuple[str, list]:
    """
    Returns (text, annotations).
    annotations = list of url_citation dicts from web_search output.
    """
    kwargs: dict = dict(
        model=model,
        instructions=system,
        input=[{"role": "user", "content": user_content}],
        max_output_tokens=max_tokens,
    )
    if tools:
        kwargs["tools"] = tools

    response    = client.responses.create(**kwargs)
    text_parts  = []
    annotations = []

    for item in response.output:
        if not hasattr(item, "content"):
            continue
        for block in item.content:
            if hasattr(block, "text"):
                text_parts.append(block.text)
            if hasattr(block, "annotations"):
                for ann in block.annotations:
                    ann_d = ann if isinstance(ann, dict) else vars(ann)
                    if ann_d.get("type") == "url_citation":
                        annotations.append(ann_d)

    return "\n".join(text_parts).strip(), annotations


# ══════════════════════════════════════════════════════════════════════════════
# MODE A — PDF / document extraction
# ══════════════════════════════════════════════════════════════════════════════

def run_mode_a(
    client: OpenAI, model: str,
    crop: str, disease: str,
    file_ids: list[str],
    doc_names: list[str],
) -> ModeResult:
    """
    Extract symptoms from uploaded doc.
    Returns ModeResult(ran=False) when no docs are available for this crop.
    Returns ran=True with empty symptoms when the disease is absent (NO_INFO).
    """
    if not file_ids:
        return ModeResult(ran=False)

    clean_crop    = crop.replace("_", " ")
    clean_disease = disease.replace("_", " ")

    user_content = [{"type": "input_file", "file_id": fid} for fid in file_ids]
    user_content.append({
        "type": "input_text",
        "text": (
            f"Crop: {clean_crop}\n"
            f"Disease: {clean_disease}\n\n"
            f"Using the attached reference document(s), extract the visual "
            f"symptom description for '{clean_disease}' on {clean_crop}.\n\n"
            f"Look carefully — the disease may appear under a slightly different "
            f"name, as part of a group, or in a table. Only reply NO_INFO if the "
            f"disease is truly absent from all attached documents."
        ),
    })

    text, _ = _call_responses(client, model, _SYSTEM_A, user_content)

    if text.strip() == "NO_INFO" or not text.strip():
        return ModeResult(
            symptoms  = "",
            citations = f"Searched: {', '.join(doc_names)} — disease not found in document",
            ran       = True,
        )

    return ModeResult(
        symptoms  = text,
        citations = "Extracted from: " + ", ".join(doc_names),
        ran       = True,
    )


# ══════════════════════════════════════════════════════════════════════════════
# MODE B — Multi-source web search with mandatory citation extraction
# ══════════════════════════════════════════════════════════════════════════════

def _format_web_citations(annotations: list[dict]) -> str:
    """
    Deduplicate url_citation annotations and return a numbered citation list:
        [1] Page Title — https://example.com
        [2] Another Source — https://another.org
    Falls back to a descriptive string when annotations are empty.
    """
    seen, lines = set(), []
    for ann in annotations:
        url   = ann.get("url", "").strip()
        title = (ann.get("title") or url).strip()
        if not url or url in seen:
            continue
        seen.add(url)
        lines.append(f"[{len(lines)+1}] {title} — {url}")
    return "\n".join(lines) if lines else "(web search ran — no url_citation annotations returned)"


def run_mode_b(
    client: OpenAI, model: str,
    crop: str, disease: str,
) -> ModeResult:
    """
    Multi-source live web search via web_search tool.
    The system prompt instructs the model to perform ≥3 searches from
    different authoritative source categories.
    url_citation annotations are mandatory; warns if none returned.
    """
    clean_crop    = crop.replace("_", " ")
    clean_disease = disease.replace("_", " ")

    text, annotations = _call_responses(
        client, model, _SYSTEM_B,
        user_content=[{
            "type": "input_text",
            "text": _PROMPT_B.format(disease=clean_disease, crop=clean_crop),
        }],
        tools=[{"type": "web_search"}],
    )

    if not annotations:
        print("\n        [WARN B] No url_citation annotations returned.", end="")

    return ModeResult(
        symptoms  = text,
        citations = _format_web_citations(annotations),
        ran       = True,
    )


# ══════════════════════════════════════════════════════════════════════════════
# WRITE BACK — save all discrete columns into xlsx
# ══════════════════════════════════════════════════════════════════════════════

def write_results_to_xlsx(
    registry_path: Path,
    row_index: dict,
    all_results: dict,
) -> Path:
    """
    Add / update discrete columns in the Summary sheet.
    Saves as *_updated.xlsx; backs up original as *_backup.xlsx.
    """
    backup = registry_path.with_name(registry_path.stem + "_backup" + registry_path.suffix)
    shutil.copy2(registry_path, backup)
    print(f"\n  [BACKUP] Original  → {backup.name}")

    wb = openpyxl.load_workbook(registry_path)
    ws = wb[SUMMARY_SHEET]

    # Find header row
    header_row_num = None
    for row in ws.iter_rows(min_row=1, max_row=15):
        for cell in row:
            if cell.value and str(cell.value).strip().lower() == "crop":
                header_row_num = cell.row
                break
        if header_row_num:
            break
    if header_row_num is None:
        raise ValueError("Cannot find header row in Summary sheet.")

    # Map existing header names → column index
    existing: dict[str, int] = {
        str(ws.cell(row=header_row_num, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=header_row_num, column=c).value
    }

    # Create or reuse each target column, in fixed order
    col_map: dict[str, int] = {}
    next_col = ws.max_column + 1

    for col_name in _NEW_COLS:
        if col_name in existing:
            col_map[col_name] = existing[col_name]
        else:
            fill_hex, font_hex = _COL_STYLES.get(col_name, ("555555", "FFFFFF"))
            cell = ws.cell(row=header_row_num, column=next_col)
            cell.value     = col_name
            cell.font      = Font(bold=True, color=font_hex)
            cell.fill      = PatternFill("solid", fgColor=fill_hex)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            col_map[col_name] = next_col
            next_col += 1

    # Set column widths
    for col_name, width in _COL_WIDTHS.items():
        if col_name in col_map:
            ws.column_dimensions[get_column_letter(col_map[col_name])].width = width

    # Write data — Summary only stores Ref Image Paths
    for key, xlsx_row in row_index.items():
        res = all_results.get(key)
        if res is None:
            continue
        _set_cell(ws, xlsx_row, col_map[COL_REF_PATHS],
                  " | ".join(res.ref_images) if res.ref_images else "")

    # If the input is already named *_updated.xlsx, write back to the same file
    stem = registry_path.stem
    if stem.endswith("_updated"):
        out_path = registry_path
    else:
        out_path = registry_path.with_name(stem + "_updated" + registry_path.suffix)
    wb.save(out_path)
    return out_path


def _set_cell(ws, row: int, col: int, value: str) -> None:
    c           = ws.cell(row=row, column=col)
    c.value     = _to_str(value)   # guarantee openpyxl-safe scalar
    c.alignment = Alignment(wrap_text=True, vertical="top")


# ══════════════════════════════════════════════════════════════════════════════
# INTERACTIVE SELECTION — let user pick which diseases to process
# ══════════════════════════════════════════════════════════════════════════════

def _print_registry_table(structure: dict) -> list[tuple[int, str, str]]:
    flat: list[tuple[int, str, str]] = []
    idx = 1
    for crop, diseases in structure.items():
        for disease in diseases:
            flat.append((idx, crop, disease))
            idx += 1

    current_crop = None
    for i, crop, disease in flat:
        if crop != current_crop:
            current_crop = crop
            print(f"\n  ── {crop} ──")
        print(f"    [{i:>3}]  {disease}")

    return flat


def _parse_selection(raw: str, max_idx: int) -> list[int]:
    raw = raw.strip().lower()
    if raw in ("all", "*"):
        return list(range(1, max_idx + 1))

    indices: set[int] = set()
    for token in re.split(r"[,\s]+", raw):
        token = token.strip()
        if not token:
            continue
        range_match = re.fullmatch(r"(\d+)-(\d+)", token)
        if range_match:
            lo, hi = int(range_match.group(1)), int(range_match.group(2))
            indices.update(range(lo, hi + 1))
        elif token.isdigit():
            indices.add(int(token))

    return sorted(i for i in indices if 1 <= i <= max_idx)


def select_diseases(structure: dict) -> dict:
    print("\n" + "═" * 70)
    print("  DISEASE SELECTION")
    print("═" * 70)

    total = sum(len(d) for d in structure.values())
    print(f"\n  Registry contains {len(structure)} crop(s) and {total} disease class(es).")
    print("  Choose how you want to select:\n")
    print("    [1]  Pick specific diseases by number / range")
    print("    [2]  Pick entire crop(s) by name")
    print("    [3]  Process ALL diseases  (original behaviour)\n")

    while True:
        strategy = input("  Your choice (1 / 2 / 3): ").strip()
        if strategy in ("1", "2", "3"):
            break
        print("  Please enter 1, 2, or 3.")

    if strategy == "3":
        print(f"\n  ✓ All {total} disease(s) selected.\n")
        return structure

    if strategy == "1":
        print("\n  Available diseases:\n")
        flat = _print_registry_table(structure)
        print(f"\n  Enter indices (e.g.  3   or   1,4,7   or   5-10   or   all):")

        while True:
            raw = input("  Selection: ").strip()
            if not raw:
                print("  Selection cannot be empty.")
                continue
            chosen = _parse_selection(raw, len(flat))
            if not chosen:
                print("  No valid indices found — please try again.")
                continue
            break

        selected_pairs: list[tuple[str, str]] = [(flat[i-1][1], flat[i-1][2]) for i in chosen]

    else:
        crop_names = list(structure.keys())
        print("\n  Available crops:\n")
        for i, crop in enumerate(crop_names, 1):
            n = len(structure[crop])
            print(f"    [{i:>3}]  {crop}  ({n} disease(s))")

        print(f"\n  Enter crop indices (e.g.  2   or   1,3   or   all):")
        while True:
            raw = input("  Selection: ").strip()
            if not raw:
                print("  Selection cannot be empty.")
                continue
            chosen_crop_idxs = _parse_selection(raw, len(crop_names))
            if not chosen_crop_idxs:
                print("  No valid indices found — please try again.")
                continue
            break

        chosen_crops = [crop_names[i-1] for i in chosen_crop_idxs]
        selected_pairs = [
            (crop, disease)
            for crop in chosen_crops
            for disease in structure[crop]
        ]

    filtered: dict[str, dict] = {}
    for crop, disease in selected_pairs:
        filtered.setdefault(crop, {})[disease] = structure[crop][disease]

    n_selected = sum(len(d) for d in filtered.values())
    print(f"\n  ✓ {n_selected} disease(s) selected across {len(filtered)} crop(s):\n")
    for crop, diseases in filtered.items():
        print(f"    {crop}:")
        for disease in diseases:
            print(f"      • {disease}")
    print()

    while True:
        confirm = input("  Proceed with this selection? [y/n]: ").strip().lower()
        if confirm in ("y", "yes"):
            break
        if confirm in ("n", "no"):
            print("\n  Restarting selection …")
            return select_diseases(structure)
        print("  Please enter y or n.")

    print()
    return filtered


# ══════════════════════════════════════════════════════════════════════════════
# REGISTRY METHOD A — extract structured registry from local docs
# ══════════════════════════════════════════════════════════════════════════════

_SYSTEM_REG_A = textwrap.dedent("""
    You are a plant pathology expert. A reference document is attached.
    Extract structured information for ONE specific disease from that document.
    Return ONLY valid JSON. No markdown, no code fences, no preamble.

    JSON schema:
    {
      "disease": "string",
      "pathogen": "string",
      "type": "Fungal | Bacterial | Viral | Nematode | Oomycete | Other",
      "affected_parts": "comma-separated list as a single string",
      "visual_description": "paragraph extracted from the document",
      "confidence": "high | medium | low"
    }

    RULES:
    - Use ONLY information present in the attached document.
    - ALL field values MUST be plain strings — never arrays or objects.
    - If the disease is not found in the document, return:
      {"disease": "<name>", "pathogen": "", "type": "", "affected_parts": "",
       "visual_description": "", "confidence": "low"}
    - Do NOT invent pathogens or descriptions.
    - visual_description: combine all visual symptom sentences from the doc into
      one coherent paragraph. No bullet points.
    - confidence: "high" if full description present, "medium" if partial,
      "low" if name only or not found.
""").strip()


def run_registry_mode_a(
    client: OpenAI,
    model: str,
    crop: str,
    diseases: list[str],
    file_ids: list[str],
    doc_names: list[str],
):
    """
    Generator — yields one RegistryRow per disease as soon as it is ready.
    """
    if not file_ids or not diseases:
        return

    import json as _json

    clean_crop = crop.replace("_", " ")
    file_blocks = [{"type": "input_file", "file_id": fid} for fid in file_ids]

    def _parse(text: str) -> dict:
        text = re.sub(r"```(?:json)?|```", "", text).strip()
        try:
            return _json.loads(text)
        except _json.JSONDecodeError:
            m = re.search(r'\{[\s\S]*\}', text)
            if m:
                try:
                    return _json.loads(m.group())
                except _json.JSONDecodeError:
                    pass
        return {}

    for disease in diseases:
        clean_disease = disease.replace("_", " ")
        print(f"      ↳ {clean_disease:<50}", end="", flush=True)

        user_content = file_blocks + [{
            "type": "input_text",
            "text": (
                f"Crop: {clean_crop}\n"
                f"Disease: {clean_disease}\n\n"
                f"Extract the registry entry for '{clean_disease}' on {clean_crop} "
                f"from the attached document. "
                f"Return ONLY the JSON object as described in the system prompt. "
                f"All field values must be plain strings, never arrays."
            ),
        }]

        try:
            text, _ = _call_responses(
                client, model, _SYSTEM_REG_A, user_content,
                max_tokens=MAX_TOKENS_REG_A,
            )
            data = _parse(text)

            if not data:
                print("⚠ no JSON")
                continue

            conf  = _to_str(data.get("confidence", "?"))
            found = "✓" if data.get("visual_description") else "∅ not in doc"
            print(f"{found}  [{conf}]")

            yield RegistryRow(
                crop               = clean_crop,
                disease            = _to_str(data.get("disease", clean_disease)),
                pathogen           = _to_str(data.get("pathogen", "")),
                type_of_disease    = _to_str(data.get("type", "")),
                affected_parts     = _to_str(data.get("affected_parts", "")),
                visual_description = _to_str(data.get("visual_description", "")),
                confidence         = _to_str(data.get("confidence", "")),
            )

        except Exception as exc:
            print(f"✗ error: {exc}")


# ══════════════════════════════════════════════════════════════════════════════
# REGISTRY METHOD B — multi-source web-search enriched registry
# ══════════════════════════════════════════════════════════════════════════════

_SYSTEM_REG_B = textwrap.dedent("""
    You are a plant pathology expert with access to web search.

    MULTI-SOURCE SEARCH STRATEGY — MANDATORY:
    Before writing your answer you MUST perform AT LEAST 3 separate searches
    from different source categories:
      Search 1: university extension pages (umn.edu, purdue.edu, iastate.edu,
                 ucdavis.edu, ncsu.edu, etc.)
      Search 2: CABI Crop Protection Compendium, USDA, or Crop Protection Network
      Search 3: peer-reviewed plant pathology journals or disease compendiums

    Cross-check all details across sources. If sources disagree on any field,
    record the conflict in the "conflicts" field.

    Return ONLY valid JSON. No markdown, no code fences, no preamble.
    ALL field values MUST be plain strings — never arrays or nested objects.

    JSON schema:
    {
      "disease": "string",
      "pathogen": "string",
      "type": "Fungal | Bacterial | Viral | Nematode | Oomycete | Other",
      "affected_parts": "comma-separated list as a single string",
      "visual_description": "5-8 sentence paragraph, visual features only",
      "confidence": "high | medium | low",
      "transmission_mode": "string",
      "look_alikes": "comma-separated disease names as a single string",
      "economic_importance": "string (yield loss %, management priority)",
      "favorable_conditions": "string (temperature, humidity, soil, timing)",
      "conflicts": "any contradictions across sources, or 'None' if all agree",
      "citations": "numbered list as a single string: [1] Title — URL\\n[2] Title — URL"
    }

    RULES:
    - visual_description: observable features only (color, shape, texture,
      elevation, halos, margins, distribution, progression). One paragraph.
    - confidence: "high" if ≥2 authoritative sources agree, "medium" if partial,
      "low" if limited info found.
    - citations: must list ALL sources used across all searches.
""").strip()

_PROMPT_REG_B = (
    "Crop: {crop}\n"
    "Disease: {disease}\n\n"
    "Perform at least 3 web searches from different authoritative source types "
    "(university extension, CABI/USDA, peer-reviewed) and return the full JSON "
    "registry entry for '{disease}' on {crop} as described in the system prompt. "
    "All JSON field values must be plain strings, never arrays."
)


def run_registry_mode_b(
    client: OpenAI,
    model: str,
    crop: str,
    disease: str,
) -> RegistryRowB:
    """
    Multi-source web-search enriched registry row for one (crop, disease) pair.
    All returned fields are normalised to plain strings via _to_str() so that
    openpyxl never receives a list or dict.
    """
    import json as _json
    clean_crop    = crop.replace("_", " ")
    clean_disease = disease.replace("_", " ")

    def _safe_parse(text: str) -> dict:
        # Strip markdown code fences if present
        text = re.sub(r"```(?:json)?|```", "", text).strip()
        try:
            return _json.loads(text)
        except _json.JSONDecodeError:
            pass
        m = re.search(r'\{[\s\S]*\}', text)
        if m:
            try:
                return _json.loads(m.group())
            except _json.JSONDecodeError:
                pass
        return {}

    try:
        text, annotations = _call_responses(
            client, model, _SYSTEM_REG_B,
            user_content=[{
                "type": "input_text",
                "text": _PROMPT_REG_B.format(crop=clean_crop, disease=clean_disease),
            }],
            tools=[{"type": "web_search"}],
            max_tokens=MAX_TOKENS_REG_B,
        )

        data = _safe_parse(text)

        if not data:
            print(f"\n        [WARN Registry-B] No parseable JSON for {crop}/{disease}")
            return RegistryRowB(crop=clean_crop, disease=clean_disease)

        # ── Build citations string ─────────────────────────────────────────
        # Start from the JSON "citations" field (normalise list → str)
        cit_text = _to_str(data.get("citations", ""))

        # Enrich / override with actual url_citation annotations from the API
        if annotations:
            ann_lines = _format_web_citations(annotations)
            if ann_lines and "(web search ran" not in ann_lines:
                # Merge: keep JSON citations if annotations duplicate them
                if not cit_text:
                    cit_text = ann_lines
                else:
                    # Append annotation citations that aren't already listed
                    existing_urls = set(re.findall(r"https?://\S+", cit_text))
                    extra = [
                        line for line in ann_lines.splitlines()
                        if not any(u in line for u in existing_urls)
                    ]
                    if extra:
                        cit_text = cit_text + "\n" + "\n".join(extra)

        # ── Return row with every field coerced to a safe string ──────────
        return RegistryRowB(
            crop                 = clean_crop,
            disease              = _to_str(data.get("disease",              clean_disease)),
            pathogen             = _to_str(data.get("pathogen",             "")),
            type_of_disease      = _to_str(data.get("type",                 "")),
            affected_parts       = _to_str(data.get("affected_parts",       "")),
            visual_description   = _to_str(data.get("visual_description",   "")),
            confidence           = _to_str(data.get("confidence",           "")),
            transmission_mode    = _to_str(data.get("transmission_mode",    "")),
            look_alikes          = _to_str(data.get("look_alikes",          "")),
            economic_importance  = _to_str(data.get("economic_importance",  "")),
            favorable_conditions = _to_str(data.get("favorable_conditions", "")),
            conflicts            = _to_str(data.get("conflicts",            "")),
            citations            = cit_text,
        )

    except Exception as exc:
        print(f"\n        [WARN Registry-B] Error for {crop}/{disease}: {exc}")
        return RegistryRowB(crop=clean_crop, disease=clean_disease)


# ══════════════════════════════════════════════════════════════════════════════
# WRITE REGISTRY SHEETS — write Registry_A and Registry_B as new xlsx sheets
# ══════════════════════════════════════════════════════════════════════════════

def _write_registry_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    col_names: list[str],
    col_styles: dict[str, tuple[str, str]],
    col_widths: dict[str, int],
    rows: list,
    row_attr_map: list[str],
    crop_col: bool = True,
) -> None:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    all_cols  = (["Crop"] + col_names)  if crop_col else col_names
    all_attrs = (["crop"] + row_attr_map) if crop_col else row_attr_map

    for ci, col_name in enumerate(all_cols, start=1):
        fill_hex, font_hex = col_styles.get(col_name, ("444444", "FFFFFF"))
        cell = ws.cell(row=1, column=ci)
        cell.value     = col_name
        cell.font      = Font(bold=True, color=font_hex)
        cell.fill      = PatternFill("solid", fgColor=fill_hex)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col_name, 20)

    for ri, row_obj in enumerate(rows, start=2):
        for ci, attr in enumerate(all_attrs, start=1):
            raw = getattr(row_obj, attr, "") or ""
            val = _to_str(raw)   # ← always a safe scalar
            cell = ws.cell(row=ri, column=ci)
            cell.value     = val
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"


def write_registry_sheets(
    registry_path: Path,
    out_path: Path,
    reg_a_rows: list[RegistryRow],
    reg_b_rows: list[RegistryRowB],
    run_reg_a: bool = True,
    run_reg_b: bool = True,
) -> None:
    wb = openpyxl.load_workbook(out_path)

    if run_reg_a:
        _write_registry_sheet(
            wb, _REG_A_SHEET,
            col_names    = _REG_A_COLS,
            col_styles   = _REG_A_STYLES,
            col_widths   = _REG_A_WIDTHS,
            rows         = reg_a_rows,
            row_attr_map = ["disease", "pathogen", "type_of_disease",
                            "affected_parts", "visual_description", "confidence"],
        )
        print(f"  ✅ Sheet '{_REG_A_SHEET}' written ({len(reg_a_rows)} rows)")

    if run_reg_b:
        _write_registry_sheet(
            wb, _REG_B_SHEET,
            col_names    = _REG_B_COLS,
            col_styles   = _REG_B_STYLES,
            col_widths   = _REG_B_WIDTHS,
            rows         = reg_b_rows,
            row_attr_map = ["disease", "pathogen", "type_of_disease",
                            "affected_parts", "visual_description", "confidence",
                            "transmission_mode", "look_alikes",
                            "economic_importance", "favorable_conditions",
                            "conflicts", "citations"],
        )
        print(f"  ✅ Sheet '{_REG_B_SHEET}' written ({len(reg_b_rows)} rows)")

    wb.save(out_path)


def _flush_registry_sheets(
    xlsx_out: Path,
    reg_a_rows: list,
    reg_b_rows: list,
    run_reg_a: bool,
    run_reg_b: bool,
) -> None:
    """Rewrite Registry sheets in-place after each disease — real-time updates."""
    try:
        wb = openpyxl.load_workbook(xlsx_out)

        if run_reg_a:
            if _REG_A_SHEET in wb.sheetnames:
                del wb[_REG_A_SHEET]
            _write_registry_sheet(
                wb, _REG_A_SHEET,
                col_names    = _REG_A_COLS,
                col_styles   = _REG_A_STYLES,
                col_widths   = _REG_A_WIDTHS,
                rows         = reg_a_rows,
                row_attr_map = ["disease", "pathogen", "type_of_disease",
                                "affected_parts", "visual_description", "confidence"],
            )

        if run_reg_b:
            if _REG_B_SHEET in wb.sheetnames:
                del wb[_REG_B_SHEET]
            _write_registry_sheet(
                wb, _REG_B_SHEET,
                col_names    = _REG_B_COLS,
                col_styles   = _REG_B_STYLES,
                col_widths   = _REG_B_WIDTHS,
                rows         = reg_b_rows,
                row_attr_map = ["disease", "pathogen", "type_of_disease",
                                "affected_parts", "visual_description", "confidence",
                                "transmission_mode", "look_alikes",
                                "economic_importance", "favorable_conditions",
                                "conflicts", "citations"],
            )

        wb.save(xlsx_out)
    except Exception as exc:
        print(f"  [WARN] real-time flush failed: {exc}")


# ══════════════════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════════════════

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description=(
            "Build disease registry xlsx:\n"
            "  Summary   — Ref Image Paths only\n"
            "  Sheet 2   — Local_Registry (local doc extraction)\n"
            "  Sheet 3   — Registry (multi-source web-search enriched)\n"
        ),
        formatter_class=argparse.RawTextHelpFormatter,
    )
    p.add_argument("--registry",      default="./crop_disease_registry.xlsx")
    p.add_argument("--curated",       default="./Curated_Dataset")
    p.add_argument("--docs",          default="./knowledge_docs")
    p.add_argument("--model",         default=DEFAULT_MODEL)
    p.add_argument("--no-registry-a", action="store_true", help="Skip Registry_A sheet")
    p.add_argument("--no-registry-b", action="store_true", help="Skip Registry_B sheet")
    p.add_argument("--no-delete",     action="store_true")
    p.add_argument("--dry-run",       action="store_true")
    return p.parse_args()


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    args      = parse_args()
    registry  = Path(args.registry)
    curated   = Path(args.curated)
    docs      = Path(args.docs)
    model     = args.model
    run_reg_a = not args.no_registry_a
    run_reg_b = not args.no_registry_b

    print("=" * 70)
    print("  DISEASE REGISTRY BUILDER")
    print("=" * 70)
    print(f"  Registry xlsx          : {registry}")
    print(f"  Curated dataset        : {curated}")
    print(f"  Knowledge docs         : {docs}")
    print(f"  Model                  : {model}")
    print(f"  Summary sheet          : Ref Image Paths only")
    print(f"  Registry_A (sheet 2)   : {'ENABLED  ← local doc extraction'       if run_reg_a else 'DISABLED'}")
    print(f"  Registry_B (sheet 3)   : {'ENABLED  ← multi-source web-search'    if run_reg_b else 'DISABLED'}")
    print("=" * 70)

    # ── Step 1: read xlsx ─────────────────────────────────────────────────
    print(f"\n[1/4] Reading registry …\n")
    try:
        structure, row_index = discover_from_xlsx(registry)
    except (FileNotFoundError, ValueError) as exc:
        print(f"\n[ERROR] {exc}")
        sys.exit(1)

    if not structure:
        print("[ERROR] No entries found in registry.")
        sys.exit(1)

    total = sum(len(d) for d in structure.values())
    print(f"  {len(structure)} crop(s), {total} disease class(es):\n")
    for crop, diseases in structure.items():
        print(f"    {crop:<42}  {len(diseases):3d} disease(s)")

    # ── Step 2: reference images ──────────────────────────────────────────
    print(f"\n[2/4] Resolving reference images …\n")
    if curated.exists():
        fill_ref_images(structure, curated)
        found = sum(1 for d in structure.values() for imgs in d.values() if imgs)
        print(f"  {found}/{total} disease(s) have reference images.")
    else:
        print(f"  [WARN] {curated} not found — no images will be resolved.")

    # ── Interactive disease selection ─────────────────────────────────────
    if not args.dry_run:
        structure = select_diseases(structure)
        total = sum(len(d) for d in structure.values())

    if args.dry_run:
        print("\n[DRY RUN] No API calls made.")
        print(f"\n  Summary sheet column : Ref Image Paths")
        print(f"\n  Registry sheets that would be created:")
        if run_reg_a:
            print(f"    {_REG_A_SHEET} — Crop, {', '.join(_REG_A_COLS)}")
        if run_reg_b:
            print(f"    {_REG_B_SHEET} — Crop, {', '.join(_REG_B_COLS)}")
        return

    # ── OpenAI client ──────────────────────────────────────────────────────
    if not os.environ.get("OPENAI_API_KEY"):
        print("[ERROR] OPENAI_API_KEY not set.")
        sys.exit(1)
    client = OpenAI()

    # ── Step 3: upload docs (for Registry_A) ──────────────────────────────
    print(f"\n[3/4] Uploading knowledge docs (Registry_A) …\n")
    file_registry, file_paths = upload_knowledge_docs(client, docs)
    print(f"\n  {len(file_registry)} doc(s) uploaded." if file_registry
          else "  No docs — Registry_A will be skipped for all crops.")

    # ── Step 4: resolve ref images and write Summary ───────────────────────
    print(f"\n[4/4] Writing Summary sheet (Ref Image Paths) …\n")

    all_results: dict[str, DiseaseResult] = {}
    for crop, diseases in structure.items():
        for disease, images in diseases.items():
            key = f"{crop}|{disease}"
            res = DiseaseResult()
            res.ref_images = images
            all_results[key] = res

    try:
        xlsx_out = write_results_to_xlsx(registry, row_index, all_results)
        print(f"  ✅ Summary xlsx saved → {xlsx_out.resolve()}")
    except Exception as exc:
        print(f"  [ERROR] xlsx write failed: {exc}")
        xlsx_out = None

    # ── Registry_A — local doc extraction ─────────────────────────────────
    reg_a_rows: list[RegistryRow]  = []
    reg_b_rows: list[RegistryRowB] = []

    if run_reg_a and xlsx_out:
        print(f"\n  [Registry_A] Extracting per-disease from local docs …\n")
        for crop, diseases in structure.items():
            crop_file_ids  = find_file_ids_for_category(crop, file_registry)
            crop_doc_names = doc_names_for_ids(crop_file_ids, file_registry)
            if not crop_file_ids:
                print(f"    {crop}: no docs — skipped")
                continue
            print(f"    {crop} ({', '.join(crop_doc_names)}) — {len(diseases)} disease(s):")
            n_before = len(reg_a_rows)
            for row in run_registry_mode_a(
                client, model, crop,
                diseases=list(diseases.keys()),
                file_ids=crop_file_ids,
                doc_names=crop_doc_names,
            ):
                reg_a_rows.append(row)
                _flush_registry_sheets(xlsx_out, reg_a_rows, reg_b_rows, run_reg_a, run_reg_b)
            print(f"    → {len(reg_a_rows) - n_before} row(s) extracted")

    # ── Registry_B — multi-source web-search enriched ─────────────────────
    if run_reg_b and xlsx_out:
        print(f"\n  [Registry_B] Building multi-source web-enriched registry …\n")
        for crop, diseases in structure.items():
            for disease in diseases:
                print(f"    {crop} / {disease:<50}", end="", flush=True)
                row = run_registry_mode_b(client, model, crop, disease)
                reg_b_rows.append(row)
                print("✓")
                _flush_registry_sheets(xlsx_out, reg_a_rows, reg_b_rows, run_reg_a, run_reg_b)

    # ── Write final Registry sheets ────────────────────────────────────────
    if xlsx_out:
        print(f"\n  Writing Registry sheets …")
        try:
            write_registry_sheets(
                registry, xlsx_out, reg_a_rows, reg_b_rows,
                run_reg_a=run_reg_a, run_reg_b=run_reg_b,
            )
        except Exception as exc:
            print(f"  [ERROR] Registry sheet write failed: {exc}")

    # ── Summary ────────────────────────────────────────────────────────────
    print(f"\n{'='*70}")
    print(f"  ✅ DONE")
    print(f"     Summary sheet          : {total} rows — Ref Image Paths written")
    if run_reg_a:
        print(f"     Registry_A (sheet 2)   : {len(reg_a_rows)} disease rows")
    if run_reg_b:
        print(f"     Registry_B (sheet 3)   : {len(reg_b_rows)} disease rows")
    if xlsx_out:
        print(f"     Output file            : {xlsx_out.resolve()}")
    print(f"{'='*70}\n")

    # ── Cleanup ────────────────────────────────────────────────────────────
    if file_paths and not args.no_delete:
        print("Cleaning up uploaded files …\n")
        delete_uploaded_files(client, file_paths)


if __name__ == "__main__":
    main()
