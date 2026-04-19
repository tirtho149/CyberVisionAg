"""Open Agentic Pipeline — Evaluation Harness.

Spawns parallel `claude -p` agents for plant disease classification.
Each agent freely reasons over test images, reference images, and symptom KB.

Usage:
    python -m CyberVisionAg.open_agentic.eval --symptom-source default --quick-test 2
"""

import argparse
import json
import os
import re
import shutil
import signal
import subprocess
import sys
import tempfile
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from .prompt import build_system_prompt, build_user_message

# ── Paths ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
CYBERVISION_DIR = SCRIPT_DIR.parent
PROJECT_ROOT = CYBERVISION_DIR.parent

DATASET_ROOT = CYBERVISION_DIR / "Curated_Local_Dataset"
TRAIN_DIR = DATASET_ROOT / "train"
TEST_DIR = DATASET_ROOT / "test"
SYMPTOMS_FILE = CYBERVISION_DIR / "disease_symptoms_crop_wise.md"
REGISTRY_OUTPUTS = CYBERVISION_DIR / "disease_registry" / "outputs"
RESULTS_DIR = CYBERVISION_DIR / "results" / "open_agentic"

# ── Config ─────────────────────────────────────────────────────────────────────
MODEL = "sonnet"
_ACTIVE_MODEL = MODEL  # set by main() from CLI args
TIMEOUT_S = 300  # 5 minutes per image
ENV_STRIP_PREFIXES = ("CLAUDE", "CURSOR", "MCP_CONNECTION", "VSCODE", "ELECTRON")

def _resolve_backend(model: str) -> tuple[str, str]:
    """Return (backend, concrete_model_id). Backend is 'claude' or 'gemini'.

    For gemini, the model name (e.g. 'gemini-flash', 'gemini-pro') is passed
    through unchanged and resolved to the underlying 2.5-flash/2.5-pro via the
    customAliases in CyberVisionAg/.gemini/settings.json (which also disables
    thinking to match Claude's default).
    """
    if model.startswith("gemini"):
        return "gemini", model
    return "claude", model


# ── KB loading ─────────────────────────────────────────────────────────────────

def load_kb(symptom_source: str, dataset_name: str, all_columns: bool = False) -> str | None:
    """Load KB text from the specified source."""
    if symptom_source == "none":
        return None

    if symptom_source == "default":
        if not SYMPTOMS_FILE.exists():
            print(f"  WARNING: {SYMPTOMS_FILE} not found, proceeding without KB")
            return None
        return _extract_crop_section(SYMPTOMS_FILE, dataset_name)

    # local or internet → read from per-crop folder
    crop = dataset_name.replace("_Diseases", "").replace("_Disease", "")
    xlsx_path = REGISTRY_OUTPUTS / crop / f"{symptom_source}.xlsx"
    if not xlsx_path.exists():
        print(f"  WARNING: {xlsx_path} not found, proceeding without KB")
        return None
    return _load_xlsx_as_markdown(xlsx_path, all_columns=all_columns)


def _extract_crop_section(kb_file: Path, dataset_name: str) -> str:
    """Extract only the target crop section from the multi-crop KB file."""
    text = kb_file.read_text()
    # Dataset names like "Soybean_Diseases" → section header "## Crop: Soybean Diseases"
    header = f"## Crop: {dataset_name.replace('_', ' ')}"
    start = text.find(header)
    if start == -1:
        print(f"  WARNING: section '{header}' not found in KB, using full file")
        return text
    # Find next "## Crop:" section or end of file
    next_section = text.find("\n## Crop:", start + len(header))
    section = text[start:next_section] if next_section != -1 else text[start:]
    return section.strip()


def kb_coverage(symptom_source: str, dataset_name: str, classes: list[str]) -> dict:
    """Check which classes have KB data. Returns {class: has_data}."""
    if symptom_source == "none":
        return {cls: False for cls in classes}

    if symptom_source in ("local", "internet"):
        crop = dataset_name.replace("_Diseases", "").replace("_Disease", "")
        xlsx_path = REGISTRY_OUTPUTS / crop / f"{symptom_source}.xlsx"
        if not xlsx_path.exists():
            return {cls: False for cls in classes}
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb.active
        kb_diseases = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[0]
            desc = row[4] if len(row) > 4 else None
            if name and desc and str(desc).strip():
                kb_diseases.add(str(name))
        wb.close()
        # Fuzzy match: class name might differ slightly from xlsx name
        return {cls: cls in kb_diseases for cls in classes}

    return {cls: True for cls in classes}  # assume coverage for unknown sources


def _load_xlsx_as_markdown(xlsx_path: Path, all_columns: bool = False) -> str:
    """Convert xlsx to markdown.

    Args:
        xlsx_path: Path to the xlsx file.
        all_columns: If True, include Pathogen, Type, Affected Parts in addition
            to Visual Description. If False, only Visual Description (legacy).
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    lines = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        disease = row[0]
        if not disease:
            continue
        description = row[4] if len(row) > 4 else None
        if not all_columns:
            if description:
                lines.append(f"### {disease}\n{description}\n")
            continue
        pathogen = row[1] if len(row) > 1 else None
        disease_type = row[2] if len(row) > 2 else None
        affected_parts = row[3] if len(row) > 3 else None
        if not description and not affected_parts:
            continue
        entry = f"### {disease}\n"
        if pathogen and str(pathogen).strip():
            entry += f"**Pathogen**: {pathogen}\n"
        if disease_type and str(disease_type).strip():
            entry += f"**Type**: {disease_type}\n"
        if affected_parts and str(affected_parts).strip():
            entry += f"**Affected parts**: {affected_parts}\n"
        if description and str(description).strip():
            entry += f"{description}\n"
        lines.append(entry)
    wb.close()
    return "\n".join(lines)


# ── Dataset loading ────────────────────────────────────────────────────────────

def load_dataset(
    dataset: str,
    num_classes: int | None,
    images_per_class: int | None,
    seed: int,
    exclude: set[str] | None = None,
    test_dir_override: Path | None = None,
) -> tuple[list[str], list[tuple[str, str]]]:
    """Discover classes and test images.

    Returns:
        (classes, test_images) where test_images is [(abs_path, ground_truth), ...]
    """
    test_dataset_dir = test_dir_override if test_dir_override else TEST_DIR / dataset
    train_dataset_dir = TRAIN_DIR / dataset

    if not test_dataset_dir.exists():
        sys.exit(f"ERROR: test directory not found: {test_dataset_dir}")
    if not train_dataset_dir.exists():
        # If using prepared dataset, train dir check is optional
        if not test_dir_override:
            sys.exit(f"ERROR: train directory not found: {train_dataset_dir}")

    # Discover all classes (folders in test dir)
    all_classes = sorted([
        d.name for d in test_dataset_dir.iterdir()
        if d.is_dir() and not d.name.startswith(".")
        and (not exclude or d.name not in exclude)
    ])

    # Select subset
    rng = __import__("random").Random(seed)
    if num_classes and num_classes < len(all_classes):
        classes = sorted(rng.sample(all_classes, num_classes))
    else:
        classes = all_classes

    # Gather test images
    test_images = []
    for cls in classes:
        cls_dir = test_dataset_dir / cls
        imgs = sorted(
            str(p) for p in cls_dir.iterdir()
            if p.suffix.lower() in (".jpg", ".jpeg", ".png", ".webp")
        )
        if images_per_class and images_per_class < len(imgs):
            imgs = sorted(rng.sample(imgs, images_per_class))
        for img_path in imgs:
            test_images.append((img_path, cls))

    return classes, test_images


def _ensure_merged_refs(ref_dir: Path, classes: list[str]) -> None:
    """Create merged/ folders by copying images from part subfolders (except rejected/).

    Runs once per eval. Skips classes where merged/ already has images.
    """
    for cls in classes:
        cls_dir = ref_dir / cls
        if not cls_dir.exists():
            continue
        merged_dir = cls_dir / "merged"
        # Skip if merged already has images
        if merged_dir.exists() and any(merged_dir.glob("*.jpg")) or any(merged_dir.glob("*.png")):
            continue
        merged_dir.mkdir(exist_ok=True)
        for part_dir in cls_dir.iterdir():
            if not part_dir.is_dir() or part_dir.name in ("rejected", "test", "merged"):
                continue
            for img in part_dir.iterdir():
                if img.suffix.lower() in (".jpg", ".jpeg", ".png", ".webp"):
                    dest = merged_dir / img.name
                    if not dest.exists():
                        shutil.copy2(img, dest)


def find_reference_images(
    dataset: str, classes: list[str], refs_per_class: int = 1,
    ref_dir: Path | None = None, use_parts: bool = True,
) -> dict[str, list[str]]:
    """Find reference images per class.

    Args:
        refs_per_class: How many reference images per class (evenly spaced).
            1 = middle image only (legacy behavior).
        ref_dir: Prepared dataset dir (class/part/ subfolders). Pools all parts.
        use_parts: If True, read from class/part/ subfolders (part info in path).
            If False, read from class/merged/ (no part info in path).

    Returns:
        {class_name: [list_of_paths]}
    """
    ref_images = {}
    for cls in classes:
        if ref_dir:
            cls_dir = ref_dir / cls
        else:
            cls_dir = TRAIN_DIR / dataset / cls
        if not cls_dir.exists():
            continue
        if ref_dir and not use_parts:
            # Read from merged/ (flat, no part info in path)
            merged_dir = cls_dir / "merged"
            imgs = sorted(
                str(p) for p in merged_dir.iterdir()
                if p.suffix.lower() in (".jpg", ".jpeg", ".png", ".webp")
            ) if merged_dir.exists() else []
        elif ref_dir:
            imgs = sorted(
                str(p) for part_dir in cls_dir.iterdir()
                if part_dir.is_dir() and part_dir.name not in ("rejected", "test", "merged")
                for p in part_dir.iterdir()
                if p.suffix.lower() in (".jpg", ".jpeg", ".png", ".webp")
            )
        else:
            imgs = sorted(
                str(p) for p in cls_dir.iterdir()
                if p.suffix.lower() in (".jpg", ".jpeg", ".png", ".webp")
            )
        if not imgs:
            continue
        if refs_per_class >= len(imgs):
            ref_images[cls] = imgs
        elif refs_per_class == 1:
            ref_images[cls] = [imgs[len(imgs) // 2]]
        else:
            # Evenly spaced selection
            step = len(imgs) / refs_per_class
            ref_images[cls] = [imgs[int(i * step)] for i in range(refs_per_class)]
    return ref_images


def make_collages(
    dataset: str, classes: list[str], n: int = 4,
    ref_dir: Path | None = None,
) -> dict[str, list[str]]:
    """Create a 2×2 collage of training images per class.

    Returns {class_name: [collage_path]} — one collage per class,
    so it uses 1 ref budget but shows n images.

    If ref_dir is provided (prepared dataset), reads from class/*/
    (all part subfolders pooled). Otherwise reads from TRAIN_DIR/dataset/class/.
    """
    from PIL import Image

    collage_dir = Path(tempfile.mkdtemp(prefix="collages_"))
    ref_images = {}
    for cls in classes:
        if ref_dir:
            cls_dir = ref_dir / cls
        else:
            cls_dir = TRAIN_DIR / dataset / cls
        if not cls_dir.exists():
            continue
        # If prepared dataset (has part subfolders), pool all part images
        if ref_dir:
            imgs = sorted(
                str(p) for part_dir in cls_dir.iterdir()
                if part_dir.is_dir() and part_dir.name not in ("rejected", "test", "merged")
                for p in part_dir.iterdir()
                if p.suffix.lower() in (".jpg", ".jpeg", ".png", ".webp")
            )
        else:
            imgs = sorted(
                str(p) for p in cls_dir.iterdir()
                if p.suffix.lower() in (".jpg", ".jpeg", ".png", ".webp")
            )
        if not imgs:
            continue
        # Pick n evenly spaced images
        selected = imgs if len(imgs) <= n else [
            imgs[int(i * len(imgs) / n)] for i in range(n)
        ]
        # Create 2×2 collage
        pil_imgs = [Image.open(p) for p in selected]
        # Resize all to same size (use smallest dimensions)
        w = min(im.width for im in pil_imgs)
        h = min(im.height for im in pil_imgs)
        size = (min(w, 800), min(h, 800))  # cap at 800px per tile
        pil_imgs = [im.resize(size) for im in pil_imgs]
        cols = 2
        rows = (len(pil_imgs) + 1) // 2
        collage = Image.new("RGB", (size[0] * cols, size[1] * rows))
        for i, im in enumerate(pil_imgs):
            x = (i % cols) * size[0]
            y = (i // cols) * size[1]
            collage.paste(im, (x, y))
        collage_path = str(collage_dir / f"{cls}.jpg")
        collage.save(collage_path, quality=85)
        ref_images[cls] = [collage_path]
    return ref_images


# ── Environment ────────────────────────────────────────────────────────────────

def _clean_env() -> dict[str, str]:
    """Strip env vars that cause claude -p to hang or use wrong credentials."""
    env = {
        k: v for k, v in os.environ.items()
        if not any(k.startswith(prefix) for prefix in ENV_STRIP_PREFIXES)
    }
    # Remove API key so claude -p uses the logged-in account, not the API key
    env.pop("ANTHROPIC_API_KEY", None)
    return env


# ── Single image classification ────────────────────────────────────────────────

def run_single_image(
    test_image: str,
    ground_truth: str,
    classes: list[str],
    ref_images: dict[str, str],
    kb_text: str | None,
    k: int | None,
    timeout: int,
    log_dir: Path | None,
    attractor_guide_dir: str | None = None,
    part_index_path: str | None = None,
) -> dict:
    """Classify one test image via a claude -p agent."""

    # Copy test image to neutral name (prevent filename leakage)
    tmp = tempfile.NamedTemporaryFile(
        suffix=".jpg", prefix="classify_", delete=False, dir="/tmp"
    )
    shutil.copy2(test_image, tmp.name)
    tmp.close()

    try:
        result = _run_agent(
            tmp.name, ground_truth, classes, ref_images, kb_text, k, timeout,
            attractor_guide_dir, part_index_path,
        )
    finally:
        os.unlink(tmp.name)

    # Log per-image result
    if log_dir:
        log_dir.mkdir(parents=True, exist_ok=True)
        img_name = Path(test_image).stem
        # Include class name to avoid collisions (e.g., Soybean_Dise_10 in multiple classes)
        log_name = f"{ground_truth}__{img_name}"
        log_file = log_dir / f"{log_name}.json"
        trace = result.pop("trace", [])
        with open(log_file, "w") as f:
            json.dump(result, f, indent=2)
        # Save trace separately (for reasoning analysis)
        if trace:
            trace_dir = log_dir / "traces"
            trace_dir.mkdir(exist_ok=True)
            with open(trace_dir / f"{log_name}.json", "w") as f:
                json.dump(trace, f, indent=2)
        result["trace"] = trace  # restore for in-memory use

    return result


def _run_agent(
    test_image_path: str,
    ground_truth: str,
    classes: list[str],
    ref_images: dict[str, str],
    kb_text: str | None,
    k: int | None,
    timeout: int,
    attractor_guide_dir: str | None = None,
    part_index_path: str | None = None,
) -> dict:
    """Dispatch to the appropriate backend based on _ACTIVE_MODEL."""
    backend, concrete_model = _resolve_backend(_ACTIVE_MODEL)
    if backend == "gemini":
        return _run_gemini_agent(
            test_image_path, ground_truth, classes, ref_images, kb_text, k, timeout,
            concrete_model, attractor_guide_dir, part_index_path,
        )
    return _run_claude_agent(
        test_image_path, ground_truth, classes, ref_images, kb_text, k, timeout,
        concrete_model, attractor_guide_dir, part_index_path,
    )


def _run_claude_agent(
    test_image_path: str,
    ground_truth: str,
    classes: list[str],
    ref_images: dict[str, str],
    kb_text: str | None,
    k: int | None,
    timeout: int,
    model: str,
    attractor_guide_dir: str | None = None,
    part_index_path: str | None = None,
) -> dict:
    """Spawn claude -p subprocess and parse results."""
    system_prompt = build_system_prompt(
        attractor_guide_dir=attractor_guide_dir,
        part_index_path=part_index_path,
        k=k,
    )
    user_message = build_user_message(
        test_image_path, classes, ref_images, kb_text, k,
        attractor_guide_dir=attractor_guide_dir,
    )

    cmd = [
        "claude", "-p",
        "--output-format", "stream-json",
        "--verbose",
        "--no-session-persistence",
        "--model", model,
        "--allowedTools", "Read",
        "--append-system-prompt", system_prompt,
    ]

    env = _clean_env()
    proc = None

    try:
        proc = subprocess.Popen(
            cmd,
            stdin=subprocess.PIPE,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            env=env,
            cwd=str(PROJECT_ROOT),
            start_new_session=True,
        )
        stdout_bytes, stderr_bytes = proc.communicate(
            input=user_message.encode(), timeout=timeout
        )
    except subprocess.TimeoutExpired:
        if proc:
            os.killpg(os.getpgid(proc.pid), signal.SIGKILL)
            proc.wait()
        return _make_result(ground_truth, error="timeout")
    except Exception as e:
        return _make_result(ground_truth, error=str(e))

    if proc.returncode != 0:
        stderr = stderr_bytes.decode(errors="replace")[:500]
        return _make_result(ground_truth, error=f"exit code {proc.returncode}: {stderr}")

    stdout = stdout_bytes.decode(errors="replace")

    # Parse stream-json
    trace, result_event = _parse_stream_json(stdout)

    # Extract prediction from final result text
    result_text = result_event.get("result", "")
    prediction = _parse_prediction(result_text, classes)

    # Count reference image reads (any Read in train dir, excluding test image)
    refs_viewed = _count_ref_reads(trace, test_image_path)

    is_correct = prediction.get("prediction") == ground_truth

    return {
        "test_image": ground_truth,  # class name (not temp path)
        "ground_truth": ground_truth,
        "prediction": prediction.get("prediction", "UNKNOWN"),
        "confidence": prediction.get("confidence", 0.0),
        "reasoning": prediction.get("reasoning", ""),
        "correct": is_correct,
        "num_turns": result_event.get("num_turns", 0),
        "cost_usd": result_event.get("total_cost_usd", 0.0),
        "duration_ms": result_event.get("duration_ms", 0),
        "refs_viewed": refs_viewed,
        "trace": trace,
        "error": None,
    }


def _run_gemini_agent(
    test_image_path: str,
    ground_truth: str,
    classes: list[str],
    ref_images: dict[str, str],
    kb_text: str | None,
    k: int | None,
    timeout: int,
    model: str,
    attractor_guide_dir: str | None = None,
    part_index_path: str | None = None,
) -> dict:
    """Spawn gemini -p subprocess and parse results."""
    if not os.environ.get("GEMINI_API_KEY"):
        return _make_result(ground_truth, error="GEMINI_API_KEY not set")

    system_prompt = build_system_prompt(
        attractor_guide_dir=attractor_guide_dir,
        part_index_path=part_index_path,
        k=k,
    )
    user_message = build_user_message(
        test_image_path, classes, ref_images, kb_text, k,
        attractor_guide_dir=attractor_guide_dir,
    )
    combined_prompt = f"{system_prompt}\n\n{user_message}"

    cmd = [
        "gemini", "-p", combined_prompt,
        "--output-format", "stream-json",
        "--approval-mode", "yolo",
        "--model", model,
        # Allow reads from anywhere: test images live in /tmp (filename-leak guard),
        # collages in $TMPDIR, ref images in CyberVisionAg, etc.
        "--include-directories", "/",
    ]

    env = _clean_env_gemini()
    proc = None

    try:
        proc = subprocess.Popen(
            cmd,
            stdin=subprocess.DEVNULL,
            stdout=subprocess.PIPE,
            stderr=subprocess.DEVNULL,  # drop YOLO warnings + startup noise
            env=env,
            cwd=str(CYBERVISION_DIR),  # avoid root .gitignore blocking image reads
            start_new_session=True,
        )
        stdout_bytes, _ = proc.communicate(timeout=timeout)
    except subprocess.TimeoutExpired:
        if proc:
            os.killpg(os.getpgid(proc.pid), signal.SIGKILL)
            proc.wait()
        return _make_result(ground_truth, error="timeout")
    except Exception as e:
        return _make_result(ground_truth, error=str(e))

    if proc.returncode != 0:
        return _make_result(ground_truth, error=f"gemini exit code {proc.returncode}")

    stdout = stdout_bytes.decode(errors="replace")
    trace, result_event = _parse_gemini_stream(stdout)

    result_text = result_event.get("result", "")
    prediction = _parse_prediction(result_text, classes)

    # Debug: on failure (no prediction), dump raw stdout for post-mortem.
    if prediction.get("prediction") is None or not trace:
        debug_dir = Path("/tmp/gemini_debug")
        debug_dir.mkdir(exist_ok=True)
        import uuid
        with open(debug_dir / f"fail_{uuid.uuid4().hex[:8]}.log", "w") as _f:
            _f.write(f"GROUND_TRUTH: {ground_truth}\nMODEL: {model}\n---STDOUT---\n{stdout}\n")

    refs_viewed = _count_ref_reads(trace, test_image_path)
    is_correct = prediction.get("prediction") == ground_truth

    return {
        "test_image": ground_truth,
        "ground_truth": ground_truth,
        "prediction": prediction.get("prediction", "UNKNOWN"),
        "confidence": prediction.get("confidence", 0.0),
        "reasoning": prediction.get("reasoning", ""),
        "correct": is_correct,
        "num_turns": result_event.get("num_turns", 0),
        "cost_usd": 0.0,  # TODO: compute from token counts + pricing table
        "duration_ms": result_event.get("duration_ms", 0),
        "refs_viewed": refs_viewed,
        "trace": trace,
        "error": None,
    }


def _clean_env_gemini() -> dict[str, str]:
    """Env for gemini subprocess: strip Claude/VSCode vars, keep GEMINI_API_KEY."""
    strip_prefixes = ENV_STRIP_PREFIXES + ("GOOGLE_",)
    env = {
        k: v for k, v in os.environ.items()
        if not any(k.startswith(p) for p in strip_prefixes)
    }
    # Keep only GEMINI_API_KEY; strip other GEMINI_* that may interfere
    for k in list(env.keys()):
        if k.startswith("GEMINI_") and k != "GEMINI_API_KEY":
            env.pop(k, None)
    env["GEMINI_API_KEY"] = os.environ.get("GEMINI_API_KEY", "")
    env.pop("ANTHROPIC_API_KEY", None)
    return env


def _parse_gemini_stream(stdout: str) -> tuple[list[dict], dict]:
    """Parse Gemini stream-json NDJSON into (trace, synthetic_result_event).

    Normalizes to Claude-compatible shape so downstream counters/parsers work:
    - trace entries: {type: "text", content: ...} or {type: "tool_use", tool: "Read", input: {...}}
    - result_event: {result: text, num_turns, duration_ms, total_cost_usd}
    """
    trace: list[dict] = []
    result_stats: dict = {}
    assistant_buffer = ""  # accumulate streamed deltas per assistant run
    assistant_active = False

    def flush_assistant():
        nonlocal assistant_buffer, assistant_active
        if assistant_buffer.strip():
            trace.append({"type": "text", "content": assistant_buffer.strip()})
        assistant_buffer = ""
        assistant_active = False

    for line in stdout.strip().split("\n"):
        if not line.strip() or not line.startswith("{"):
            continue
        try:
            obj = json.loads(line)
        except json.JSONDecodeError:
            continue

        etype = obj.get("type")
        if etype == "message" and obj.get("role") == "assistant":
            assistant_buffer += obj.get("content", "")
            assistant_active = True
        elif etype == "tool_use":
            if assistant_active:
                flush_assistant()
            tool_name = obj.get("tool_name", "")
            # Normalize read_file -> Read for downstream compatibility
            norm_tool = "Read" if tool_name == "read_file" else tool_name
            params = obj.get("parameters", {}) or {}
            trace.append({
                "type": "tool_use",
                "tool": norm_tool,
                "input": params,
            })
        elif etype == "result":
            if assistant_active:
                flush_assistant()
            result_stats = obj.get("stats", {}) or {}

    # Flush any trailing assistant content
    if assistant_active:
        flush_assistant()

    # Final text is the last assistant run in trace
    final_text = ""
    for entry in reversed(trace):
        if entry.get("type") == "text":
            final_text = entry.get("content", "")
            break

    result_event = {
        "result": final_text,
        "num_turns": result_stats.get("tool_calls", 0) + 1,
        "duration_ms": result_stats.get("duration_ms", 0),
        "total_cost_usd": 0.0,  # Gemini does not report cost; compute later
    }
    return trace, result_event


def _make_result(ground_truth: str, error: str) -> dict:
    """Create an error result dict."""
    return {
        "test_image": ground_truth,
        "ground_truth": ground_truth,
        "prediction": "UNKNOWN",
        "confidence": 0.0,
        "reasoning": "",
        "correct": False,
        "num_turns": 0,
        "cost_usd": 0.0,
        "duration_ms": 0,
        "refs_viewed": 0,
        "trace": [],
        "error": error,
    }


# ── Output parsing ─────────────────────────────────────────────────────────────

def _parse_stream_json(stdout: str) -> tuple[list[dict], dict]:
    """Parse stream-json NDJSON output into trace + result event."""
    trace = []
    result_event = {}

    for line in stdout.strip().split("\n"):
        if not line.strip():
            continue
        try:
            obj = json.loads(line)
        except json.JSONDecodeError:
            continue

        if obj.get("type") == "assistant":
            for c in obj.get("message", {}).get("content", []):
                if c.get("type") == "text" and c.get("text", "").strip():
                    trace.append({
                        "type": "text",
                        "content": c["text"].strip(),
                    })
                elif c.get("type") == "tool_use":
                    trace.append({
                        "type": "tool_use",
                        "tool": c.get("name", ""),
                        "input": c.get("input", {}),
                    })

        elif obj.get("type") == "result":
            result_event = obj

    return trace, result_event


def _parse_prediction(text: str, classes: list[str]) -> dict:
    """Extract the prediction JSON from the agent's final text response."""
    # Look for JSON block (last one wins)
    json_blocks = re.findall(
        r'```(?:json)?\s*\n?\s*(\{[^`]+\})\s*\n?\s*```', text, re.DOTALL
    )

    # Also try bare JSON at end of text
    bare = re.findall(r'(\{"prediction"[^}]+\})', text)

    candidates = json_blocks + bare

    for candidate in reversed(candidates):  # last one is most likely the final
        try:
            parsed = json.loads(candidate)
            if "prediction" in parsed:
                return parsed
        except json.JSONDecodeError:
            continue

    # Fallback: try to find class name mentioned near end of text
    for cls in classes:
        if cls in text[-200:]:
            return {"prediction": cls, "confidence": 0.0, "reasoning": "parsed from text"}

    return {"prediction": "UNKNOWN", "confidence": 0.0, "reasoning": "no prediction found"}


def _count_ref_reads(trace: list[dict], test_image: str) -> int:
    """Count Read tool calls for reference images (not the test image)."""
    count = 0
    for entry in trace:
        if entry.get("type") != "tool_use" or entry.get("tool") != "Read":
            continue
        file_path = entry.get("input", {}).get("file_path", "")
        if file_path == test_image:
            continue
        # Any image that isn't the test image is a reference
        if file_path.lower().endswith((".jpg", ".jpeg", ".png", ".webp")):
            count += 1
    return count


# ── Metrics ────────────────────────────────────────────────────────────────────

def compute_metrics(results: list[dict]) -> dict:
    """Compute aggregate metrics from all image results."""
    n = len(results)
    if n == 0:
        return {}

    correct = sum(1 for r in results if r["correct"])
    errors = sum(1 for r in results if r["error"])
    total_turns = sum(r["num_turns"] for r in results)
    total_cost = sum(r["cost_usd"] for r in results)
    total_duration = sum(r["duration_ms"] for r in results)
    total_refs = sum(r["refs_viewed"] for r in results)

    # Per-class accuracy
    from collections import defaultdict
    class_stats = defaultdict(lambda: {"correct": 0, "total": 0})
    for r in results:
        gt = r["ground_truth"]
        class_stats[gt]["total"] += 1
        if r["correct"]:
            class_stats[gt]["correct"] += 1

    per_class = {
        cls: stats["correct"] / stats["total"] * 100
        for cls, stats in sorted(class_stats.items())
    }

    return {
        "accuracy": correct / n * 100,
        "correct": correct,
        "total": n,
        "errors": errors,
        "avg_turns": total_turns / n,
        "avg_cost_usd": total_cost / n,
        "total_cost_usd": total_cost,
        "avg_duration_s": total_duration / n / 1000,
        "avg_refs_viewed": total_refs / n,
        "per_class_accuracy": per_class,
    }


def print_metrics(metrics: dict) -> None:
    """Print formatted metrics."""
    if not metrics:
        print("  No results.")
        return
    print(f"\n  Accuracy         : {metrics['correct']}/{metrics['total']} "
          f"({metrics['accuracy']:.1f}%)")
    print(f"  Avg turns        : {metrics['avg_turns']:.1f}")
    print(f"  Avg refs viewed  : {metrics['avg_refs_viewed']:.1f}")
    print(f"  Avg cost/image   : ${metrics['avg_cost_usd']:.4f}")
    print(f"  Total cost       : ${metrics['total_cost_usd']:.4f}")
    print(f"  Avg duration     : {metrics['avg_duration_s']:.1f}s")
    print(f"  Errors           : {metrics['errors']}")

    if metrics.get("per_class_accuracy"):
        print("\n  Per-class accuracy:")
        for cls, acc in metrics["per_class_accuracy"].items():
            print(f"    {cls:30s} {acc:.0f}%")


# ── Main ───────────────────────────────────────────────────────────────────────

def run_eval(
    classes: list[str],
    test_images: list[tuple[str, str]],
    ref_images: dict[str, str],
    kb_text: str | None,
    k: int | None,
    parallel: int,
    timeout: int,
    log_dir: Path,
    attractor_guide_dir: str | None = None,
    part_index_path: str | None = None,
) -> list[dict]:
    """Run evaluation on all test images with parallel dispatch."""
    results = []
    total = len(test_images)

    def _process(idx_and_item):
        idx, (img_path, gt) = idx_and_item
        label = f"[{idx+1}/{total}]"
        print(f"  {label} {gt} ...", end="", flush=True)
        t0 = time.time()
        result = run_single_image(
            img_path, gt, classes, ref_images, kb_text, k, timeout, log_dir,
            attractor_guide_dir, part_index_path,
        )
        elapsed = time.time() - t0
        status = "OK" if result["correct"] else "WRONG"
        if result["error"]:
            status = f"ERR({result['error'][:20]})"
        pred = result["prediction"]
        print(f" → {pred} ({status}, {elapsed:.0f}s)", flush=True)
        return result

    if parallel <= 1:
        for item in enumerate(test_images):
            results.append(_process(item))
    else:
        with ThreadPoolExecutor(max_workers=parallel) as pool:
            futures = {
                pool.submit(_process, item): item
                for item in enumerate(test_images)
            }
            for future in as_completed(futures):
                results.append(future.result())

    return results


def main():
    parser = argparse.ArgumentParser(
        description="Open Agentic Pipeline — plant disease classification"
    )
    parser.add_argument("--model", default=MODEL,
                        help=f"Claude model to use (default: {MODEL})")
    parser.add_argument("--symptom-source", default="none",
                        choices=["none", "local", "internet"],
                        help="KB source (default: none)")
    parser.add_argument("--dataset", default="Soybean_Diseases",
                        help="Dataset folder name (default: Soybean_Diseases)")
    parser.add_argument("--num-classes", type=int, default=None,
                        help="Limit to N random classes")
    parser.add_argument("--images-per-class", type=int, default=None,
                        help="Test images per class")
    parser.add_argument("--quick-test", type=int, default=None, metavar="N",
                        help="Shortcut: num-classes=N, images-per-class=1")
    parser.add_argument("--no-collage", action="store_true",
                        help="Use single training images instead of collages (default: collage)")
    parser.add_argument("--ref-dir", type=str, default=None,
                        help="Prepared dataset dir for references (e.g., Prepared_Dataset/Soybean). "
                             "Reads from class/part/ subfolders, pools all parts.")
    parser.add_argument("--test-dir", type=str, default=None,
                        help="Override test image directory (e.g., Prepared_Dataset/Soybean_test). "
                             "Expects class/img.jpg structure.")
    parser.add_argument("--part-index", type=str, default=None,
                        help="Path to part_index.md file. Agent reads it to narrow candidates "
                             "by plant part (e.g., Prepared_Dataset/Soybean/part_index.md).")
    parser.add_argument("--refs-per-class", type=int, default=1,
                        help="Reference images per class (default: 1, max: 5)")
    parser.add_argument("--k", type=int, default=None,
                        help="Max reference images to view (default: unlimited)")
    parser.add_argument("--parallel", type=int, default=1,
                        help="Concurrent workers (default: 1)")
    parser.add_argument("--seed", type=int, default=42,
                        help="Random seed (default: 42)")
    parser.add_argument("--timeout", type=int, default=TIMEOUT_S,
                        help=f"Per-image timeout in seconds (default: {TIMEOUT_S})")
    parser.add_argument("--exclude", type=str, default=None,
                        help="Comma-separated class names to exclude")
    parser.add_argument("--all-kb-columns", action="store_true",
                        help="Include Pathogen, Type, Affected Parts in KB (not just Visual Description)")
    parser.add_argument("--kb-file", type=str, default=None,
                        help="Custom KB markdown file (overrides --symptom-source)")
    parser.add_argument("--confusion-guide", type=str, default=None,
                        help="Path to confusion guide markdown file (agent reads it as a separate tool call)")
    args = parser.parse_args()

    # Set active model from CLI
    global _ACTIVE_MODEL
    _ACTIVE_MODEL = args.model

    # Quick-test shortcut
    if args.quick_test:
        args.num_classes = args.quick_test
        args.images_per_class = 1

    exclude = set(args.exclude.split(",")) if args.exclude else None

    # Load dataset
    classes, test_images = load_dataset(
        args.dataset, args.num_classes, args.images_per_class, args.seed, exclude,
        Path(args.test_dir) if args.test_dir else None,
    )
    ref_dir = Path(args.ref_dir) if args.ref_dir else None
    # No-KB uses merged/ refs (no part info in path); KB uses part/ subfolders
    has_kb = args.symptom_source != "none" or args.kb_file
    use_parts = has_kb
    if ref_dir and not use_parts:
        _ensure_merged_refs(ref_dir, classes)
    if args.no_collage:
        ref_images = find_reference_images(args.dataset, classes, args.refs_per_class,
                                           ref_dir=ref_dir, use_parts=use_parts)
    else:
        ref_images = make_collages(args.dataset, classes, ref_dir=ref_dir)
    if ref_dir:
        print(f"  Ref source: {ref_dir} ({'parts' if use_parts else 'merged'})")

    # Load KB
    if args.kb_file:
        kb_text = Path(args.kb_file).read_text()
        kb_label = Path(args.kb_file).stem  # e.g., "kb_v1"
    else:
        kb_text = load_kb(args.symptom_source, args.dataset, all_columns=args.all_kb_columns)
        kb_label = args.symptom_source

    # Setup logging — dir includes crop/kb/model/k so configs don't clobber
    k_label = f"k{args.k}" if args.k is not None else "kunlimited"
    if args.confusion_guide:
        k_label += "_cg"  # distinguish confusion-guide runs
    log_dir = RESULTS_DIR / args.dataset / kb_label / _ACTIVE_MODEL / k_label
    # Clear only this specific config's prior results
    if log_dir.exists():
        for old_file in log_dir.glob("*.json"):
            old_file.unlink()
        # Also clear traces
        trace_dir = log_dir / "traces"
        if trace_dir.exists():
            for old_file in trace_dir.glob("*.json"):
                old_file.unlink()

    # Print config
    _backend, _ = _resolve_backend(_ACTIVE_MODEL)
    _binary = "gemini -p" if _backend == "gemini" else "claude -p"
    print(f"OPEN AGENTIC CLASSIFICATION ({_binary})")
    print(f"  Model         : {_ACTIVE_MODEL}")
    print(f"  Symptom source: {args.kb_file or args.symptom_source}")
    print(f"  Dataset       : {args.dataset}")
    print(f"  Classes       : {len(classes)}")
    print(f"  Test images   : {len(test_images)}")
    print(f"  Ref budget (k): {args.k or 'unlimited'}")
    print(f"  Parallel      : {args.parallel}")
    print(f"  Seed          : {args.seed}")
    print(f"  Timeout       : {args.timeout}s")
    print(f"  KB loaded     : {'yes' if kb_text else 'no'}")
    if kb_text:
        print(f"  KB size       : {len(kb_text)} chars")
    # KB coverage
    if args.symptom_source != "none":
        coverage = kb_coverage(args.symptom_source, args.dataset, classes)
        covered = sum(1 for v in coverage.values() if v)
        print(f"  KB coverage   : {covered}/{len(classes)} classes have symptom data")
        missing = [c for c, v in coverage.items() if not v]
        if missing:
            print(f"  KB missing    : {', '.join(missing)}")
    print(f"  Logs          : {log_dir}/")
    print()

    # Resolve attractor guide directory
    attractor_guide_dir = None
    if args.confusion_guide:
        cg = Path(args.confusion_guide)
        if cg.is_dir():
            attractor_guide_dir = str(cg.resolve())
            n_files = len(list(cg.glob("*.md")))
            print(f"  Attractor guide: {attractor_guide_dir} ({n_files} classes)")
        else:
            print(f"  WARNING: attractor guide dir not found: {cg}")

    # Run
    # Resolve part index
    part_index_path = None
    if args.part_index:
        pi = Path(args.part_index)
        if pi.exists():
            part_index_path = str(pi.resolve())
            print(f"  Part index: {part_index_path}")
        else:
            print(f"  WARNING: part index not found: {pi}")

    results = run_eval(
        classes, test_images, ref_images, kb_text,
        args.k, args.parallel, args.timeout, log_dir,
        attractor_guide_dir, part_index_path,
    )

    # Metrics
    metrics = compute_metrics(results)
    print_metrics(metrics)

    # Save summary
    log_dir.mkdir(parents=True, exist_ok=True)
    summary_file = log_dir / "summary.json"
    with open(summary_file, "w") as f:
        json.dump({
            "config": {
                "symptom_source": args.symptom_source,
                "dataset": args.dataset,
                "num_classes": len(classes),
                "num_images": len(test_images),
                "k": args.k,
                "parallel": args.parallel,
                "seed": args.seed,
                "model": _ACTIVE_MODEL,
                "confusion_guide": args.confusion_guide,
            },
            "metrics": {k: v for k, v in metrics.items()
                        if k != "per_class_accuracy"},
            "per_class_accuracy": metrics.get("per_class_accuracy", {}),
        }, f, indent=2)
    print(f"\n  Summary saved: {summary_file}")


if __name__ == "__main__":
    main()
