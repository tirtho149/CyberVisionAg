"""
Disease Dataset Curator & Report Generator
===========================================
Handles BOTH local datasets (dataloader.py) AND online Kaggle/Zenodo datasets
(data_loader.py) in one unified pipeline.

FEATURES
─────────
• Interactive prompt: asks how many REFERENCE (train) and BENCHMARK (test) images
  per class before doing anything.
• Smart downloading: skips any dataset whose output folder already exists.
• Two saved sets per class:
    samples/<DatasetName>/<ClassName>/reference/   ← train images
    samples/<DatasetName>/<ClassName>/benchmark/   ← test  images
  Files named:  <ClassName>_reference_001.jpg / <ClassName>_benchmark_001.jpg
• Deletes raw download after sampling to save disk space.
• PDF report with two image rows per class (Reference | Benchmark).
• Works with LOCAL folders (set LOCAL_SOURCE_ROOT) and/or ONLINE datasets.

REQUIREMENTS
────────────
  pip install kaggle reportlab Pillow pandas scikit-learn requests tqdm openpyxl
  pip install datasets          # required for LeafNet (HuggingFace)
  pip install kagglehub         # required for Banana Leaf dataset
  pip install tensorflow tensorflow_datasets  # required for PlantVillage

USAGE
─────
  1.  Set LOCAL_SOURCE_ROOT below (or leave "" to skip local datasets).
  2.  Place ~/.kaggle/kaggle.json for Kaggle datasets.
  3.  python disease_report_generator.py
"""

# ═══════════════════════════════════════════════════════════════════════════
#  IMPORTS
# ═══════════════════════════════════════════════════════════════════════════

import os, sys, zipfile, shutil, difflib, random, json, re, time, hashlib
import multiprocessing
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import requests
from io import BytesIO

import pandas as pd
import numpy as np
from PIL import Image
from sklearn.utils import shuffle as sk_shuffle

# ── HuggingFace / datasets (optional — only needed for LeafNet) ──────────────
try:
    # Patch md5 for environments where usedforsecurity kwarg causes errors
    import hashlib as _hashlib
    _orig_md5 = _hashlib.md5
    def _patched_md5(*a, **kw):
        kw.pop("usedforsecurity", None)
        return _orig_md5(*a, **kw)
    _hashlib.md5 = _patched_md5

    from datasets import load_dataset as hf_load_dataset
    HF_AVAILABLE = True
except ImportError:
    HF_AVAILABLE = False

# ── tqdm shim ───────────────────────────────────────────────────────────────
try:
    from tqdm import tqdm as _tqdm
    def tqdm(iterable=None, **kw): return _tqdm(iterable, **kw)
except ImportError:
    class tqdm:
        def __init__(self, iterable=None, desc="", total=None, **kw):
            self._it = iter(iterable) if iterable is not None else None
            if desc: print(f"  {desc} ...")
        def __iter__(self):  return self._it
        def __next__(self):  return next(self._it)
        def __enter__(self): return self
        def __exit__(self, *a): pass
        def update(self, n=1): pass

# ── Kaggle ───────────────────────────────────────────────────────────────────
try:
    from kaggle.api.kaggle_api_extended import KaggleApi
    KAGGLE_AVAILABLE = True
except ImportError:
    KAGGLE_AVAILABLE = False

# ── kagglehub (used by Banana Leaf dataset) ──────────────────────────────────
try:
    import kagglehub
    KAGGLEHUB_AVAILABLE = True
except ImportError:
    KAGGLEHUB_AVAILABLE = False

# ── TensorFlow Datasets (optional — only needed for PlantVillage) ────────────
try:
    import tensorflow_datasets as tfds
    import tensorflow as tf
    TFDS_AVAILABLE = True
except ImportError:
    TFDS_AVAILABLE = False

# ── ReportLab ────────────────────────────────────────────────────────────────
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Image as RLImage, HRFlowable,
)

# ═══════════════════════════════════════════════════════════════════════════
#  ★  USER CONFIGURATION  ★
# ═══════════════════════════════════════════════════════════════════════════

# Path to your LOCAL disease image folder (set "" to skip local datasets)
LOCAL_SOURCE_ROOT = "/home/user/Desktop/Disease Images for Arti/InternalData"  # local disease image root
CDDM_SOURCE_ROOT  = "/home/user/Desktop/Disease Images for Arti/CDDM-images/images"  # CDDM flat comma-named dataset

DATA_ROOT    = "./data"                      # raw download staging area (deleted after use)
CURATED_DIR  = "./Curated_Dataset"           # final output root
REF_DIR      = os.path.join(CURATED_DIR, "Reference_Image")   # reference images
BENCH_DIR    = os.path.join(CURATED_DIR, "Benchmark")         # benchmark images
SAMPLES_DIR  = CURATED_DIR                  # alias kept for internal skip-checks
OUTPUT_PDF   = "./disease_dataset_report.pdf"
OUTPUT_XLSX  = "./crop_disease_registry.xlsx"
RANDOM_STATE = 42
IMAGE_EXT    = ('.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif', '.webp')

# ── Worker threads: use ALL available logical cores ───────────────────────────
NUM_WORKERS = multiprocessing.cpu_count()
print(f"  [SYSTEM] {NUM_WORKERS} logical cores detected — threads will use all of them")

# ═══════════════════════════════════════════════════════════════════════════
#  INTERACTIVE PROMPT
# ═══════════════════════════════════════════════════════════════════════════

def ask_sample_counts():
    """
    Ask user how many images per class for reference and benchmark sets.
    This count is applied uniformly to EVERY class from EVERY source —
    online datasets, InternalData, CDDM, LeafNet — all get the same n.
    """
    print("\n" + "="*60)
    print("  DISEASE DATASET CURATOR — SETUP")
    print("="*60)
    print("  Enter images PER CLASS (same count applied to every class,")
    print("  regardless of which dataset or source it comes from).")
    print("  Two sets are saved per class:")
    print("    REFERENCE  — training / reference samples")
    print("    BENCHMARK  — test / evaluation samples")
    print()

    def ask_int(prompt, default):
        while True:
            raw = input(f"  {prompt} [default: {default}]: ").strip()
            if raw == "":
                return default
            try:
                v = int(raw)
                if v > 0:
                    return v
                print("  Please enter a positive integer.")
            except ValueError:
                print("  Invalid input — enter a whole number.")

    n_ref   = ask_int("Reference images per class?", 10)
    n_bench = ask_int("Benchmark  images per class?", 5)

    print(f"\n  -> {n_ref} reference  +  {n_bench} benchmark  per class  "
          f"(uniform across all sources)")
    print("="*60 + "\n")
    return n_ref, n_bench

# ═══════════════════════════════════════════════════════════════════════════
#  COLOURS
# ═══════════════════════════════════════════════════════════════════════════

DATASET_COLORS = {
    "SBRD":               "#2e9e5e",
    "Mango Leaf":         "#e8794a",
    "Soybean PNAS":       "#4a8ec2",
    "Bean Leaf":          "#9b6fd4",
    "Yellow Rust":        "#d4a82a",
    "FUSARIUM 22":        "#d45858",
    "Banana Leaf":        "#f5c518",
    "Cauliflower":        "#7cb342",
    "Lettuce":            "#00897b",
    "LeafNet":            "#1B5E20",
    "Alfalfa_Diseases":   "#4CAF50",
    "Corn_Diseases":      "#FF9800",
    "Soybean_Diseases":   "#2196F3",
    "Wheat_Diseases":     "#9C27B0",
    "Mango_Leaf_Disease": "#F44336",
}

def _hex(h):
    h = h.lstrip("#")
    r,g,b = (int(h[i:i+2],16)/255 for i in (0,2,4))
    return colors.Color(r,g,b)

PAGE_W, PAGE_H = A4
MARGIN = 1.8 * cm

# ═══════════════════════════════════════════════════════════════════════════
#  UTILITIES
# ═══════════════════════════════════════════════════════════════════════════

def safe_name(s):
    return re.sub(r'[^A-Za-z0-9_\-]', '_', s.replace(' ', '_'))

def get_closest_match(name, options):
    m = difflib.get_close_matches(name, options, n=1, cutoff=0.2)
    return m[0] if m else None

def rename_folders(base, expected):
    if not os.path.isdir(base): return
    for f in list(os.listdir(base)):
        fp = os.path.join(base, f)
        if not os.path.isdir(fp): continue
        m = get_closest_match(f, expected)
        if m and m != f:
            np_ = os.path.join(base, m)
            if not os.path.exists(np_): os.rename(fp, np_)

def rename_folders_dict(base, rmap):
    if not os.path.isdir(base): return
    for f in list(os.listdir(base)):
        fp = os.path.join(base, f)
        if os.path.isdir(fp) and f in rmap:
            np_ = os.path.join(base, rmap[f])
            if not os.path.exists(np_): os.rename(fp, np_)

def find_best_class_dir(root, classes, min_match=1):
    """
    Walk `root` recursively and return the directory whose immediate
    subdirectories best match `classes` (case-insensitive prefix/fuzzy).
    Useful when the zip structure is unknown or varies across dataset versions.

    Parameters
    ----------
    root      : str  – top of the extracted archive
    classes   : list – expected class folder names
    min_match : int  – minimum number of class matches required (default 1)

    Returns
    -------
    best directory path, or None if no suitable folder found.
    """
    classes_lower = [c.lower() for c in classes]
    best_path  = None
    best_score = 0

    for dirpath, dirnames, _ in os.walk(root):
        # Skip hidden / system dirs
        dirnames[:] = [d for d in dirnames if not d.startswith('.')]
        subdirs_lower = [d.lower() for d in dirnames if not d.startswith('.')]
        if not subdirs_lower:
            continue
        # Count how many expected classes appear (exact or as a prefix/suffix)
        score = sum(
            1 for cl in classes_lower
            if any(cl in sl or sl in cl for sl in subdirs_lower)
        )
        # Also give credit for directories that contain image files directly
        img_count = sum(
            1 for d in dirnames
            if any(
                f.lower().endswith(IMAGE_EXT)
                for f in os.listdir(os.path.join(dirpath, d))
                if os.path.isfile(os.path.join(dirpath, d, f))
            )
        )
        combined = score * 10 + min(img_count, len(classes))
        if combined > best_score:
            best_score = combined
            best_path  = dirpath

    if best_path and best_score >= min_match * 10:
        return best_path
    return None


def collect_images_flat(folder):
    """All image files directly inside folder (non-recursive)."""
    if not os.path.isdir(folder): return []
    return [os.path.join(folder, f) for f in os.listdir(folder)
            if f.lower().endswith(IMAGE_EXT)
            and os.path.isfile(os.path.join(folder, f))]

def collect_images_df(base_directory):
    """Walk class subdirs -> DataFrame(path=0, label=1)."""
    rows = []
    for sub in os.listdir(base_directory):
        if sub == ".DS_Store": continue
        sp = os.path.join(base_directory, sub)
        if not os.path.isdir(sp): continue
        for f in os.listdir(sp):
            if f.lower().endswith(IMAGE_EXT):
                rows.append({0: os.path.join(sp, f), 1: sub})
    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=[0,1])

# ─── Download helpers ────────────────────────────────────────────────────────

def download_file(url, filename):
    try:
        r = requests.get(url, stream=True, timeout=60)
        total = int(r.headers.get('content-length', 0))
        with open(filename, 'wb') as f, tqdm(desc=os.path.basename(filename),
                total=total, unit='iB', unit_scale=True) as bar:
            for chunk in r.iter_content(1024):
                f.write(chunk); bar.update(len(chunk))
    except Exception as e:
        print(f"  [ERROR] {url}: {e}")

def get_zenodo_urls(record_id):
    try:
        r = requests.get(f"https://zenodo.org/api/records/{record_id}", timeout=30)
        if r.status_code == 200:
            return [(f['links']['self'], f['key']) for f in r.json().get('files',[])]
    except Exception as e:
        print(f"  [ERROR] Zenodo: {e}")
    return []

def extract_zip(zpath, dest):
    try:
        with zipfile.ZipFile(zpath, 'r') as zf: zf.extractall(dest)
    except Exception as e:
        print(f"  [ERROR] Extract {zpath}: {e}")

def kaggle_download(dataset_name, path):
    if not KAGGLE_AVAILABLE:
        print(f"  [SKIP] kaggle not installed — {dataset_name}"); return False
    try:
        api = KaggleApi(); api.authenticate()
        print(f"  Downloading: {dataset_name}")
        api.dataset_download_files(dataset_name, path=path, unzip=True, quiet=False)
        return True
    except Exception as e:
        print(f"  [ERROR] Kaggle: {e}"); return False

# ─── Sample & save ───────────────────────────────────────────────────────────

def sample_per_class(data_df, classes, n):
    """Sample exactly n images per class. Classes with fewer than n are skipped entirely."""
    if data_df is None or len(data_df) == 0:
        return pd.DataFrame(columns=[0,1])
    parts = []
    for cls in classes:
        cd = data_df[data_df[1] == cls]
        if len(cd) < n:
            if len(cd) > 0:
                print(f"  [SKIP CLASS] '{cls}': only {len(cd)} images, need {n} -- excluded")
            continue
        parts.append(cd.sample(n=n, random_state=RANDOM_STATE))
    if not parts:
        return pd.DataFrame(columns=[0,1])
    return sk_shuffle(pd.concat(parts, ignore_index=True),
                      random_state=RANDOM_STATE).reset_index(drop=True)


def filter_viable_classes(data_df, classes, n_ref, n_bench):
    """
    Return only classes that have at least (n_ref + n_bench) images available.
    Classes below this threshold are dropped entirely -- no partial saves.
    """
    viable, dropped = [], []
    for cls in classes:
        count = len(data_df[data_df[1] == cls]) if data_df is not None and len(data_df) else 0
        if count >= n_ref + n_bench:
            viable.append(cls)
        else:
            dropped.append((cls, count))
    if dropped:
        for cls, count in dropped:
            print(f"  [DROP CLASS] '{cls}': {count} imgs < {n_ref + n_bench} required "
                  f"(need {n_ref} ref + {n_bench} bench) -- excluded entirely")
    return viable


# ─── Universal crop/disease resolver (used by save_split, PDF, XLSX) ────────
DATASET_CROP = {
    "SBRD":            "Rice",
    "Mango Leaf":      "Mango",
    "Soybean PNAS":    "Soybean",
    "Bean Leaf":       "Bean",
    "Yellow Rust":     "Wheat",
    "FUSARIUM 22":     "Chickpea",
    "Banana Leaf":     "Banana",
    "Cauliflower":     "Cauliflower",
    "Lettuce":         "Lettuce",
    "Cucumber":            "Cucumber",
    "Eggplant Disease":    "Eggplant",
    "Cotton Disease":      "Cotton",
    "Pumpkin Leaf":        "Pumpkin",
    "Rose Leaf":           "Rose",
    "Coconut Disease":     "Coconut",
    "Vanilla Disease":     "Vanilla",
    "SugarLeaf IDN":       "Sugarcane",
    "Cucumber Zenodo":     "Cucumber",
    "Durian Leaf":         "Durian",
}

def _parse_crop_disease_from_label(ds_name, cls_label):
    """
    Universal (crop, disease) resolver used by save_split, PDF, and XLSX.

    Rules:
      LeafNet / CDDM  — class label is safe_name("Crop_Disease"):
                         first token = crop, remainder = disease.
      Fixed-crop sets — crop from DATASET_CROP, label = disease.
      Local / unknown — dataset category name = crop, label = disease.
    """
    if ds_name in ("LeafNet", "CDDM", "PlantVillage", "New Plant Diseases", "Plant Diseases Dataset"):
        parts = cls_label.split("_", 1)
        if len(parts) == 2:
            return parts[0].strip().title(), parts[1].replace("_", " ").strip().title()
        return cls_label.replace("_", " ").title(), "Unknown"

    # PlantDoc: "Apple Scab Leaf" -> crop="Apple", disease="Scab Leaf"
    if ds_name == "PlantDoc":
        parts = cls_label.replace("_", " ").strip().split(" ", 1)
        if len(parts) == 2:
            return parts[0].title(), parts[1].title()
        return cls_label.replace("_", " ").title(), "Unknown"

    crop = DATASET_CROP.get(ds_name)
    if crop:
        return crop, cls_label.replace("_", " ").strip().title()

    # Local/internal datasets: ds_name is e.g. "Soybean_Diseases"
    # Strip trailing _Diseases/_Disease so it merges with online datasets of the same crop
    raw_crop = ds_name.replace("_", " ").title()
    for suffix in (" Diseases", " Disease"):
        if raw_crop.endswith(suffix):
            raw_crop = raw_crop[:-len(suffix)].strip()
            break
    return raw_crop, cls_label.replace("_", " ").strip().title()

def save_split(sampled_df, dataset_name, split_name):
    """
    Copy images to the universal crop/disease hierarchy:
      Curated_Dataset/Reference_Image/<Crop>/<Disease>/
      Curated_Dataset/Benchmark/<Crop>/<Disease>/

    The dataset name is NOT part of the path — all sources merge into the
    same crop/disease folders.  Source tracking is handled in the XLSX only.

    Filename pattern:  <Disease>_<split>_<source>_001.jpg
    Returns updated DataFrame with saved paths in column 0.
    """
    if sampled_df is None or len(sampled_df) == 0:
        return sampled_df

    base_dir = REF_DIR if split_name == "reference" else BENCH_DIR
    src_tag  = safe_name(dataset_name)[:12]   # short source tag in filename only

    counters = {}   # (crop_dir, disease_dir) -> count
    tasks    = []   # (idx, src, dest)

    for idx, row in sampled_df.iterrows():
        src = row[0]
        cls = str(row[1])
        crop, disease = _parse_crop_disease_from_label(dataset_name, cls)
        crop_dir    = safe_name(crop)
        disease_dir = safe_name(disease)
        dest_dir    = os.path.join(base_dir, crop_dir, disease_dir)
        os.makedirs(dest_dir, exist_ok=True)
        key = (crop_dir, disease_dir)
        counters[key] = counters.get(key, 0) + 1
        ext   = os.path.splitext(src)[1].lower() or ".jpg"
        fname = f"{disease_dir}_{split_name}_{src_tag}_{counters[key]:03d}{ext}"
        dest  = os.path.join(dest_dir, fname)
        tasks.append((idx, src, dest))

    results = {}
    def _copy(args):
        idx, src, dest = args
        try:
            shutil.copy2(src, dest)
            return idx, dest
        except Exception as e:
            print(f"  [COPY ERROR] {src}: {e}")
            return idx, None

    with ThreadPoolExecutor(max_workers=NUM_WORKERS) as ex:
        for idx, dest in ex.map(_copy, tasks):
            results[idx] = dest

    result    = sampled_df.copy()
    result[0] = [results.get(i) for i in sampled_df.index]
    return result

def split_save_cleanup(data_df, classes, dataset_name, download_path, n_ref, n_bench):
    """
    1. Filter to only classes with >= (n_ref + n_bench) images.
    2. Sample n_ref reference images per class (no overlap with benchmark).
    3. Sample n_bench benchmark images per class from remaining pool.
    4. Save both sets. Delete download_path to free space.
    Returns (ref_df, bench_df).
    """
    # Hard filter: drop any class that cannot satisfy both quotas
    classes = filter_viable_classes(data_df, classes, n_ref, n_bench)
    if not classes:
        print(f"  [ERROR] No viable classes after filtering for {dataset_name}")
        return pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1])

    ref_df = sample_per_class(data_df, classes, n_ref)

    # Exclude reference paths from benchmark pool
    used      = set(ref_df[0].tolist())
    remaining = data_df[~data_df[0].isin(used)]
    bench_df  = sample_per_class(remaining, classes, n_bench)

    ref_df   = save_split(ref_df,   dataset_name, "reference")
    bench_df = save_split(bench_df, dataset_name, "benchmark")

    n_r = sum(1 for p in ref_df[0]   if p)
    n_b = sum(1 for p in bench_df[0] if p)
    print(f"  [SAVED] reference={n_r}  benchmark={n_b} "
          f"-> Curated_Dataset/Reference_Image|Benchmark/{safe_name(dataset_name)}/")

    if download_path and os.path.exists(download_path):
        shutil.rmtree(download_path, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {download_path}")

    return ref_df, bench_df

# ─── Smart skip: already sampled? ────────────────────────────────────────────

def already_sampled(dataset_name, n_ref=None, n_bench=None):
    """
    Check if this dataset has already been sampled into the Curated_Dataset.
    Since all datasets now share a flat Crop/Disease hierarchy, we detect
    presence by looking for any image file whose filename contains the
    dataset source tag.

    If n_ref / n_bench provided, also checks every disease folder that
    belongs to this dataset has enough images.
    """
    src_tag = safe_name(dataset_name)[:12]

    # Collect all disease-level folders that have files tagged with src_tag
    found_dirs = {}   # disease_dir_path -> count of matching images
    for base_dir in [REF_DIR, BENCH_DIR]:
        if not os.path.isdir(base_dir):
            continue
        for crop_dir in os.listdir(base_dir):
            crop_path = os.path.join(base_dir, crop_dir)
            if not os.path.isdir(crop_path):
                continue
            for disease_dir in os.listdir(crop_path):
                disease_path = os.path.join(crop_path, disease_dir)
                if not os.path.isdir(disease_path):
                    continue
                count = sum(
                    1 for f in os.listdir(disease_path)
                    if f.lower().endswith(IMAGE_EXT) and src_tag in f
                )
                if count > 0:
                    found_dirs[disease_path] = count

    if not found_dirs:
        return False   # nothing saved yet for this dataset

    if n_ref is None and n_bench is None:
        return True    # just existence check — passed

    # Count-aware: every disease folder for this source must meet threshold
    required = n_ref or 0
    for disease_path, count in found_dirs.items():
        if count < required:
            print(f"  [RESAMPLE] {dataset_name} -> {os.path.basename(disease_path)}: "
                  f"{count} images < {required} required — will re-sample")
            return False
    return True

def load_from_samples(dataset_name, classes):
    """
    Rebuild ref_df / bench_df from already-saved Curated_Dataset folders.
    Images now live at  <REF_DIR>/<Crop>/<Disease>/
    and are tagged with the dataset source in their filename.
    """
    src_tag   = safe_name(dataset_name)[:12]
    ref_rows  = []
    bench_rows = []

    for base_dir, rows in [(REF_DIR, ref_rows), (BENCH_DIR, bench_rows)]:
        if not os.path.isdir(base_dir):
            continue
        for crop_dir in sorted(os.listdir(base_dir)):
            crop_path = os.path.join(base_dir, crop_dir)
            if not os.path.isdir(crop_path):
                continue
            for disease_dir in sorted(os.listdir(crop_path)):
                disease_path = os.path.join(crop_path, disease_dir)
                if not os.path.isdir(disease_path):
                    continue
                for f in sorted(os.listdir(disease_path)):
                    if f.lower().endswith(IMAGE_EXT) and src_tag in f:
                        # Reconstruct class label as "Crop_Disease" for consistency
                        cls = f"{crop_dir}__{disease_dir}"
                        rows.append({0: os.path.join(disease_path, f), 1: cls})

    ref_df   = pd.DataFrame(ref_rows)   if ref_rows   else pd.DataFrame(columns=[0,1])
    bench_df = pd.DataFrame(bench_rows) if bench_rows else pd.DataFrame(columns=[0,1])
    return ref_df, bench_df

# ═══════════════════════════════════════════════════════════════════════════
#  ONLINE DATASET LOADERS
# ═══════════════════════════════════════════════════════════════════════════

def load_SBRD(n_ref, n_bench):
    name    = "SBRD"
    classes = ['Healthy','Mild Bacterial Blight','Mild Blast','Mild Brownspot',
               'Mild Tungro','Severe Bacterial Blight','Severe Blast',
               'Severe Brownspot','Severe Tungro']
    desc    = ("Rice leaf disease dataset with severity levels covering Bacterial "
               "Blight, Blast, Brownspot and Tungro at Mild and Severe stages.")
    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc
    dl   = os.path.join(DATA_ROOT, "SBRD")
    base = os.path.join(dl, "Leaf Disease Dataset", "train")
    if not os.path.exists(dl):
        kaggle_download("isaacritharson/severity-based-rice-leaf-diseases-dataset", dl)
    rename_folders(base, classes)
    data = collect_images_df(base)
    ref, bench = split_save_cleanup(data, classes, name, dl, n_ref, n_bench)
    return name, ref, bench, classes, desc


def load_MangoLeaf(n_ref, n_bench):
    name    = "Mango Leaf"
    classes = ['Anthracnose','Bacterial Canker','Cutting Weevil','Die Back',
               'Gall Midge','Healthy','Powdery Mildew','Sooty Mould']
    desc    = ("Mango leaf disease dataset covering 7 diseases plus Healthy. "
               "Includes fungal, bacterial and pest-related leaf conditions.")
    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc
    dl = os.path.join(DATA_ROOT, "mango-leaf-disease-dataset")
    if not os.path.exists(dl):
        kaggle_download("aryashah2k/mango-leaf-disease-dataset", dl)
    rename_folders(dl, classes)
    data = collect_images_df(dl)
    ref, bench = split_save_cleanup(data, classes, name, dl, n_ref, n_bench)
    return name, ref, bench, classes, desc


def load_SoybeanPNAS(n_ref, n_bench):
    name    = "Soybean PNAS"
    classes = ['Bacterial Blight','Bacterial Pustule','Frogeye Leaf Spot',
               'Healthy','Herbicide Injury','Iron Deficiency Chlorosis',
               'Potassium Deficiency','Septoria Brown Spot','Sudden Death Syndrome']
    rmap    = {str(i): c for i, c in enumerate(classes)}
    desc    = ("Soybean stress identification from PNAS. Covers 8 stress/disease "
               "conditions plus Healthy including nutrient deficiencies and "
               "fungal, bacterial and environmental stressors.")
    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc
    dl   = os.path.join(DATA_ROOT, "Soybean-PNAS")
    base = os.path.join(dl, "Training Samples")
    if not os.path.exists(dl):
        os.makedirs(dl, exist_ok=True)
        orig = os.getcwd(); os.chdir(dl)
        for url, fname in get_zenodo_urls("12747481"): download_file(url, fname)
        os.chdir(orig)
        for f in os.listdir(dl):
            if f.lower() == "soybean_stress_identification.zip":
                extract_zip(os.path.join(dl, f), dl)
    rename_folders_dict(base, rmap)
    data = collect_images_df(base)
    ref, bench = split_save_cleanup(data, classes, name, dl, n_ref, n_bench)
    return name, ref, bench, classes, desc


def load_BeanLeaf(n_ref, n_bench):
    name    = "Bean Leaf"
    classes = ['Angular Leaf Spot','Bean Rust','Healthy']
    desc    = ("Bean leaf lesion classification: Angular Leaf Spot (bacterial), "
               "Bean Rust (fungal) and Healthy. Compact well-balanced dataset.")
    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc
    dl   = os.path.join(DATA_ROOT, "bean-leaf-lesions-classification")
    base = os.path.join(dl, "train")
    if not os.path.exists(dl):
        kaggle_download("marquis03/bean-leaf-lesions-classification", dl)
    rename_folders(base, classes)
    data = collect_images_df(base)
    ref, bench = split_save_cleanup(data, classes, name, dl, n_ref, n_bench)
    return name, ref, bench, classes, desc


def load_YellowRust(n_ref, n_bench):
    name = "Yellow Rust"
    # Numeric folder names found in the actual zip (RAW/RAW/0, 1, 2 ...)
    # mapped to meaningful class labels.
    rmap = {
        '0': 'No Disease',
        '1': 'Resistant (R)',
        '2': 'Moderately Resistant (MR)',
        '3': 'MRMS',
        '4': 'Moderately Susceptible (MS)',
        '5': 'Susceptible (S)',
        # legacy letter-named folders (some versions of the dataset)
        'MR':   'Moderately Resistant (MR)',
        'MS':   'Moderately Susceptible (MS)',
        'MRMS': 'MRMS',
        'R':    'Resistant (R)',
        'S':    'Susceptible (S)',
    }
    classes = list(dict.fromkeys(rmap.values()))   # unique, insertion-ordered
    desc    = ("Yellow Rust 19 wheat disease dataset. Labelled by resistance level: "
               "Resistant -> Moderately Resistant -> MRMS -> Moderately Susceptible "
               "-> Susceptible, plus No Disease.")

    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc

    dl = os.path.join(DATA_ROOT, "yellowrust19")
    if not os.path.exists(dl):
        kaggle_download("tolgahayit/yellowrust19-yellow-rust-disease-in-wheat", dl)

    # The zip extracts to varying paths across versions:
    #   YELLOW-RUST-19/YELLOW-RUST-19/  (old)
    #   RAW/RAW/                        (new — confirmed in screenshot)
    # Use find_best_class_dir to locate whichever folder holds the class subfolders.
    base = find_best_class_dir(dl, list(rmap.keys()))
    if base is None:
        # Fallback: find the deepest directory that has numeric or letter subfolders
        for dirpath, dirnames, _ in os.walk(dl):
            dirnames[:] = [d for d in dirnames if not d.startswith('.')]
            subs = [d for d in dirnames
                    if d.isdigit() or d in rmap]
            if subs:
                base = dirpath
                break
    if base is None:
        print(f"  [ERROR] Could not locate YellowRust class folders in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    print(f"  [YellowRust] Using folder: {base}")
    print(f"  [YellowRust] Subfolders: {sorted(os.listdir(base))}")

    rename_folders_dict(base, rmap)

    # After renaming, discover which classes actually exist on disk
    actual_classes = [
        c for c in classes
        if os.path.isdir(os.path.join(base, safe_name(c))) or
           os.path.isdir(os.path.join(base, c))
    ]
    if not actual_classes:
        actual_classes = classes   # fall back to full list

    data = collect_images_df(base)
    if len(data) == 0:
        print(f"  [ERROR] No images found in {base}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    present = data[1].unique().tolist()
    actual_classes = [c for c in classes if c in present]

    ref, bench = split_save_cleanup(data, actual_classes, name, dl, n_ref, n_bench)
    return name, ref, bench, actual_classes, desc


def load_BananaLeaf(n_ref, n_bench):
    """
    Banana Leaf Disease Dataset v1.1 — gimrillozarita/banana-leaf-disease-dataset-v1-1
    4 classes (actual Kaggle structure):
      Cordana, Healthy, Panama Disease, Yellow and Black Sigatoka
    Class folders sit directly in the dataset root — there is no augmented subfolder.
    """
    name    = "Banana Leaf"
    classes = ['Cordana', 'Healthy', 'Panama Disease', 'Yellow and Black Sigatoka']
    desc    = ("Banana leaf disease dataset (v1.1) covering three banana leaf diseases: "
               "Cordana, Panama Disease, Yellow and Black Sigatoka, plus Healthy. "
               "Class images are organised directly under the dataset root folder.")

    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc

    base = None

    # ── Download via kagglehub (preferred) ──────────────────────────────────
    if KAGGLEHUB_AVAILABLE:
        try:
            print(f"  [kagglehub] Downloading banana-leaf-disease-dataset-v1-1 ...")
            dl_path = kagglehub.dataset_download(
                "gimrillozarita/banana-leaf-disease-dataset-v1-1")
            print(f"  Path to dataset files: {dl_path}")

            # Class folders (Cordana, Healthy, Panama Disease, …) sit directly
            # under the dataset root — there is NO augmented subfolder.
            # Use find_best_class_dir to handle any nesting the cache may add.
            best = find_best_class_dir(dl_path, classes)
            base = best if best else dl_path
        except Exception as e:
            print(f"  [kagglehub ERROR] {e}")

    # ── Fallback: kaggle API ──────────────────────────────────────────────────
    if base is None:
        print(f"  [FALLBACK] Trying kaggle API ...")
        dl = os.path.join(DATA_ROOT, "banana-leaf-disease")
        if not os.path.exists(dl):
            kaggle_download("gimrillozarita/banana-leaf-disease-dataset-v1-1", dl)

        # Use best-match search — no augmented subfolder in this dataset
        base = find_best_class_dir(dl, classes)
        if base is None:
            base = dl   # absolute last resort

    if base is None or not os.path.isdir(base):
        print(f"  [ERROR] Could not locate Banana Leaf dataset folder.")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    print(f"  [BananaLeaf] Using folder: {base}")
    # Show what's inside to aid debugging
    try:
        contents = [d for d in os.listdir(base)
                    if os.path.isdir(os.path.join(base, d)) and not d.startswith('.')]
        print(f"  [BananaLeaf] Subfolders found: {contents}")
    except Exception:
        pass

    # Normalise folder names to match classes list (case-insensitive)
    rename_folders(base, classes)

    data = collect_images_df(base)
    if len(data) == 0:
        print(f"  [ERROR] No images found in {base}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    # kagglehub downloads to a cache dir — we don't own it, so don't delete it.
    # Pass download_path=None to skip cleanup.
    ref_df  = sample_per_class(data, classes, n_ref)
    used    = set(ref_df[0].tolist())
    rem     = data[~data[0].isin(used)]
    bench_df = sample_per_class(rem, classes, n_bench)
    ref_df   = save_split(ref_df,   name, "reference")
    bench_df = save_split(bench_df, name, "benchmark")

    n_r = sum(1 for p in ref_df[0]   if p)
    n_b = sum(1 for p in bench_df[0] if p)
    print(f"  [SAVED] reference={n_r}  benchmark={n_b} "
          f"-> samples/{safe_name(name)}/")

    return name, ref_df, bench_df, classes, desc


def load_Lettuce(n_ref, n_bench):
    """
    Lettuce Diseases dataset — ashishjstar/lettuce-diseases
    Downloaded via kagglehub (falls back to kaggle API).
    Root: Lettuce_disease_datasets/
    8 classes (auto-discovered from subfolders):
      Bacterial, Downy_mildew_on_lettuce, Healthy,
      Powdery_mildew_on_lettuce, Septoria_blight_on_lettuce,
      Shepherd_purse_weed, Viral, Wilt_and_leaf_blight_or_rot
    Images per class vary — uses whatever is available.
    """
    name = "Lettuce"
    desc = ("Lettuce leaf disease dataset with 8 classes covering bacterial, "
            "fungal and viral conditions: Bacterial, Downy Mildew, Powdery Mildew, "
            "Septoria Blight, Viral, Wilt/Leaf Blight, Shepherd's Purse Weed, "
            "and Healthy. Image counts vary across classes.")

    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, [])
        src_tag = safe_name(name)[:12]
        saved_classes = []
        for crop_d in (os.listdir(REF_DIR) if os.path.isdir(REF_DIR) else []):
            cp = os.path.join(REF_DIR, crop_d)
            if not os.path.isdir(cp): continue
            for dis_d in os.listdir(cp):
                dp = os.path.join(cp, dis_d)
                if not os.path.isdir(dp): continue
                if any(src_tag in f for f in os.listdir(dp)
                       if f.lower().endswith(IMAGE_EXT)):
                    saved_classes.append(f"{crop_d}__{dis_d}")
        return name, ref, bench, sorted(saved_classes), desc

    base = None

    # ── Download via kagglehub (preferred) ───────────────────────────────────
    if KAGGLEHUB_AVAILABLE:
        try:
            print(f"  [kagglehub] Downloading lettuce-diseases ...")
            dl_path = kagglehub.dataset_download("ashishjstar/lettuce-diseases")
            print(f"  Path to dataset files: {dl_path}")
            # Root folder is Lettuce_disease_datasets/ or the dl_path itself
            for candidate in [
                os.path.join(dl_path, "Lettuce_disease_datasets"),
                dl_path,
            ]:
                if os.path.isdir(candidate):
                    subs = [d for d in os.listdir(candidate)
                            if os.path.isdir(os.path.join(candidate, d))
                            and not d.startswith('.')]
                    if subs:
                        base = candidate
                        break
        except Exception as e:
            print(f"  [kagglehub ERROR] {e}")

    # ── Fallback: kaggle API ──────────────────────────────────────────────────
    if base is None:
        print(f"  [FALLBACK] Trying kaggle API ...")
        dl = os.path.join(DATA_ROOT, "lettuce-diseases")
        if not os.path.exists(dl):
            kaggle_download("ashishjstar/lettuce-diseases", dl)
        for candidate in [
            os.path.join(dl, "Lettuce_disease_datasets"),
            dl,
        ]:
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))
                        and not d.startswith('.')]
                if subs:
                    base = candidate
                    break

    if base is None or not os.path.isdir(base):
        print(f"  [ERROR] Could not locate Lettuce dataset folder.")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    # Auto-discover classes from subfolders — handles any truncation/naming
    classes = sorted([
        d for d in os.listdir(base)
        if os.path.isdir(os.path.join(base, d)) and not d.startswith('.')
    ])
    print(f"  [Lettuce] Found {len(classes)} classes: {classes}")

    data = collect_images_df(base)
    if len(data) == 0:
        print(f"  [ERROR] No images found in {base}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    # kagglehub cache — don't delete; pass None to skip cleanup
    ref_df   = sample_per_class(data, classes, n_ref)
    used     = set(ref_df[0].tolist())
    rem      = data[~data[0].isin(used)]
    bench_df = sample_per_class(rem, classes, n_bench)
    ref_df   = save_split(ref_df,   name, "reference")
    bench_df = save_split(bench_df, name, "benchmark")

    n_r = sum(1 for p in ref_df[0]   if p)
    n_b = sum(1 for p in bench_df[0] if p)
    print(f"  [SAVED] reference={n_r}  benchmark={n_b} "
          f"-> samples/{safe_name(name)}/")

    return name, ref_df, bench_df, classes, desc


def load_Cucumber(n_ref, n_bench):
    """
    Cucumber Plant Diseases Dataset
    kagglehub: kareem3egm/cucumber-plant-diseases-dataset

    Structure:
      Cucumber plant diseases/
        training/  Ill_cucumber/  good_Cucumber/   <- REFERENCE pool
        testing/   Ill_cucumber/  good_Cucumber/   <- BENCHMARK pool
        single_prediction/                         <- ignored
    """
    name    = "Cucumber"
    classes = ["Ill_cucumber", "good_Cucumber"]
    desc    = ("Cucumber plant disease dataset (kaggle: kareem3egm). "
               "Two classes: Ill_cucumber (diseased) and good_Cucumber (healthy). "
               "Training split used for reference images, testing split for benchmark.")

    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc

    dl = os.path.join(DATA_ROOT, "cucumber-plant-diseases")
    if not os.path.exists(dl):
        if KAGGLEHUB_AVAILABLE:
            try:
                print(f"  [kagglehub] Downloading Cucumber dataset ...")
                dl_path = kagglehub.dataset_download(
                    "kareem3egm/cucumber-plant-diseases-dataset")
                print(f"  Path to dataset files: {dl_path}")
                os.makedirs(dl, exist_ok=True)
                shutil.copytree(dl_path, dl, dirs_exist_ok=True)
            except Exception as e:
                print(f"  [kagglehub ERROR] {e} — trying kaggle API fallback")
                kaggle_download("kareem3egm/cucumber-plant-diseases-dataset", dl)
        else:
            kaggle_download("kareem3egm/cucumber-plant-diseases-dataset", dl)

    def _find_split(split_name):
        for candidate in [
            os.path.join(dl, "Cucumber plant diseases", split_name),
            os.path.join(dl, split_name),
        ]:
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))
                        and not d.startswith('.')]
                if subs:
                    return candidate
        for dirpath, dirnames, _ in os.walk(dl):
            dirnames[:] = [d for d in dirnames if not d.startswith('.')]
            if os.path.basename(dirpath).lower() == split_name.lower():
                subs = [d for d in os.listdir(dirpath)
                        if os.path.isdir(os.path.join(dirpath, d))]
                if subs:
                    return dirpath
        return None

    train_dir = _find_split("training")
    test_dir  = _find_split("testing")

    if train_dir is None and test_dir is None:
        print(f"  [ERROR] Could not locate Cucumber training/testing folders in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    print(f"  [Cucumber] training (reference): {train_dir}")
    print(f"  [Cucumber] testing  (benchmark): {test_dir}")

    train_data = collect_images_df(train_dir) if train_dir else pd.DataFrame(columns=[0,1])
    test_data  = collect_images_df(test_dir)  if test_dir  else pd.DataFrame(columns=[0,1])

    actual_classes = sorted(set(
        (train_data[1].unique().tolist() if len(train_data) else []) +
        (test_data[1].unique().tolist()  if len(test_data)  else [])
    )) or classes

    # Drop classes that cannot meet quota in EITHER pool
    _viable = []
    for _c in actual_classes:
        _rc = len(train_data[train_data[1] == _c]) if len(train_data) else 0
        _bc = len(test_data[test_data[1] == _c]) if len(test_data) else 0
        if _rc >= n_ref and _bc >= n_bench:
            _viable.append(_c)
        else:
            print(f'  [DROP CLASS] {_c!r}: ref={_rc}/{n_ref}, bench={_bc}/{n_bench} -- excluded')
    actual_classes = _viable
    if not actual_classes:
        print(f'  [DROP ALL] No classes meet quota for both pools in {name}')
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    ref_df   = sample_per_class(train_data, actual_classes, n_ref)
    bench_df = sample_per_class(test_data,  actual_classes, n_bench)
    ref_df   = save_split(ref_df,   name, "reference")
    bench_df = save_split(bench_df, name, "benchmark")

    n_r = sum(1 for p in ref_df[0] if p)
    n_b = sum(1 for p in bench_df[0] if p)
    print(f"  [SAVED] {name}: reference={n_r}  benchmark={n_b}")
    if os.path.exists(dl):
        shutil.rmtree(dl, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {dl}")
    return name, ref_df, bench_df, actual_classes, desc


def load_DurianLeaf(n_ref, n_bench):
    """
    Durian Leaf Disease Dataset
    kagglehub: cthng123/durian-leaf-disease-dataset

    Structure:
      DLD_FinalDataset_224_sp.../
        train/  ALGAL_LEAF_SPOT/ ALLOCARIDARA_ATT/ HEALTHY_LEAF/
                LEAF_BLIGHT/ PHOMOPSIS_LEAF_SPOT/  <- REFERENCE pool
        test/   (same classes)                     <- BENCHMARK pool
        val/    (same classes)                     <- ignored

    Crop: Durian. Class names are UPPER_SNAKE_CASE.
    """
    name    = "Durian Leaf"
    desc    = ("Durian Leaf Disease Dataset (kaggle: cthng123). "
               "5 classes: Algal Leaf Spot, Allocaridara Attack, Healthy Leaf, "
               "Leaf Blight, Phomopsis Leaf Spot. Train/test/val splits.")

    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, [])
        return name, ref, bench, [], desc

    dl = os.path.join(DATA_ROOT, "durian-leaf-disease")
    if not os.path.exists(dl):
        if KAGGLEHUB_AVAILABLE:
            try:
                print(f"  [kagglehub] Downloading durian-leaf-disease-dataset ...")
                dl_path = kagglehub.dataset_download(
                    "cthng123/durian-leaf-disease-dataset")
                print(f"  Path to dataset files: {dl_path}")
                os.makedirs(dl, exist_ok=True)
                shutil.copytree(dl_path, dl, dirs_exist_ok=True)
            except Exception as e:
                print(f"  [kagglehub ERROR] {e} — trying kaggle API fallback")
                kaggle_download("cthng123/durian-leaf-disease-dataset", dl)
        else:
            kaggle_download("cthng123/durian-leaf-disease-dataset", dl)

    def _find_split(split_name):
        """Find train/test/val under the DLD_FinalDataset_... root."""
        # Try direct
        for candidate in [os.path.join(dl, split_name)]:
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))
                        and not d.startswith('.')]
                if subs:
                    return candidate
        # One level deep (DLD_FinalDataset_224_sp.../train/)
        for top in os.listdir(dl):
            top_path = os.path.join(dl, top)
            if not os.path.isdir(top_path) or top.startswith('.'):
                continue
            candidate = os.path.join(top_path, split_name)
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))
                        and not d.startswith('.')]
                if subs:
                    return candidate
        # Walk fallback
        for dirpath, dirnames, _ in os.walk(dl):
            dirnames[:] = [d for d in dirnames if not d.startswith('.')]
            if os.path.basename(dirpath).lower() == split_name.lower():
                subs = [d for d in os.listdir(dirpath)
                        if os.path.isdir(os.path.join(dirpath, d))]
                if subs:
                    return dirpath
        return None

    train_dir = _find_split("train")
    test_dir  = _find_split("test")

    if train_dir is None and test_dir is None:
        print(f"  [ERROR] Could not locate Durian train/test folders in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    print(f"  [DurianLeaf] train (reference): {train_dir}")
    print(f"  [DurianLeaf] test  (benchmark): {test_dir}")

    train_data = collect_images_df(train_dir) if train_dir else pd.DataFrame(columns=[0,1])
    test_data  = collect_images_df(test_dir)  if test_dir  else pd.DataFrame(columns=[0,1])

    actual_classes = sorted(set(
        (train_data[1].unique().tolist() if len(train_data) else []) +
        (test_data[1].unique().tolist()  if len(test_data)  else [])
    ))

    if not actual_classes:
        print(f"  [ERROR] No classes found in DurianLeaf")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    print(f"  [DurianLeaf] {len(actual_classes)} classes: {actual_classes}")

    # Drop classes that cannot meet quota in EITHER pool
    _viable = []
    for _c in actual_classes:
        _rc = len(train_data[train_data[1] == _c]) if len(train_data) else 0
        _bc = len(test_data[test_data[1] == _c]) if len(test_data) else 0
        if _rc >= n_ref and _bc >= n_bench:
            _viable.append(_c)
        else:
            print(f'  [DROP CLASS] {_c!r}: ref={_rc}/{n_ref}, bench={_bc}/{n_bench} -- excluded')
    actual_classes = _viable
    if not actual_classes:
        print(f'  [DROP ALL] No classes meet quota for both pools in {name}')
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    ref_df   = sample_per_class(train_data, actual_classes, n_ref)
    bench_df = sample_per_class(test_data,  actual_classes, n_bench)
    ref_df   = save_split(ref_df,   name, "reference")
    bench_df = save_split(bench_df, name, "benchmark")

    n_r = sum(1 for p in ref_df[0]   if p)
    n_b = sum(1 for p in bench_df[0] if p)
    print(f"  [SAVED] {name}: reference={n_r}  benchmark={n_b}")

    if os.path.exists(dl):
        shutil.rmtree(dl, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {dl}")

    return name, ref_df, bench_df, actual_classes, desc


def load_EggplantDisease(n_ref, n_bench):
    """
    Eggplant Disease Recognition Dataset
    kagglehub: kamalmoha/eggplant-disease-recognition-dataset

    Structure:
      Eggplant Disease Recogn.../
        Original Images/
          Healthy Leaf/
          Insect Pest Disease/
          Leaf Spot Disease/
          Mosaic Virus Disease/
          Small Leaf Disease/
          White Mold Disease/
          Wilt Disease/
        Augmented Images/   <- ignored

    Crop: Eggplant. Single pool split into ref/bench.
    """
    name = "Eggplant Disease"
    desc = ("Eggplant Disease Recognition Dataset (kaggle: kamalmoha). "
            "7 classes: Healthy Leaf, Insect Pest Disease, Leaf Spot Disease, "
            "Mosaic Virus Disease, Small Leaf Disease, White Mold Disease, Wilt Disease. "
            "Original Images only (Augmented Images folder ignored).")

    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, [])
        return name, ref, bench, [], desc

    dl = os.path.join(DATA_ROOT, "eggplant-disease-dataset")
    if not os.path.exists(dl):
        if KAGGLEHUB_AVAILABLE:
            try:
                print(f"  [kagglehub] Downloading eggplant-disease-recognition-dataset ...")
                dl_path = kagglehub.dataset_download(
                    "kamalmoha/eggplant-disease-recognition-dataset")
                print(f"  Path to dataset files: {dl_path}")
                os.makedirs(dl, exist_ok=True)
                shutil.copytree(dl_path, dl, dirs_exist_ok=True)
            except Exception as e:
                print(f"  [kagglehub ERROR] {e} — trying kaggle API fallback")
                kaggle_download("kamalmoha/eggplant-disease-recognition-dataset", dl)
        else:
            kaggle_download("kamalmoha/eggplant-disease-recognition-dataset", dl)

    # ── Locate "Original Images" folder ───────────────────────────────────────
    orig_dir = None
    for candidate in [os.path.join(dl, "Original Images")]:
        if os.path.isdir(candidate):
            orig_dir = candidate
            break
    if orig_dir is None:
        for dirpath, dirnames, _ in os.walk(dl):
            dirnames[:] = [d for d in dirnames
                           if d != "Augmented Images" and not d.startswith(".")]
            if os.path.basename(dirpath) == "Original Images":
                orig_dir = dirpath
                break
        # Fallback: find any parent that has class-like subfolders (not Augmented)
        if orig_dir is None:
            for dirpath, dirnames, _ in os.walk(dl):
                dirnames[:] = [d for d in dirnames if not d.startswith(".")]
                subs = [d for d in dirnames if d not in ("Augmented Images",)]
                if subs and any("Disease" in d or "Leaf" in d for d in subs):
                    orig_dir = dirpath
                    break

    if orig_dir is None:
        print(f"  [ERROR] Could not locate Original Images folder in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    print(f"  [EggplantDisease] Original Images: {orig_dir}")

    data = collect_images_df(orig_dir)
    if len(data) == 0:
        print(f"  [ERROR] No images found in {orig_dir}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    actual_classes = sorted(data[1].unique().tolist())
    print(f"  [EggplantDisease] {len(actual_classes)} classes: {actual_classes}")

    ref, bench = split_save_cleanup(data, actual_classes, name, dl, n_ref, n_bench)
    return name, ref, bench, actual_classes, desc


def load_CottonDisease(n_ref, n_bench):
    """
    Cotton Disease Multi Transformation Dataset
    kaggle: shuvokumarbasak2030/cotton-disease-multi-transformation-dataset

    Structure:
      dataset/dataset/
        train/  Aphids/ Army_Worm/ Bacterial_Blight/ Healthy/
                Powdery_Mildew/ Target_Spot/          <- REFERENCE pool
        test/   (same classes)                        <- BENCHMARK pool
      augmentation/                                   <- ignored entirely

    Crop: Cotton. Train -> reference, test -> benchmark.
    """
    name    = "Cotton Disease"
    classes = ["Aphids", "Army_Worm", "Bacterial_Blight",
               "Healthy", "Powdery_Mildew", "Target_Spot"]
    desc    = ("Cotton Disease Multi Transformation Dataset (kaggle: shuvokumarbasak2030). "
               "6 classes: Aphids, Army Worm, Bacterial Blight, Healthy, "
               "Powdery Mildew, Target Spot. Train/test splits; augmentation folder ignored.")

    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc

    dl = os.path.join(DATA_ROOT, "cotton-disease-dataset")
    if not os.path.exists(dl):
        if KAGGLEHUB_AVAILABLE:
            try:
                print(f"  [kagglehub] Downloading cotton-disease-multi-transformation-dataset ...")
                dl_path = kagglehub.dataset_download(
                    "shuvokumarbasak2030/cotton-disease-multi-transformation-dataset")
                print(f"  Path to dataset files: {dl_path}")
                os.makedirs(dl, exist_ok=True)
                shutil.copytree(dl_path, dl, dirs_exist_ok=True)
            except Exception as e:
                print(f"  [kagglehub ERROR] {e} — trying kaggle API fallback")
                kaggle_download(
                    "shuvokumarbasak2030/cotton-disease-multi-transformation-dataset", dl)
        else:
            kaggle_download(
                "shuvokumarbasak2030/cotton-disease-multi-transformation-dataset", dl)

    def _find_split(split_name):
        """Locate train/ or test/ under dataset/dataset/, ignoring augmentation/."""
        for candidate in [
            os.path.join(dl, "dataset", "dataset", split_name),
            os.path.join(dl, "dataset", split_name),
            os.path.join(dl, split_name),
        ]:
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))
                        and not d.startswith(".")]
                if subs:
                    return candidate
        for dirpath, dirnames, _ in os.walk(dl):
            # Never descend into augmentation folder
            dirnames[:] = [d for d in dirnames
                           if not d.startswith(".") and d.lower() != "augmentation"]
            if os.path.basename(dirpath).lower() == split_name.lower():
                subs = [d for d in os.listdir(dirpath)
                        if os.path.isdir(os.path.join(dirpath, d))]
                if subs:
                    return dirpath
        return None

    train_dir = _find_split("train")
    test_dir  = _find_split("test")

    if train_dir is None and test_dir is None:
        print(f"  [ERROR] Could not locate Cotton train/test folders in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    print(f"  [CottonDisease] train (reference): {train_dir}")
    print(f"  [CottonDisease] test  (benchmark): {test_dir}")

    train_data = collect_images_df(train_dir) if train_dir else pd.DataFrame(columns=[0,1])
    test_data  = collect_images_df(test_dir)  if test_dir  else pd.DataFrame(columns=[0,1])

    actual_classes = sorted(set(
        (train_data[1].unique().tolist() if len(train_data) else []) +
        (test_data[1].unique().tolist()  if len(test_data)  else [])
    )) or classes

    # Drop classes that cannot meet quota in both pools
    _viable = []
    for _c in actual_classes:
        _rc = len(train_data[train_data[1] == _c]) if len(train_data) else 0
        _bc = len(test_data[test_data[1]   == _c]) if len(test_data)  else 0
        if _rc >= n_ref and _bc >= n_bench:
            _viable.append(_c)
        else:
            print(f"  [DROP CLASS] {_c!r}: ref={_rc}/{n_ref}, bench={_bc}/{n_bench} -- excluded")
    actual_classes = _viable
    if not actual_classes:
        print(f"  [DROP ALL] No viable classes for {name}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    ref_df   = sample_per_class(train_data, actual_classes, n_ref)
    bench_df = sample_per_class(test_data,  actual_classes, n_bench)
    ref_df   = save_split(ref_df,   name, "reference")
    bench_df = save_split(bench_df, name, "benchmark")

    n_r = sum(1 for p in ref_df[0]   if p)
    n_b = sum(1 for p in bench_df[0] if p)
    print(f"  [SAVED] {name}: reference={n_r}  benchmark={n_b}")

    if os.path.exists(dl):
        shutil.rmtree(dl, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {dl}")

    return name, ref_df, bench_df, actual_classes, desc


def load_PumpkinLeaf(n_ref, n_bench):
    """
    Pumpkin Leaf Disease Multi Transformation Dataset
    kaggle: shuvokumarbasak2030/pumpkin-leaf-disease-multi-transformation-dataset

    Structure:
      dataset/dataset/
        train/  Bacterial_Leaf_Spot/ Downy_Mildew/ Healthy_Leaf/
                Mosaic_Disease/ Powdery_Mildew/     <- REFERENCE pool
        test/   (same classes)                      <- BENCHMARK pool
        val/    (same classes)                      <- ignored
      augmentation/                                 <- ignored

    Crop: Pumpkin. Train -> reference, test -> benchmark.
    """
    name    = "Pumpkin Leaf"
    classes = ["Bacterial_Leaf_Spot", "Downy_Mildew", "Healthy_Leaf",
               "Mosaic_Disease", "Powdery_Mildew"]
    desc    = ("Pumpkin Leaf Disease Multi Transformation Dataset (kaggle: shuvokumarbasak2030). "
               "5 classes: Bacterial Leaf Spot, Downy Mildew, Healthy Leaf, "
               "Mosaic Disease, Powdery Mildew. Train/test/val splits; "
               "augmentation folder and val split ignored.")

    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc

    dl = os.path.join(DATA_ROOT, "pumpkin-leaf-disease-dataset")
    if not os.path.exists(dl):
        if KAGGLEHUB_AVAILABLE:
            try:
                print(f"  [kagglehub] Downloading pumpkin-leaf-disease-multi-transformation-dataset ...")
                dl_path = kagglehub.dataset_download(
                    "shuvokumarbasak2030/pumpkin-leaf-disease-multi-transformation-dataset")
                print(f"  Path to dataset files: {dl_path}")
                os.makedirs(dl, exist_ok=True)
                shutil.copytree(dl_path, dl, dirs_exist_ok=True)
            except Exception as e:
                print(f"  [kagglehub ERROR] {e} — trying kaggle API fallback")
                kaggle_download(
                    "shuvokumarbasak2030/pumpkin-leaf-disease-multi-transformation-dataset", dl)
        else:
            kaggle_download(
                "shuvokumarbasak2030/pumpkin-leaf-disease-multi-transformation-dataset", dl)

    def _find_split(split_name):
        for candidate in [
            os.path.join(dl, "dataset", "dataset", split_name),
            os.path.join(dl, "dataset", split_name),
            os.path.join(dl, split_name),
        ]:
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))
                        and not d.startswith(".")]
                if subs:
                    return candidate
        for dirpath, dirnames, _ in os.walk(dl):
            dirnames[:] = [d for d in dirnames
                           if not d.startswith(".")
                           and d.lower() not in ("augmentation", "val")]
            if os.path.basename(dirpath).lower() == split_name.lower():
                subs = [d for d in os.listdir(dirpath)
                        if os.path.isdir(os.path.join(dirpath, d))]
                if subs:
                    return dirpath
        return None

    train_dir = _find_split("train")
    test_dir  = _find_split("test")

    if train_dir is None and test_dir is None:
        print(f"  [ERROR] Could not locate Pumpkin train/test folders in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    print(f"  [PumpkinLeaf] train (reference): {train_dir}")
    print(f"  [PumpkinLeaf] test  (benchmark): {test_dir}")

    train_data = collect_images_df(train_dir) if train_dir else pd.DataFrame(columns=[0,1])
    test_data  = collect_images_df(test_dir)  if test_dir  else pd.DataFrame(columns=[0,1])

    actual_classes = sorted(set(
        (train_data[1].unique().tolist() if len(train_data) else []) +
        (test_data[1].unique().tolist()  if len(test_data)  else [])
    )) or classes

    _viable = []
    for _c in actual_classes:
        _rc = len(train_data[train_data[1] == _c]) if len(train_data) else 0
        _bc = len(test_data[test_data[1]   == _c]) if len(test_data)  else 0
        if _rc >= n_ref and _bc >= n_bench:
            _viable.append(_c)
        else:
            print(f"  [DROP CLASS] {_c!r}: ref={_rc}/{n_ref}, bench={_bc}/{n_bench} -- excluded")
    actual_classes = _viable
    if not actual_classes:
        print(f"  [DROP ALL] No viable classes for {name}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    ref_df   = sample_per_class(train_data, actual_classes, n_ref)
    bench_df = sample_per_class(test_data,  actual_classes, n_bench)
    ref_df   = save_split(ref_df,   name, "reference")
    bench_df = save_split(bench_df, name, "benchmark")

    n_r = sum(1 for p in ref_df[0]   if p)
    n_b = sum(1 for p in bench_df[0] if p)
    print(f"  [SAVED] {name}: reference={n_r}  benchmark={n_b}")
    if os.path.exists(dl):
        shutil.rmtree(dl, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {dl}")
    return name, ref_df, bench_df, actual_classes, desc


def load_RoseLeaf(n_ref, n_bench):
    """
    Rose Leaf Disease Multi Transformation Dataset
    kaggle: shuvokumarbasak2030/rose-leaf-disease-multi-transformation-dataset

    Structure:
      dataset/dataset/
        train/  Black_Spot/ Downy_Mildew/ Dry_Leaf/ Healthy_Leaf/ Leaf_Holes/
        test/   (same classes)   <- BENCHMARK pool
        val/                     <- ignored
      augmentation/              <- ignored
      raw_mixed/                 <- ignored

    Crop: Rose. Train -> reference, test -> benchmark.
    """
    name    = "Rose Leaf"
    classes = ["Black_Spot", "Downy_Mildew", "Dry_Leaf", "Healthy_Leaf", "Leaf_Holes"]
    desc    = ("Rose Leaf Disease Multi Transformation Dataset (kaggle: shuvokumarbasak2030). "
               "5 classes: Black Spot, Downy Mildew, Dry Leaf, Healthy Leaf, Leaf Holes. "
               "Train/test splits used; augmentation, val, raw_mixed ignored.")

    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc

    dl = os.path.join(DATA_ROOT, "rose-leaf-disease-dataset")
    if not os.path.exists(dl):
        if KAGGLEHUB_AVAILABLE:
            try:
                print(f"  [kagglehub] Downloading rose-leaf-disease-multi-transformation-dataset ...")
                dl_path = kagglehub.dataset_download(
                    "shuvokumarbasak2030/rose-leaf-disease-multi-transformation-dataset")
                print(f"  Path to dataset files: {dl_path}")
                os.makedirs(dl, exist_ok=True)
                shutil.copytree(dl_path, dl, dirs_exist_ok=True)
            except Exception as e:
                print(f"  [kagglehub ERROR] {e} — trying kaggle API fallback")
                kaggle_download(
                    "shuvokumarbasak2030/rose-leaf-disease-multi-transformation-dataset", dl)
        else:
            kaggle_download(
                "shuvokumarbasak2030/rose-leaf-disease-multi-transformation-dataset", dl)

    _IGNORE = {"augmentation", "val", "raw_mixed"}

    def _find_split(split_name):
        for candidate in [
            os.path.join(dl, "dataset", "dataset", split_name),
            os.path.join(dl, "dataset", split_name),
            os.path.join(dl, split_name),
        ]:
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))
                        and not d.startswith(".")]
                if subs:
                    return candidate
        for dirpath, dirnames, _ in os.walk(dl):
            dirnames[:] = [d for d in dirnames
                           if not d.startswith(".") and d.lower() not in _IGNORE]
            if os.path.basename(dirpath).lower() == split_name.lower():
                subs = [d for d in os.listdir(dirpath)
                        if os.path.isdir(os.path.join(dirpath, d))]
                if subs:
                    return dirpath
        return None

    train_dir = _find_split("train")
    test_dir  = _find_split("test")

    if train_dir is None and test_dir is None:
        print(f"  [ERROR] Could not locate Rose train/test folders in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    print(f"  [RoseLeaf] train (reference): {train_dir}")
    print(f"  [RoseLeaf] test  (benchmark): {test_dir}")

    train_data = collect_images_df(train_dir) if train_dir else pd.DataFrame(columns=[0,1])
    test_data  = collect_images_df(test_dir)  if test_dir  else pd.DataFrame(columns=[0,1])

    actual_classes = sorted(set(
        (train_data[1].unique().tolist() if len(train_data) else []) +
        (test_data[1].unique().tolist()  if len(test_data)  else [])
    )) or classes

    _viable = []
    for _c in actual_classes:
        _rc = len(train_data[train_data[1] == _c]) if len(train_data) else 0
        _bc = len(test_data[test_data[1]   == _c]) if len(test_data)  else 0
        if _rc >= n_ref and _bc >= n_bench:
            _viable.append(_c)
        else:
            print(f"  [DROP CLASS] {_c!r}: ref={_rc}/{n_ref}, bench={_bc}/{n_bench} -- excluded")
    actual_classes = _viable
    if not actual_classes:
        print(f"  [DROP ALL] No viable classes for {name}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    ref_df   = sample_per_class(train_data, actual_classes, n_ref)
    bench_df = sample_per_class(test_data,  actual_classes, n_bench)
    ref_df   = save_split(ref_df,   name, "reference")
    bench_df = save_split(bench_df, name, "benchmark")

    n_r = sum(1 for p in ref_df[0]   if p)
    n_b = sum(1 for p in bench_df[0] if p)
    print(f"  [SAVED] {name}: reference={n_r}  benchmark={n_b}")
    if os.path.exists(dl):
        shutil.rmtree(dl, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {dl}")
    return name, ref_df, bench_df, actual_classes, desc


def load_CoconutDisease(n_ref, n_bench):
    """
    Coconut Disease Multi Transformation STTV Dataset
    kaggle: shuvokumarbasak2030/coconut-disease-multi-transformation-sttv-dataset

    Structure:
      dataset/
        test/  CCI_Caterpillars/ CCI_Leaflets/ Healthy_Leaves/
               WCLWD_DryingofLeaves/ WCLWD_Flaccidity/ WCLWD_Yellowing/  <- BENCHMARK
        train/ (same classes)                                             <- REFERENCE
        val/                                                              <- ignored

    Crop: Coconut. Train -> reference, test -> benchmark.
    """
    name    = "Coconut Disease"
    classes = ["CCI_Caterpillars", "CCI_Leaflets", "Healthy_Leaves",
               "WCLWD_DryingofLeaves", "WCLWD_Flaccidity", "WCLWD_Yellowing"]
    desc    = ("Coconut Disease Multi Transformation STTV Dataset (kaggle: shuvokumarbasak2030). "
               "6 classes: CCI Caterpillars, CCI Leaflets, Healthy Leaves, "
               "WCLWD Drying of Leaves, WCLWD Flaccidity, WCLWD Yellowing. "
               "Train/test splits used; val ignored.")

    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc

    dl = os.path.join(DATA_ROOT, "coconut-disease-dataset")
    if not os.path.exists(dl):
        if KAGGLEHUB_AVAILABLE:
            try:
                print(f"  [kagglehub] Downloading coconut-disease-multi-transformation-sttv-dataset ...")
                dl_path = kagglehub.dataset_download(
                    "shuvokumarbasak2030/coconut-disease-multi-transformation-sttv-dataset")
                print(f"  Path to dataset files: {dl_path}")
                os.makedirs(dl, exist_ok=True)
                shutil.copytree(dl_path, dl, dirs_exist_ok=True)
            except Exception as e:
                print(f"  [kagglehub ERROR] {e} — trying kaggle API fallback")
                kaggle_download(
                    "shuvokumarbasak2030/coconut-disease-multi-transformation-sttv-dataset", dl)
        else:
            kaggle_download(
                "shuvokumarbasak2030/coconut-disease-multi-transformation-sttv-dataset", dl)

    _IGNORE = {"val", "augmentation"}

    def _find_split(split_name):
        for candidate in [
            os.path.join(dl, "dataset", split_name),
            os.path.join(dl, split_name),
        ]:
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))
                        and not d.startswith(".")]
                if subs:
                    return candidate
        for dirpath, dirnames, _ in os.walk(dl):
            dirnames[:] = [d for d in dirnames
                           if not d.startswith(".") and d.lower() not in _IGNORE]
            if os.path.basename(dirpath).lower() == split_name.lower():
                subs = [d for d in os.listdir(dirpath)
                        if os.path.isdir(os.path.join(dirpath, d))]
                if subs:
                    return dirpath
        return None

    train_dir = _find_split("train")
    test_dir  = _find_split("test")

    if train_dir is None and test_dir is None:
        print(f"  [ERROR] Could not locate Coconut train/test folders in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    print(f"  [CoconutDisease] train (reference): {train_dir}")
    print(f"  [CoconutDisease] test  (benchmark): {test_dir}")

    train_data = collect_images_df(train_dir) if train_dir else pd.DataFrame(columns=[0,1])
    test_data  = collect_images_df(test_dir)  if test_dir  else pd.DataFrame(columns=[0,1])

    actual_classes = sorted(set(
        (train_data[1].unique().tolist() if len(train_data) else []) +
        (test_data[1].unique().tolist()  if len(test_data)  else [])
    )) or classes

    _viable = []
    for _c in actual_classes:
        _rc = len(train_data[train_data[1] == _c]) if len(train_data) else 0
        _bc = len(test_data[test_data[1]   == _c]) if len(test_data)  else 0
        if _rc >= n_ref and _bc >= n_bench:
            _viable.append(_c)
        else:
            print(f"  [DROP CLASS] {_c!r}: ref={_rc}/{n_ref}, bench={_bc}/{n_bench} -- excluded")
    actual_classes = _viable
    if not actual_classes:
        print(f"  [DROP ALL] No viable classes for {name}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    ref_df   = sample_per_class(train_data, actual_classes, n_ref)
    bench_df = sample_per_class(test_data,  actual_classes, n_bench)
    ref_df   = save_split(ref_df,   name, "reference")
    bench_df = save_split(bench_df, name, "benchmark")

    n_r = sum(1 for p in ref_df[0]   if p)
    n_b = sum(1 for p in bench_df[0] if p)
    print(f"  [SAVED] {name}: reference={n_r}  benchmark={n_b}")
    if os.path.exists(dl):
        shutil.rmtree(dl, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {dl}")
    return name, ref_df, bench_df, actual_classes, desc


def load_VanillaDisease(n_ref, n_bench):
    """
    Vanilla Disease Multi Transformation Dataset
    kaggle: shuvokumarbasak2030/vanilla-disease-multi-transformation-dataset

    Structure (different from other shuvokumarbasak2030 datasets):
      <class>/          e.g. black_spots/ healthy/ rotten_stem/ ...
        original/, augmented/, ...  <- ALL 17 subfolders used

    No train/test split — single pool (all variants) split into ref/bench here.
    Crop: Vanilla.
    """
    name = "Vanilla Disease"
    desc = ("Vanilla Disease Multi Transformation Dataset (kaggle: shuvokumarbasak2030). "
            "Classes include black spots, healthy, rotten stem etc. "
            "All image variants across all subfolders used; single pool split into ref/bench.")

    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, [])
        return name, ref, bench, [], desc

    dl = os.path.join(DATA_ROOT, "vanilla-disease-dataset")
    if not os.path.exists(dl):
        if KAGGLEHUB_AVAILABLE:
            try:
                print(f"  [kagglehub] Downloading vanilla-disease-multi-transformation-dataset ...")
                dl_path = kagglehub.dataset_download(
                    "shuvokumarbasak2030/vanilla-disease-multi-transformation-dataset")
                print(f"  Path to dataset files: {dl_path}")
                os.makedirs(dl, exist_ok=True)
                shutil.copytree(dl_path, dl, dirs_exist_ok=True)
            except Exception as e:
                print(f"  [kagglehub ERROR] {e} — trying kaggle API fallback")
                kaggle_download(
                    "shuvokumarbasak2030/vanilla-disease-multi-transformation-dataset", dl)
        else:
            kaggle_download(
                "shuvokumarbasak2030/vanilla-disease-multi-transformation-dataset", dl)

    # ── Locate dataset root — folder whose subdirs are class names ────────────
    # Structure: dl/<root>/<class>/original/*.jpg
    # Find a folder containing subdirs that each have an "original" subfolder
    dataset_root = None
    for dirpath, dirnames, _ in os.walk(dl):
        dirnames[:] = [d for d in dirnames if not d.startswith(".")]
        has_original = [d for d in dirnames
                        if os.path.isdir(os.path.join(dirpath, d, "original"))]
        if has_original:
            dataset_root = dirpath
            break
    # Fallback: any folder containing class-like subdirs with image subfolders
    if dataset_root is None:
        for dirpath, dirnames, _ in os.walk(dl):
            dirnames[:] = [d for d in dirnames if not d.startswith(".")]
            if len(dirnames) >= 2:
                dataset_root = dirpath
                break

    if dataset_root is None:
        print(f"  [ERROR] Could not locate Vanilla dataset root in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    print(f"  [VanillaDisease] Dataset root: {dataset_root}")

    # ── Collect images from ALL subfolders of each class ────────────────────
    rows = []
    classes_found = []
    for cls_dir in sorted(os.listdir(dataset_root)):
        cls_path = os.path.join(dataset_root, cls_dir)
        if not os.path.isdir(cls_path) or cls_dir.startswith("."):
            continue

        # Collect recursively from all subfolders (all augmentation variants included)
        cls_key = safe_name(cls_dir)
        cls_imgs = collect_images_recursive(cls_path)
        if not cls_imgs:
            print(f"  [SKIP CLASS] '{cls_dir}': no images found")
            continue

        classes_found.append(cls_key)
        for img_path in cls_imgs:
            rows.append({0: img_path, 1: cls_key})


    if not rows:
        print(f"  [ERROR] No images found for VanillaDisease")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    data = pd.DataFrame(rows)
    print(f"  [VanillaDisease] {len(classes_found)} classes, {len(data)} images total")

    ref, bench = split_save_cleanup(data, classes_found, name, dl, n_ref, n_bench)
    return name, ref, bench, classes_found, desc


def load_SugarLeafIDN(n_ref, n_bench):
    """
    SugarLeaf-IDN Dataset (sugarcane leaf diseases, Indonesia)
    kagglehub: bettydpuspasari/sugarleafidn

    Structure:
      dataset_Resize_224x224_NB_test100/
        train/  0 Healthy/ 1 Pokkahboeng/ 2 Common Rust (Karat Daun)/
                3 Eye Spot (cincin)/ 4 yellow Spot (noda Kuning)/
                5 red spot (noda Merah)/ 6 Mosaic/
                7 Streak Mosaic SCSMV (bergaris)/ 8 Leaf Scald (blendok)/
        test/   (same 9 classes, 100 images each)  <- BENCHMARK
        validation/                                <- ignored

    Numeric prefix stripped from class names (e.g. "1 Pokkahboeng" -> "Pokkahboeng").
    Crop: Sugarcane. Train -> reference, test -> benchmark.
    """
    name = "SugarLeaf IDN"
    desc = ("SugarLeaf-IDN dataset (kaggle: bettydpuspasari). 9 classes of sugarcane "
            "leaf diseases from East Java, Indonesia. 224x224 px, field conditions. "
            "Classes: Healthy, Pokkahboeng, Common Rust, Eye Spot, Yellow Spot, "
            "Red Spot, Mosaic, Streak Mosaic SCSMV, Leaf Scald.")

    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, [])
        return name, ref, bench, [], desc

    dl = os.path.join(DATA_ROOT, "sugarleafidn")
    if not os.path.exists(dl):
        if KAGGLEHUB_AVAILABLE:
            try:
                print(f"  [kagglehub] Downloading sugarleafidn ...")
                dl_path = kagglehub.dataset_download("bettydpuspasari/sugarleafidn")
                print(f"  Path to dataset files: {dl_path}")
                os.makedirs(dl, exist_ok=True)
                shutil.copytree(dl_path, dl, dirs_exist_ok=True)
            except Exception as e:
                print(f"  [kagglehub ERROR] {e} — trying kaggle API fallback")
                kaggle_download("bettydpuspasari/sugarleafidn", dl)
        else:
            kaggle_download("bettydpuspasari/sugarleafidn", dl)

    _IGNORE = {"validation", "val", "augmentation"}

    def _find_split(split_name):
        # Try one and two levels deep first
        for candidate in [
            os.path.join(dl, split_name),
        ]:
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))
                        and not d.startswith(".")]
                if subs:
                    return candidate
        # One level deep (dataset_Resize_.../train/)
        for top in os.listdir(dl):
            top_path = os.path.join(dl, top)
            if not os.path.isdir(top_path) or top.startswith("."): continue
            candidate = os.path.join(top_path, split_name)
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))]
                if subs:
                    return candidate
        # Walk fallback
        for dirpath, dirnames, _ in os.walk(dl):
            dirnames[:] = [d for d in dirnames
                           if not d.startswith(".") and d.lower() not in _IGNORE]
            if os.path.basename(dirpath).lower() == split_name.lower():
                subs = [d for d in os.listdir(dirpath)
                        if os.path.isdir(os.path.join(dirpath, d))]
                if subs:
                    return dirpath
        return None

    train_dir = _find_split("train")
    test_dir  = _find_split("test")

    if train_dir is None and test_dir is None:
        print(f"  [ERROR] Could not locate SugarLeafIDN train/test folders in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    print(f"  [SugarLeafIDN] train (reference): {train_dir}")
    print(f"  [SugarLeafIDN] test  (benchmark): {test_dir}")

    def _collect_strip_prefix(split_dir):
        """Collect images; strip leading numeric prefix from class folder names."""
        if not split_dir or not os.path.isdir(split_dir):
            return pd.DataFrame(columns=[0, 1])
        rows = []
        for cls_dir in os.listdir(split_dir):
            cls_path = os.path.join(split_dir, cls_dir)
            if not os.path.isdir(cls_path) or cls_dir.startswith("."):
                continue
            # Strip leading "N " prefix: "1 Pokkahboeng" -> "Pokkahboeng"
            import re as _re
            clean = _re.sub(r"^\d+\s+", "", cls_dir).strip()
            cls_key = safe_name(clean) if clean else safe_name(cls_dir)
            for f in os.listdir(cls_path):
                if f.lower().endswith(IMAGE_EXT):
                    rows.append({0: os.path.join(cls_path, f), 1: cls_key})
        return pd.DataFrame(rows) if rows else pd.DataFrame(columns=[0, 1])

    train_data = _collect_strip_prefix(train_dir)
    test_data  = _collect_strip_prefix(test_dir)

    actual_classes = sorted(set(
        (train_data[1].unique().tolist() if len(train_data) else []) +
        (test_data[1].unique().tolist()  if len(test_data)  else [])
    ))

    if not actual_classes:
        print(f"  [ERROR] No classes found for SugarLeafIDN")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    print(f"  [SugarLeafIDN] {len(actual_classes)} classes: {actual_classes}")

    _viable = []
    for _c in actual_classes:
        _rc = len(train_data[train_data[1] == _c]) if len(train_data) else 0
        _bc = len(test_data[test_data[1]   == _c]) if len(test_data)  else 0
        if _rc >= n_ref and _bc >= n_bench:
            _viable.append(_c)
        else:
            print(f"  [DROP CLASS] {_c!r}: ref={_rc}/{n_ref}, bench={_bc}/{n_bench} -- excluded")
    actual_classes = _viable
    if not actual_classes:
        print(f"  [DROP ALL] No viable classes for {name}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    ref_df   = sample_per_class(train_data, actual_classes, n_ref)
    bench_df = sample_per_class(test_data,  actual_classes, n_bench)
    ref_df   = save_split(ref_df,   name, "reference")
    bench_df = save_split(bench_df, name, "benchmark")

    n_r = sum(1 for p in ref_df[0]   if p)
    n_b = sum(1 for p in bench_df[0] if p)
    print(f"  [SAVED] {name}: reference={n_r}  benchmark={n_b}")
    if os.path.exists(dl):
        shutil.rmtree(dl, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {dl}")
    return name, ref_df, bench_df, actual_classes, desc


def load_CucumberZenodo(n_ref, n_bench):
    """
    Cucumber Disease and Freshness Classification Dataset – Curated Annotations
    Zenodo DOI: 10.5281/zenodo.16816441
    URL: https://zenodo.org/records/16816441

    Structure (flat class folders, no train/test split):
      Anthracnose/
      Bacterial_Wilt/
      Belly_Rot/
      Downy_Mildew/
      Fresh_Cucumber/
      Fresh_Leaf/
      Pythium_Fruit_Rot/

    Downloaded as a zip via Zenodo REST API.
    Crop: Cucumber. Single pool split into ref/bench.
    """
    name    = "Cucumber Zenodo"
    classes = ["Anthracnose", "Bacterial_Wilt", "Belly_Rot", "Downy_Mildew",
               "Fresh_Cucumber", "Fresh_Leaf", "Pythium_Fruit_Rot"]
    desc    = ("Cucumber Disease and Freshness Classification Dataset – Curated Annotations "
               "(Zenodo DOI: 10.5281/zenodo.16816441). 7 classes: Anthracnose, Bacterial Wilt, "
               "Belly Rot, Downy Mildew, Fresh Cucumber, Fresh Leaf, Pythium Fruit Rot. "
               "Manually annotated with Label Studio.")

    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc

    dl = os.path.join(DATA_ROOT, "cucumber-zenodo-dataset")
    zip_path = dl + ".zip"

    if not os.path.exists(dl):
        os.makedirs(dl, exist_ok=True)
        # ── Download via Zenodo REST API ──────────────────────────────────────
        RECORD_ID = "16816441"
        api_url   = f"https://zenodo.org/api/records/{RECORD_ID}"
        try:
            import urllib.request, json as _json
            print(f"  [Zenodo] Fetching record metadata from {api_url} ...")
            with urllib.request.urlopen(api_url, timeout=30) as resp:
                meta = _json.loads(resp.read().decode())
            files = meta.get("files", [])
            zip_files = [f for f in files if f.get("key","").endswith(".zip")]
            dl_files  = zip_files if zip_files else files
            if not dl_files:
                print(f"  [ERROR] No files found in Zenodo record {RECORD_ID}")
                return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc
            for file_info in dl_files:
                file_url  = file_info.get("links", {}).get("self") or file_info.get("download_url")
                file_name = file_info.get("key", "dataset.zip")
                dest      = os.path.join(dl, file_name)
                print(f"  [Zenodo] Downloading {file_name} ...")
                urllib.request.urlretrieve(file_url, dest)
                if dest.endswith(".zip"):
                    import zipfile
                    print(f"  [Zenodo] Extracting {file_name} ...")
                    with zipfile.ZipFile(dest, "r") as z:
                        z.extractall(dl)
                    os.remove(dest)
        except Exception as e:
            print(f"  [Zenodo ERROR] {e}")
            return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    # ── Locate dataset root — folder whose subdirs match class names ──────────
    dataset_root = None
    for dirpath, dirnames, _ in os.walk(dl):
        dirnames[:] = [d for d in dirnames if not d.startswith(".") and d != "__MACOSX"]
        hits = [d for d in dirnames
                if any(kw in d.lower() for kw in
                       ("anthracnose","bacterial","belly","downy","fresh","pythium","cucumber"))]
        if len(hits) >= 3:
            dataset_root = dirpath
            break

    if dataset_root is None:
        dataset_root = dl  # best-effort fallback
    print(f"  [CucumberZenodo] Dataset root: {dataset_root}")

    data = collect_images_df(dataset_root)
    if len(data) == 0:
        print(f"  [ERROR] No images found in {dataset_root}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    actual_classes = sorted(data[1].unique().tolist())
    print(f"  [CucumberZenodo] {len(actual_classes)} classes: {actual_classes}")

    ref, bench = split_save_cleanup(data, actual_classes, name, dl, n_ref, n_bench)
    return name, ref, bench, actual_classes, desc


def load_Cauliflower(n_ref, n_bench):
    """
    Cauliflower Disease Multi Transformation Dataset
    kaggle: shuvokumarbasak2030/cauliflower-disease-multi-transformation-dataset
    Structure: dataset/dataset/train/{Bacterial_Spot_Rot, Black_Rot, Downy_Mildew, No_disease}
    4 classes; uses train split for sampling.
    """
    name    = "Cauliflower"
    classes = ['Bacterial_Spot_Rot', 'Black_Rot', 'Downy_Mildew', 'No_disease']
    desc    = ("Cauliflower leaf disease dataset with multi-transformation augmentation. "
               "Covers Bacterial Spot Rot, Black Rot, Downy Mildew and No Disease "
               "(healthy). Train/val/test splits available; train split used here.")

    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc

    dl   = os.path.join(DATA_ROOT, "cauliflower-disease")

    if not os.path.exists(dl):
        kaggle_download(
            "shuvokumarbasak2030/cauliflower-disease-multi-transformation-dataset", dl)

    # Confirmed structure (Kaggle Data Explorer):
    #   dataset/dataset/train/  <-- REFERENCE images
    #   dataset/dataset/test/   <-- BENCHMARK images
    #   dataset/dataset/val/    <-- ignored
    #   augmentation/           <-- ignored entirely

    def _find_cauliflower_split(split_name):
        # Try explicit paths first
        for candidate in [
            os.path.join(dl, "dataset", "dataset", split_name),
            os.path.join(dl, "dataset", split_name),
            os.path.join(dl, split_name),
        ]:
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))
                        and not d.startswith('.')]
                if subs:
                    return candidate
        # Walk fallback — only match folder named split_name inside a dataset parent
        for dirpath, dirnames, _ in os.walk(dl):
            dirnames[:] = [d for d in dirnames if not d.startswith('.')]
            if (os.path.basename(dirpath) == split_name
                    and "dataset" in dirpath.lower()):
                subs = [d for d in os.listdir(dirpath)
                        if os.path.isdir(os.path.join(dirpath, d))
                        and not d.startswith('.')]
                if subs:
                    return dirpath
        return None

    train_dir = _find_cauliflower_split("train")
    test_dir  = _find_cauliflower_split("test")

    if train_dir is None and test_dir is None:
        print(f"  [ERROR] Could not locate Cauliflower train or test folders in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    print(f"  [Cauliflower] train (reference) : {train_dir}")
    print(f"  [Cauliflower] test  (benchmark) : {test_dir}")

    train_data = collect_images_df(train_dir) if train_dir else pd.DataFrame(columns=[0,1])
    test_data  = collect_images_df(test_dir)  if test_dir  else pd.DataFrame(columns=[0,1])

    if len(train_data) == 0 and len(test_data) == 0:
        print(f"  [ERROR] No images found in train or test folders")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    # Sample reference images from train/, benchmark images from test/
    # Drop classes that cannot meet quota in EITHER pool
    _viable = []
    for _c in classes:
        _rc = len(train_data[train_data[1] == _c]) if len(train_data) else 0
        _bc = len(test_data[test_data[1] == _c]) if len(test_data) else 0
        if _rc >= n_ref and _bc >= n_bench:
            _viable.append(_c)
        else:
            print(f'  [DROP CLASS] {_c!r}: ref={_rc}/{n_ref}, bench={_bc}/{n_bench} -- excluded')
    classes = _viable
    if not classes:
        print(f'  [DROP ALL] No classes meet quota for both pools in {name}')
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    ref_df   = sample_per_class(train_data, classes, n_ref)
    bench_df = sample_per_class(test_data,  classes, n_bench)

    ref_df   = save_split(ref_df,   name, "reference")
    bench_df = save_split(bench_df, name, "benchmark")

    n_r = sum(1 for p in ref_df[0]   if p)
    n_b = sum(1 for p in bench_df[0] if p)
    print(f"  [SAVED] reference={n_r}  benchmark={n_b} "
          f"-> samples/{safe_name(name)}/")

    # Clean up raw download
    if os.path.exists(dl):
        shutil.rmtree(dl, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {dl}")

    return name, ref_df, bench_df, classes, desc


def load_NewPlantDiseases(n_ref, n_bench):
    """
    New Plant Diseases Dataset (vipoooool/new-plant-diseases-dataset) — 1.43 GB
    kagglehub: vipoooool/new-plant-diseases-dataset

    Structure:
      New Plant Diseases Dataset/New Plant Diseases Da.../
        train/  <Crop>___<Disease>/   <- class subfolders, REFERENCE pool
        test/test/                    <- flat folder, filenames encode class, BENCHMARK pool
          AppleCedarRust1.JPG
          AppleScab1.JPG

    Class format: "Apple___Apple_scab" (triple underscore, same as PlantVillage tfds).
    Train split -> reference. Flat test folder -> benchmark via filename prefix matching.
    """
    name = "New Plant Diseases"
    desc = ("New Plant Diseases Dataset (Kaggle: vipoooool). 1.43 GB, 38 classes "
            "across 14 crop species using triple-underscore Crop___Disease naming. "
            "Train split used for reference; flat test folder for benchmark.")

    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, [])
        src_tag = safe_name(name)[:12]
        saved_classes = []
        for crop_d in (os.listdir(REF_DIR) if os.path.isdir(REF_DIR) else []):
            cp = os.path.join(REF_DIR, crop_d)
            if not os.path.isdir(cp): continue
            for dis_d in os.listdir(cp):
                dp = os.path.join(cp, dis_d)
                if not os.path.isdir(dp): continue
                if any(src_tag in f for f in os.listdir(dp)
                       if f.lower().endswith(IMAGE_EXT)):
                    saved_classes.append(f"{crop_d}__{dis_d}")
        return name, ref, bench, sorted(saved_classes), desc

    dl = os.path.join(DATA_ROOT, "new-plant-diseases")
    if not os.path.exists(dl):
        if KAGGLEHUB_AVAILABLE:
            try:
                print(f"  [kagglehub] Downloading New Plant Diseases dataset ...")
                dl_path = kagglehub.dataset_download(
                    "vipoooool/new-plant-diseases-dataset")
                print(f"  Path to dataset files: {dl_path}")
                os.makedirs(dl, exist_ok=True)
                shutil.copytree(dl_path, dl, dirs_exist_ok=True)
            except Exception as e:
                print(f"  [kagglehub ERROR] {e} — trying kaggle API fallback")
                kaggle_download("vipoooool/new-plant-diseases-dataset", dl)
        else:
            kaggle_download("vipoooool/new-plant-diseases-dataset", dl)

    # ── Locate train folder (has Crop___Disease subfolders) ───────────────────
    train_dir = None
    for candidate in [
        os.path.join(dl, "New Plant Diseases Dataset", "New Plant Diseases Dataset", "train"),
        os.path.join(dl, "New Plant Diseases Dataset", "train"),
        os.path.join(dl, "train"),
    ]:
        if os.path.isdir(candidate):
            train_dir = candidate; break
    if train_dir is None:
        for dirpath, dirnames, _ in os.walk(dl):
            dirnames[:] = [d for d in dirnames if not d.startswith('.')]
            if os.path.basename(dirpath) == "train":
                if any("___" in d for d in dirnames):
                    train_dir = dirpath; break

    if train_dir is None:
        print(f"  [ERROR] Could not locate train folder in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
    print(f"  [NewPlantDiseases] train (reference): {train_dir}")

    # ── Locate flat test folder ────────────────────────────────────────────────
    test_dir = None
    for candidate in [
        os.path.join(dl, "New Plant Diseases Dataset", "New Plant Diseases Dataset", "test", "test"),
        os.path.join(dl, "New Plant Diseases Dataset", "test", "test"),
        os.path.join(dl, "test", "test"),
        os.path.join(dl, "test"),
    ]:
        if os.path.isdir(candidate):
            imgs = [f for f in os.listdir(candidate) if f.lower().endswith(IMAGE_EXT)]
            if imgs:
                test_dir = candidate; break
    if test_dir is None:
        for dirpath, dirnames, filenames in os.walk(dl):
            dirnames[:] = [d for d in dirnames if not d.startswith('.')]
            imgs = [f for f in filenames if f.lower().endswith(IMAGE_EXT)]
            if imgs and not dirnames:   # leaf folder with images only
                test_dir = dirpath; break
    print(f"  [NewPlantDiseases] test  (benchmark): {test_dir}")

    # ── Reference DataFrame from train subfolders ─────────────────────────────
    train_data = collect_images_df(train_dir)
    def _norm(lbl):
        if "___" in lbl:
            parts = lbl.split("___", 1)
            return safe_name(f"{parts[0].strip()}_{parts[1].strip()}")
        return safe_name(lbl)
    if len(train_data):
        train_data[1] = train_data[1].apply(_norm)
    classes = sorted(train_data[1].unique().tolist()) if len(train_data) else []

    # ── Benchmark DataFrame from flat test folder (filename prefix matching) ───
    bench_rows = []
    if test_dir and classes:
        cls_lookup = {c.lower().replace("_","").replace(" ",""): c for c in classes}
        for fname in os.listdir(test_dir):
            if not fname.lower().endswith(IMAGE_EXT):
                continue
            stem_norm = os.path.splitext(fname)[0].lower().rstrip("0123456789").replace("_","").replace(" ","")
            matched = next(
                (cls_key for cls_bare, cls_key in cls_lookup.items()
                 if stem_norm == cls_bare or stem_norm.startswith(cls_bare)),
                None
            )
            if matched:
                bench_rows.append({0: os.path.join(test_dir, fname), 1: matched})
    bench_data = pd.DataFrame(bench_rows) if bench_rows else pd.DataFrame(columns=[0,1])
    print(f"  [NewPlantDiseases] {len(classes)} classes | "
          f"{len(train_data)} train imgs | {len(bench_data)} test imgs matched")

    if not classes:
        print(f"  [ERROR] No classes found")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    # Drop classes that cannot meet quota in EITHER pool
    _viable = []
    for _c in classes:
        _rc = len(train_data[train_data[1] == _c]) if len(train_data) else 0
        _bc = len(bench_data[bench_data[1] == _c]) if len(bench_data) else 0
        if _rc >= n_ref and _bc >= n_bench:
            _viable.append(_c)
        else:
            print(f'  [DROP CLASS] {_c!r}: ref={_rc}/{n_ref}, bench={_bc}/{n_bench} -- excluded')
    classes = _viable
    if not classes:
        print(f'  [DROP ALL] No classes meet quota for both pools in {name}')
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    ref_df   = sample_per_class(train_data, classes, n_ref)
    bench_df = sample_per_class(bench_data, classes, n_bench) if len(bench_data) else pd.DataFrame(columns=[0,1])
    ref_df   = save_split(ref_df,   name, "reference")
    bench_df = save_split(bench_df, name, "benchmark")

    n_r = sum(1 for p in ref_df[0]   if p)
    n_b = sum(1 for p in bench_df[0] if p)
    print(f"  [SAVED] {name}: reference={n_r}  benchmark={n_b}")
    if os.path.exists(dl):
        shutil.rmtree(dl, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {dl}")
    return name, ref_df, bench_df, classes, desc


def load_PlantDoc(n_ref, n_bench):
    """
    PlantDoc Dataset — GitHub: pratikkayal/PlantDoc-Dataset
    https://github.com/pratikkayal/PlantDoc-Dataset.git

    Structure:
      PlantDoc-Dataset/
        train/  <Class Name with spaces>/  <- REFERENCE pool
        test/   <Class Name with spaces>/  <- BENCHMARK pool

    Class names e.g. "Apple Scab Leaf" — first word = crop, rest = disease.
    Train -> reference, test -> benchmark. Cloned via git.
    """
    name = "PlantDoc"
    desc = ("PlantDoc Dataset (GitHub: pratikkayal). 27 classes across 13 crop "
            "species. Real-world plant disease images with natural lighting. "
            "Train split for reference, test split for benchmark.")

    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, [])
        src_tag = safe_name(name)[:12]
        saved_classes = []
        for crop_d in (os.listdir(REF_DIR) if os.path.isdir(REF_DIR) else []):
            cp = os.path.join(REF_DIR, crop_d)
            if not os.path.isdir(cp): continue
            for dis_d in os.listdir(cp):
                dp = os.path.join(cp, dis_d)
                if not os.path.isdir(dp): continue
                if any(src_tag in f for f in os.listdir(dp)
                       if f.lower().endswith(IMAGE_EXT)):
                    saved_classes.append(f"{crop_d}__{dis_d}")
        return name, ref, bench, sorted(saved_classes), desc

    dl = os.path.join(DATA_ROOT, "PlantDoc-Dataset")
    if not os.path.exists(dl):
        print(f"  [git] Cloning PlantDoc-Dataset ...")
        ret = os.system(
            f'git clone --depth=1 '
            f'https://github.com/pratikkayal/PlantDoc-Dataset.git "{dl}"')
        if ret != 0 or not os.path.isdir(dl):
            print(f"  [ERROR] git clone failed (exit {ret}). "
                  "Check internet access and that git is installed.")
            return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
        print(f"  [git] Cloned to {dl}")

    def _find_split(split_name):
        for candidate in [
            os.path.join(dl, split_name),
            os.path.join(dl, "PlantDoc-Dataset", split_name),
        ]:
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))
                        and not d.startswith(".")]
                if subs:
                    return candidate
        for dirpath, dirnames, _ in os.walk(dl):
            dirnames[:] = [d for d in dirnames if not d.startswith(".")]
            if os.path.basename(dirpath).lower() == split_name.lower():
                subs = [d for d in os.listdir(dirpath)
                        if os.path.isdir(os.path.join(dirpath, d))]
                if subs:
                    return dirpath
        return None

    train_dir = _find_split("train")
    test_dir  = _find_split("test")

    if train_dir is None and test_dir is None:
        print(f"  [ERROR] Could not locate PlantDoc train/test folders in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    print(f"  [PlantDoc] train (reference): {train_dir}")
    print(f"  [PlantDoc] test  (benchmark): {test_dir}")

    train_data = collect_images_df(train_dir) if train_dir else pd.DataFrame(columns=[0,1])
    test_data  = collect_images_df(test_dir)  if test_dir  else pd.DataFrame(columns=[0,1])

    actual_classes = sorted(set(
        (train_data[1].unique().tolist() if len(train_data) else []) +
        (test_data[1].unique().tolist()  if len(test_data)  else [])
    ))

    if not actual_classes:
        print(f"  [ERROR] No classes found in PlantDoc")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    print(f"  [PlantDoc] {len(actual_classes)} classes discovered")

    # Drop classes that cannot meet quota in EITHER pool
    _viable = []
    for _c in actual_classes:
        _rc = len(train_data[train_data[1] == _c]) if len(train_data) else 0
        _bc = len(test_data[test_data[1] == _c]) if len(test_data) else 0
        if _rc >= n_ref and _bc >= n_bench:
            _viable.append(_c)
        else:
            print(f'  [DROP CLASS] {_c!r}: ref={_rc}/{n_ref}, bench={_bc}/{n_bench} -- excluded')
    actual_classes = _viable
    if not actual_classes:
        print(f'  [DROP ALL] No classes meet quota for both pools in {name}')
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    ref_df   = sample_per_class(train_data, actual_classes, n_ref)
    bench_df = sample_per_class(test_data,  actual_classes, n_bench)
    ref_df   = save_split(ref_df,   name, "reference")
    bench_df = save_split(bench_df, name, "benchmark")

    n_r = sum(1 for p in ref_df[0]   if p)
    n_b = sum(1 for p in bench_df[0] if p)
    print(f"  [SAVED] {name}: reference={n_r}  benchmark={n_b}")
    if os.path.exists(dl):
        shutil.rmtree(dl, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {dl}")
    return name, ref_df, bench_df, actual_classes, desc



def load_PlantDiseasesDataset(n_ref, n_bench):
    """
    Plant Diseases Dataset (abdallahalidev/plantdisease-dataset)
    kagglehub: abdallahalidev/plantdisease-dataset

    Structure:
      PlantDiseasesDataset/
        <Crop>/
          train/  <Crop Disease>/ ...   <- REFERENCE pool
          valid/  <Crop Disease>/ ...   <- BENCHMARK pool

    Top-level folder = crop (Apple, Corn, Grape, ...).
    Class folder names already include the crop e.g. "Apple Black rot".
    -> crop  = top-level dir name
    -> disease = class folder name (with crop prefix stripped for cleanliness)
    train -> reference, valid -> benchmark.
    """
    name = "Plant Diseases Dataset"
    desc = ("Plant Diseases Dataset (kaggle: abdallahalidev). 2.77 GB, multi-crop "
            "dataset organised by crop then disease. Train/valid splits per crop. "
            "Covers Apple, Corn, Grape, Peach, Pepper, Potato, Strawberry, Tomato etc.")

    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, [])
        src_tag = safe_name(name)[:12]
        saved_classes = []
        for crop_d in (os.listdir(REF_DIR) if os.path.isdir(REF_DIR) else []):
            cp = os.path.join(REF_DIR, crop_d)
            if not os.path.isdir(cp): continue
            for dis_d in os.listdir(cp):
                dp = os.path.join(cp, dis_d)
                if not os.path.isdir(dp): continue
                if any(src_tag in f for f in os.listdir(dp)
                       if f.lower().endswith(IMAGE_EXT)):
                    saved_classes.append(f"{crop_d}__{dis_d}")
        return name, ref, bench, sorted(saved_classes), desc

    dl = os.path.join(DATA_ROOT, "plantdisease-dataset")
    if not os.path.exists(dl):
        if KAGGLEHUB_AVAILABLE:
            try:
                print(f"  [kagglehub] Downloading plantdisease-dataset ...")
                dl_path = kagglehub.dataset_download(
                    "abdallahalidev/plantdisease-dataset")
                print(f"  Path to dataset files: {dl_path}")
                os.makedirs(dl, exist_ok=True)
                shutil.copytree(dl_path, dl, dirs_exist_ok=True)
            except Exception as e:
                print(f"  [kagglehub ERROR] {e} — trying kaggle API fallback")
                kaggle_download("abdallahalidev/plantdisease-dataset", dl)
        else:
            kaggle_download("abdallahalidev/plantdisease-dataset", dl)

    # ── Locate the root that has <Crop>/train/ substructure ──────────────────
    dataset_root = None
    for candidate in [
        os.path.join(dl, "PlantDiseasesDataset"),
        dl,
    ]:
        if os.path.isdir(candidate):
            # Check if any subdir has train/ or valid/ inside it
            for sub in os.listdir(candidate):
                sub_path = os.path.join(candidate, sub)
                if os.path.isdir(os.path.join(sub_path, "train")) or                    os.path.isdir(os.path.join(sub_path, "valid")):
                    dataset_root = candidate
                    break
        if dataset_root:
            break

    if dataset_root is None:
        # Walk fallback — find a folder that has crop subdirs with train/valid
        for dirpath, dirnames, _ in os.walk(dl):
            dirnames[:] = [d for d in dirnames if not d.startswith(".")]
            for d in dirnames:
                crop_path = os.path.join(dirpath, d)
                if os.path.isdir(os.path.join(crop_path, "train")) or                    os.path.isdir(os.path.join(crop_path, "valid")):
                    dataset_root = dirpath
                    break
            if dataset_root:
                break

    if dataset_root is None:
        print(f"  [ERROR] Could not locate PlantDiseasesDataset root in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    print(f"  [PlantDiseasesDataset] Root: {dataset_root}")

    # ── Walk crop dirs, collect train and valid images ────────────────────────
    ref_rows, bench_rows = [], []
    crops_found = []

    for crop_dir in sorted(os.listdir(dataset_root)):
        crop_path = os.path.join(dataset_root, crop_dir)
        if not os.path.isdir(crop_path) or crop_dir.startswith("."):
            continue

        train_path = os.path.join(crop_path, "train")
        valid_path = os.path.join(crop_path, "valid")

        for split_path, rows in [(train_path, ref_rows),
                                  (valid_path, bench_rows)]:
            if not os.path.isdir(split_path):
                continue
            for cls_dir in sorted(os.listdir(split_path)):
                cls_path = os.path.join(split_path, cls_dir)
                if not os.path.isdir(cls_path) or cls_dir.startswith("."):
                    continue
                # Encode as "<Crop>_<ClassName>" so _parse_crop_disease_from_label
                # correctly splits crop from disease on first underscore.
                # e.g. Apple dir + "Cedar apple rust" -> "Apple_Cedar_apple_rust"
#                   -> crop="Apple", disease="Cedar Apple Rust"  checkmark
                cls_key = safe_name(crop_dir) + "_" + safe_name(cls_dir)
                for f in os.listdir(cls_path):
                    if f.lower().endswith(IMAGE_EXT):
                        rows.append({0: os.path.join(cls_path, f), 1: cls_key})

        crops_found.append(crop_dir)

    print(f"  [PlantDiseasesDataset] {len(crops_found)} crops: {crops_found}")

    ref_data   = pd.DataFrame(ref_rows)   if ref_rows   else pd.DataFrame(columns=[0,1])
    bench_data = pd.DataFrame(bench_rows) if bench_rows else pd.DataFrame(columns=[0,1])

    if len(ref_data) == 0:
        print(f"  [ERROR] No reference images found")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    all_classes = sorted(set(
        ref_data[1].unique().tolist() +
        (bench_data[1].unique().tolist() if len(bench_data) else [])
    ))
    print(f"  [PlantDiseasesDataset] {len(all_classes)} classes total")

    # Drop classes that cannot meet quota in EITHER pool
    _viable = []
    for _c in all_classes:
        _rc = len(ref_data[ref_data[1] == _c]) if len(ref_data) else 0
        _bc = len(bench_data[bench_data[1] == _c]) if len(bench_data) else 0
        if _rc >= n_ref and _bc >= n_bench:
            _viable.append(_c)
        else:
            print(f'  [DROP CLASS] {_c!r}: ref={_rc}/{n_ref}, bench={_bc}/{n_bench} -- excluded')
    all_classes = _viable
    if not all_classes:
        print(f'  [DROP ALL] No classes meet quota for both pools in {name}')
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    ref_df   = sample_per_class(ref_data,   all_classes, n_ref)
    bench_df = sample_per_class(bench_data, all_classes, n_bench)                if len(bench_data) else pd.DataFrame(columns=[0,1])

    ref_df   = save_split(ref_df,   name, "reference")
    bench_df = save_split(bench_df, name, "benchmark")

    n_r = sum(1 for p in ref_df[0]   if p)
    n_b = sum(1 for p in bench_df[0] if p)
    print(f"  [SAVED] {name}: reference={n_r}  benchmark={n_b}")

    if os.path.exists(dl):
        shutil.rmtree(dl, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {dl}")

    return name, ref_df, bench_df, all_classes, desc


def load_FUSARIUM22(n_ref, n_bench):
    name    = "FUSARIUM 22"
    rmap    = {'1(HR)':'Highly Resistant','9(HS)':'Highly Susceptible',
               '5(MR)':'Moderately Resistant','3(R)':'Resistant','7(S)':'Susceptible'}
    classes = list(rmap.values())
    desc    = ("Fusarium Wilt disease severity in chickpea. Classes range from Highly "
               "Resistant to Highly Susceptible — ideal for ordinal severity tasks.")
    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc
    dl   = os.path.join(DATA_ROOT, "fusarium22")
    base = os.path.join(dl, "FUSARIUM-22", "dataset_raw")
    if not os.path.exists(dl):
        kaggle_download("tolgahayit/fusarium-wilt-disease-in-chickpea-dataset", dl)
    rename_folders_dict(base, rmap)
    data = collect_images_df(base)
    ref, bench = split_save_cleanup(data, classes, name, dl, n_ref, n_bench)
    return name, ref, bench, classes, desc

# ═══════════════════════════════════════════════════════════════════════════
#  LEAFNET LOADER  (HuggingFace streaming — enalis/LeafNet)
# ═══════════════════════════════════════════════════════════════════════════

def _extract_crop_disease(caption):
    """Parse crop + disease label from a LeafNet caption string."""
    c  = str(caption).strip()
    cl = c.lower()

    # "a image of <crop> healthy leaves"
    m = re.search(r'a image of ([a-z]+) healthy leaves?', cl)
    if m:
        return m.group(1).capitalize(), "Healthy"

    # "a image of <crop> leaves diseased by <Disease> [with ...]"
    m = re.search(r'a image of ([a-z]+) leaves? diseased by ([a-z][a-z\s]*?)(?:\s+with|\s+disease|$)', cl)
    if m:
        return m.group(1).capitalize(), m.group(2).strip().title()

    # "a image of <crop> leaves with <Disease> [with symptoms ...]"
    m = re.search(r'a image of ([a-z]+) leaves? with ([a-z][a-z\s]*?)(?:\s+with|\s+symptoms|$)', cl)
    if m:
        disease = m.group(2).strip().title()
        if len(disease.split()) <= 5:
            return m.group(1).capitalize(), disease

    # Fallback: "<crop> leaves diseased by <Disease>"
    m = re.search(r'([a-z]+) leaves? diseased by ([a-z][a-z\s]*?)(?:\s+with|$)', cl)
    if m:
        return m.group(1).capitalize(), m.group(2).strip().title()

    return None


def _save_pil_image(pil_img, dest_path):
    """Convert a PIL image (any mode) to JPEG and save to dest_path."""
    try:
        img = pil_img.convert("RGB").resize((224, 224), Image.BILINEAR)
        img.save(dest_path, "JPEG", quality=85)
        return True
    except Exception as e:
        print(f"  [IMG SAVE] {e}")
        return False


def load_LeafNet(n_ref, n_bench):
    """
    Stream enalis/LeafNet from HuggingFace.
    Downloads all images to a temp folder on disk (data/leafnet-raw/<ClassName>/)
    then passes through the standard collect_images_df -> split_save_cleanup
    pipeline — identical to every other loader.

    Requires:  pip install datasets
    Optional:  set HUGGINGFACE_HUB_TOKEN env var for private/gated access.
    """
    name = "LeafNet"
    desc = ("LeafNet — HuggingFace dataset (enalis/LeafNet). "
            "Multi-crop leaf disease classification derived from natural-language "
            "captions. Classes are (Crop, Disease) pairs spanning diverse species "
            "and conditions including Healthy controls.")

    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, [])
        src_tag = safe_name(name)[:12]
        saved_classes = []
        for crop_d in (os.listdir(REF_DIR) if os.path.isdir(REF_DIR) else []):
            cp = os.path.join(REF_DIR, crop_d)
            if not os.path.isdir(cp): continue
            for dis_d in os.listdir(cp):
                dp = os.path.join(cp, dis_d)
                if not os.path.isdir(dp): continue
                if any(src_tag in f for f in os.listdir(dp)
                       if f.lower().endswith(IMAGE_EXT)):
                    saved_classes.append(f"{crop_d}__{dis_d}")
        return name, ref, bench, sorted(saved_classes), desc

    if not HF_AVAILABLE:
        print("  [SKIP] LeafNet — 'datasets' package not installed. "
              "Run:  pip install datasets")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    # Temp staging folder — deleted by split_save_cleanup after sampling
    dl = os.path.join(DATA_ROOT, "leafnet-raw")
    # Wipe any previous staging run so class folders start clean
    if os.path.isdir(dl):
        shutil.rmtree(dl, ignore_errors=True)
    os.makedirs(dl, exist_ok=True)

    # Wipe stale curated output (wrong structure from previous runs)
    for stale_base in [os.path.join(REF_DIR, "LeafNet"),
                       os.path.join(BENCH_DIR, "LeafNet")]:
        if os.path.isdir(stale_base):
            shutil.rmtree(stale_base, ignore_errors=True)
            print(f"  [CLEANUP] Removed stale LeafNet output: {stale_base}")

    print(f"  Streaming LeafNet from HuggingFace -> {dl}")
    t0 = time.time()

    if "HUGGINGFACE_HUB_TOKEN" in os.environ:
        os.environ["HF_TOKEN"] = os.environ["HUGGINGFACE_HUB_TOKEN"]

    try:
        ds = hf_load_dataset("enalis/LeafNet", split="train", streaming=True)
    except Exception as e:
        print(f"  [ERROR] Could not load LeafNet: {e}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    needed  = n_ref + n_bench
    counters = {}   # cls_label -> count of images saved so far
    scanned  = 0

    def _save_row(row):
        """Parse caption, save image to dl/<cls_label>/<n>.jpg. Returns cls_label or None."""
        result = _extract_crop_disease(row.get("caption", ""))
        if not result:
            return None
        crop, disease = result
        # Use safe_name immediately so staging folder name == final saved folder name.
        # Format: "Crop__Disease"  (double-underscore separator, spaces -> underscores)
        cls_label = safe_name(f"{crop}_{disease}")

        # Single flat folder directly inside dl/ — no subdirectories
        cls_dir = os.path.join(dl, cls_label)
        os.makedirs(cls_dir, exist_ok=True)

        try:
            pil = row["image"]
            if isinstance(pil, np.ndarray):
                pil = Image.fromarray(pil)
            count = len([f for f in os.listdir(cls_dir) if f.lower().endswith(IMAGE_EXT)])
            if count >= needed:
                return cls_label   # already have enough for this class
            dest = os.path.join(cls_dir, f"{count+1:05d}.jpg")
            pil.convert("RGB").save(dest, "JPEG", quality=85)
        except Exception:
            return None
        return cls_label

    lock = threading.Lock()
    new_classes_seen = set()

    with ThreadPoolExecutor(max_workers=NUM_WORKERS) as ex:
        batch = []
        BATCH = NUM_WORKERS * 4

        for row in ds:
            scanned += 1
            batch.append(row)

            if len(batch) >= BATCH:
                for cls_label in ex.map(_save_row, batch):
                    if cls_label and cls_label not in new_classes_seen:
                        with lock:
                            new_classes_seen.add(cls_label)
                            print(f"  [{scanned:>8,}]  New class {len(new_classes_seen):>3}: "
                                  f"{cls_label}  ({time.time()-t0:.0f}s)")
                batch = []

            if scanned % 20_000 == 0:
                print(f"  [{scanned:>8,}]  {len(new_classes_seen)} classes found  "
                      f"({time.time()-t0:.0f}s)")

        # Flush remaining
        for cls_label in ex.map(_save_row, batch):
            if cls_label and cls_label not in new_classes_seen:
                new_classes_seen.add(cls_label)

    print(f"\n  Stream complete: {len(new_classes_seen)} classes, "
          f"{scanned:,} rows in {time.time()-t0:.1f}s")

    # ── Standard pipeline from here — identical to all other loaders ──────────
    data    = collect_images_df(dl)   # walks dl/<ClassName>/*.jpg
    classes = sorted(data[1].unique().tolist()) if len(data) else []

    if len(data) == 0:
        print(f"  [ERROR] No images saved to {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    ref, bench = split_save_cleanup(data, classes, name, dl, n_ref, n_bench)
    return name, ref, bench, classes, desc


# ═══════════════════════════════════════════════════════════════════════════
#  LOCAL DATASET LOADER
# ═══════════════════════════════════════════════════════════════════════════

def collect_images_recursive(folder):
    """
    Return ALL image files found anywhere under `folder` (recursive).
    Used for local datasets where images may be spread across sub-subfolders.
    """
    found = []
    if not os.path.isdir(folder):
        return found
    for dirpath, _, filenames in os.walk(folder):
        for f in filenames:
            if f.lower().endswith(IMAGE_EXT):
                found.append(os.path.join(dirpath, f))
    return found


def build_local_class_map(root):
    """
    Recursively discover all leaf-level disease classes under `root`.

    Strategy
    --------
    The top-level subdirectories of `root` become **categories**
    (e.g. "Corn Diseases", "Soybean Diseases").  Within each category
    every subdirectory — at ANY depth — that directly contains image
    files is treated as a **class**.  The class name is built by joining
    all folder-name segments below the category root with '__', so nested
    paths become unique, flat keys:

        Corn Diseases/Ear rots/Aspergillus ear rot  →  Ear_rots__Aspergillus_ear_rot

    Folders that contain only sub-folders (no images of their own) are
    NOT included as classes — only the deepest image-bearing folders are.
    This avoids double-counting parent folders.

    Returns
    -------
    dict  category_name -> { class_key: absolute_path }
    """
    if not os.path.isdir(root):
        print(f"  [ERROR] LOCAL_SOURCE_ROOT does not exist: {root}")
        return {}

    class_map = {}

    for cat_dir in sorted(os.listdir(root)):
        cat_path = os.path.join(root, cat_dir)
        if not os.path.isdir(cat_path) or cat_dir.startswith('.'):
            continue

        cat_key = safe_name(cat_dir)   # e.g. "Corn_Diseases"
        classes = {}

        for dirpath, dirnames, filenames in os.walk(cat_path):
            # Skip hidden dirs
            dirnames[:] = [d for d in dirnames if not d.startswith('.')]

            # Check if this folder contains images directly
            imgs_here = [f for f in filenames if f.lower().endswith(IMAGE_EXT)]
            if not imgs_here:
                continue  # no images here — keep walking deeper

            # Build class key: relative path from cat_path, separators → '__'
            rel = os.path.relpath(dirpath, cat_path)
            if rel == '.':
                # Images sitting directly in the category root — skip,
                # they don't belong to a named class folder
                continue
            cls_key = safe_name(rel.replace(os.sep, '__'))
            classes[cls_key] = dirpath

        if classes:
            class_map[cat_key] = classes
            print(f"  [LOCAL] {cat_key}: {len(classes)} class folders discovered")
        else:
            print(f"  [LOCAL] {cat_key}: no image-bearing subfolders found — skipped")

    return class_map


def load_local_category(category_name, cls_map, n_ref, n_bench, source_root):
    """
    Process one auto-discovered local category.
    Images are collected recursively from each class folder.
    Returns (name, ref_df, bench_df, classes, desc).
    """
    name = category_name
    desc = (f"Local disease dataset — {category_name.replace('_', ' ')}. "
            f"Source: {source_root}")

    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        classes = list(cls_map.keys())
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc

    total_needed = n_ref + n_bench
    rows, included = [], []

    for cls_name, src_folder in sorted(cls_map.items()):
        # Collect recursively — images may be spread across sub-subfolders
        imgs = collect_images_recursive(src_folder)
        if len(imgs) < total_needed:
            print(f"  [SKIP CLASS] {cls_name}: {len(imgs)} imgs (need {total_needed})")
            continue
        included.append(cls_name)
        for p in imgs:
            rows.append({0: p, 1: cls_name})

    if not rows:
        print(f"  [WARN] No valid classes in {category_name}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    data = pd.DataFrame(rows)
    # Local data: never delete source files
    ref_df   = sample_per_class(data, included, n_ref)
    used     = set(ref_df[0].tolist())
    rem      = data[~data[0].isin(used)]
    bench_df = sample_per_class(rem, included, n_bench)

    ref_df   = save_split(ref_df,   name, "reference")
    bench_df = save_split(bench_df, name, "benchmark")

    n_r = sum(1 for p in ref_df[0]   if p)
    n_b = sum(1 for p in bench_df[0] if p)
    print(f"  [SAVED] {name}: reference={n_r}  benchmark={n_b} "
          f"-> samples/{safe_name(name)}/")
    return name, ref_df, bench_df, included, desc
# ═══════════════════════════════════════════════════════════════════════════
#  PDF REPORT  — formal, Times New Roman, grouped by Crop → Disease
# ═══════════════════════════════════════════════════════════════════════════

# ── Palette: black / white / light greys only ────────────────────────────────
C_BLACK      = colors.HexColor("#000000")
C_DARK       = colors.HexColor("#1a1a1a")
C_MID        = colors.HexColor("#555555")
C_LIGHT_GREY = colors.HexColor("#dddddd")
C_ROW_ALT    = colors.HexColor("#f5f5f5")
C_WHITE      = colors.white
C_HDR_BG     = colors.HexColor("#222222")   # near-black header band

TN   = "Times-Roman"
TNB  = "Times-Bold"
TNI  = "Times-Italic"
TNBI = "Times-BoldItalic"


def build_styles():
    base = getSampleStyleSheet()
    return {
        "ReportTitle": ParagraphStyle("ReportTitle", parent=base["Title"],
            fontSize=22, leading=28, spaceAfter=4, alignment=TA_CENTER,
            textColor=C_BLACK, fontName=TNB),
        "ReportSubtitle": ParagraphStyle("ReportSubtitle", parent=base["Normal"],
            fontSize=10, leading=14, spaceAfter=4, alignment=TA_CENTER,
            textColor=C_MID, fontName=TNI),
        "CropHeading": ParagraphStyle("CropHeading", parent=base["Normal"],
            fontSize=13, leading=17, spaceBefore=10, spaceAfter=2,
            textColor=C_BLACK, fontName=TNB),
        "DatasetTitle": ParagraphStyle("DatasetTitle", parent=base["Heading1"],
            fontSize=12, leading=16, textColor=C_WHITE,
            fontName=TNB, alignment=TA_LEFT),
        "DatasetDesc": ParagraphStyle("DatasetDesc", parent=base["Normal"],
            fontSize=9, leading=13, spaceAfter=4,
            textColor=C_DARK, fontName=TN),
        "ClassLabel": ParagraphStyle("ClassLabel", parent=base["Normal"],
            fontSize=6, leading=8, spaceAfter=0, alignment=TA_CENTER,
            textColor=C_MID, fontName=TN),
        "SplitLabel": ParagraphStyle("SplitLabel", parent=base["Normal"],
            fontSize=8, leading=11, alignment=TA_LEFT,
            textColor=C_DARK, fontName=TNB),
        "DiseaseLabel": ParagraphStyle("DiseaseLabel", parent=base["Normal"],
            fontSize=9, leading=12, spaceBefore=4, spaceAfter=1,
            textColor=C_BLACK, fontName=TNB),
        "SectionLabel": ParagraphStyle("SectionLabel", parent=base["Normal"],
            fontSize=9, leading=11, spaceBefore=3, spaceAfter=2,
            textColor=C_BLACK, fontName=TNB),
        "InfoCellBold": ParagraphStyle("InfoCellBold", parent=base["Normal"],
            fontSize=8, leading=10, textColor=C_DARK, fontName=TNB),
        "InfoCell": ParagraphStyle("InfoCell", parent=base["Normal"],
            fontSize=8, leading=10, textColor=C_DARK, fontName=TN),
        "TOCEntry": ParagraphStyle("TOCEntry", parent=base["Normal"],
            fontSize=9, leading=13, leftIndent=8,
            textColor=C_DARK, fontName=TN),
        "NoDataMsg": ParagraphStyle("NoDataMsg", parent=base["Normal"],
            fontSize=9, leading=12, spaceAfter=8, alignment=TA_CENTER,
            textColor=C_MID, fontName=TNI),
        "FooterStyle": ParagraphStyle("FooterStyle", parent=base["Normal"],
            fontSize=7, leading=9, alignment=TA_CENTER,
            textColor=C_MID, fontName=TNI),
    }


def thumb(img_path, w=130, h=130):
    if not img_path or not os.path.isfile(img_path): return None
    try:
        with Image.open(img_path) as im:
            im.verify()          # catch truncated / broken files early
    except Exception:
        return None              # silently skip broken images
    try:
        with Image.open(img_path) as im:   # must re-open after verify()
            im = im.convert("RGB")
            im.thumbnail((w, h), Image.LANCZOS)
            buf = BytesIO()
            im.save(buf, "JPEG", quality=80)
            if buf.tell() == 0:
                return None      # empty buffer — something went wrong
            buf.seek(0)
            return buf
    except Exception as e:
        return None              # silently skip any other decode errors


def _thin_rule(W):
    return HRFlowable(width=W, thickness=0.5, color=C_LIGHT_GREY, spaceAfter=3)


def make_banner(title, width, sty):
    """Near-black banner with white Times-Bold title — no colour."""
    tbl = Table([[Paragraph(title, sty["DatasetTitle"])]],
                colWidths=[width], rowHeights=[22])
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), C_HDR_BG),
        ("LEFTPADDING",   (0,0),(-1,-1), 10),
        ("RIGHTPADDING",  (0,0),(-1,-1), 10),
        ("TOPPADDING",    (0,0),(-1,-1), 3),
        ("BOTTOMPADDING", (0,0),(-1,-1), 3),
    ]))
    return tbl


def make_summary_table(all_datasets, sty, n_ref, n_bench):
    TYPE = {
        "SBRD":"Disease + Severity",    "Mango Leaf":"Disease ID",
        "Soybean PNAS":"Disease ID",    "Bean Leaf":"Disease ID",
        "Yellow Rust":"Severity/Resistance",
        "FUSARIUM 22":"Severity/Resistance",
        "Banana Leaf":"Disease ID",     "Cauliflower":"Disease ID",
        "Lettuce":"Disease ID",         "LeafNet":"Disease ID (HuggingFace)",
        "Alfalfa_Diseases":"Local",     "Corn_Diseases":"Local",
        "Soybean_Diseases":"Local",     "Wheat_Diseases":"Local",
        "Mango_Leaf_Disease":"Local",
    }
    rows = [["#", "Dataset", "Classes", "Ref / cls", "Bench / cls", "Task Type"]]
    for i, (nm, ref, bench, cls, _) in enumerate(all_datasets, 1):
        ok = hasattr(ref, '__len__') and len(ref) > 0
        rows.append([str(i), nm, str(len(cls)),
                     str(n_ref) if ok else "N/A",
                     str(n_bench) if ok else "N/A",
                     TYPE.get(nm, "Disease")])
    cw = [0.7*cm, 4.2*cm, 1.6*cm, 1.8*cm, 2.2*cm, 3.8*cm]
    t  = Table(rows, colWidths=cw, repeatRows=1)
    t.setStyle(TableStyle([
        # Header
        ("BACKGROUND",    (0,0),(-1,0), C_HDR_BG),
        ("TEXTCOLOR",     (0,0),(-1,0), C_WHITE),
        ("FONTNAME",      (0,0),(-1,0), TNB),
        ("FONTSIZE",      (0,0),(-1,0), 9),
        ("ALIGN",         (0,0),(-1,0), "CENTER"),
        ("TOPPADDING",    (0,0),(-1,0), 5),
        ("BOTTOMPADDING", (0,0),(-1,0), 5),
        # Body
        ("FONTNAME",      (0,1),(-1,-1), TN),
        ("FONTSIZE",      (0,1),(-1,-1), 8),
        ("TOPPADDING",    (0,1),(-1,-1), 3),
        ("BOTTOMPADDING", (0,1),(-1,-1), 3),
        ("ALIGN",         (0,1),(0,-1),  "CENTER"),
        ("ALIGN",         (2,1),(4,-1),  "CENTER"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [C_WHITE, C_ROW_ALT]),
        ("GRID",          (0,0),(-1,-1), 0.4, C_LIGHT_GREY),
    ]))
    return t


def _build_grouped_structure(all_datasets):
    """
    Returns an OrderedDict:
      crop -> { disease -> [(ds_name, ref_paths, bench_paths)] }
    Sorted crop A-Z, disease A-Z within each crop.
    """
    from collections import defaultdict
    grouped = defaultdict(lambda: defaultdict(list))

    for ds_name, ref_df, bench_df, classes, _ in all_datasets:
        if not classes:
            continue
        ref_map   = {}
        bench_map = {}
        if hasattr(ref_df, '__len__') and len(ref_df) > 0:
            for cls_label, grp in ref_df.groupby(1):
                ref_map[cls_label] = grp[0].tolist()
        if hasattr(bench_df, '__len__') and len(bench_df) > 0:
            for cls_label, grp in bench_df.groupby(1):
                bench_map[cls_label] = grp[0].tolist()

        for cls in classes:
            crop, disease = _parse_crop_disease_from_label(ds_name, cls)
            ref_paths   = ref_map.get(cls, [])[:PDF_DISPLAY_MAX]
            bench_paths = bench_map.get(cls, [])[:PDF_DISPLAY_MAX]
            grouped[crop][disease].append((ds_name, ref_paths, bench_paths))

    # Sort
    return {
        crop: dict(sorted(diseases.items()))
        for crop, diseases in sorted(grouped.items())
    }


def _image_row(paths, label, W, sty):
    """5-column image row with filenames below each thumbnail."""
    PDF_COLS = 5
    IW = (W - 0.4*cm) / PDF_COLS
    IH = IW
    LH = 0.7*cm
    cells = []
    for p in (paths or [])[:5]:
        buf   = thumb(p, int(IW), int(IH))
        fname = os.path.basename(p)[:16] if p else ""
        if buf:
            cells.append([RLImage(buf, width=IW, height=IH),
                          Paragraph(fname, sty["ClassLabel"])])
        else:
            cells.append([Paragraph("(missing)", sty["ClassLabel"])])
    while len(cells) < 5:
        cells.append([Paragraph("", sty["ClassLabel"])])
    CW  = IW + 0.1*cm
    tbl = Table([cells], colWidths=[CW]*5, rowHeights=[IH + LH])
    tbl.setStyle(TableStyle([
        ("ALIGN",         (0,0),(-1,-1), "CENTER"),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0),(-1,-1), 2),
        ("BOTTOMPADDING", (0,0),(-1,-1), 2),
        ("LEFTPADDING",   (0,0),(-1,-1), 2),
        ("RIGHTPADDING",  (0,0),(-1,-1), 2),
    ]))
    return tbl


PDF_DISPLAY_MAX = 5


def page_footer(canvas_obj, doc):
    canvas_obj.saveState()
    canvas_obj.setFont(TNI, 7)
    canvas_obj.setFillColor(C_MID)
    canvas_obj.drawCentredString(PAGE_W/2, 0.55*cm,
        f"Disease Dataset Report  \u2022  Page {doc.page}")
    canvas_obj.restoreState()


def generate_pdf(all_datasets, output_path, n_ref, n_bench):
    n_ref   = min(n_ref,   PDF_DISPLAY_MAX)
    n_bench = min(n_bench, PDF_DISPLAY_MAX)

    doc = SimpleDocTemplate(output_path, pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=MARGIN,  bottomMargin=1.4*cm,
        title="Disease Dataset Report")
    W   = PAGE_W - 2*MARGIN
    sty = build_styles()
    story = []

    # ── Cover page ────────────────────────────────────────────────────────────
    story.append(Spacer(1, 2.0*cm))

    # Title box: thin top/bottom rules only, no fill
    ttbl = Table([[Paragraph(
        "Disease Classification &amp; Severity<br/>Dataset Report",
        sty["ReportTitle"])]], colWidths=[W])
    ttbl.setStyle(TableStyle([
        ("TOPPADDING",    (0,0),(-1,-1), 14),
        ("BOTTOMPADDING", (0,0),(-1,-1), 14),
        ("LEFTPADDING",   (0,0),(-1,-1), 0),
        ("RIGHTPADDING",  (0,0),(-1,-1), 0),
        ("LINEABOVE",     (0,0),(-1,0),  1.5, C_BLACK),
        ("LINEBELOW",     (0,0),(-1,-1), 1.5, C_BLACK),
    ]))
    story.append(ttbl)
    story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph(
        f"Reference: {n_ref} images/class  \u2022  "
        f"Benchmark: {n_bench} images/class  \u2022  "
        f"{len(all_datasets)} source datasets",
        sty["ReportSubtitle"]))
    story.append(Spacer(1, 0.8*cm))
    story.append(_thin_rule(W))
    story.append(Spacer(1, 0.5*cm))

    # Dataset list (plain text, no coloured dots)
    story.append(Paragraph("Source Datasets", sty["SectionLabel"]))
    story.append(Spacer(1, 3))
    ds_names = [nm for nm,*_ in all_datasets]
    half = len(ds_names)//2 + len(ds_names)%2
    for i in range(half):
        left  = ds_names[i]
        right = ds_names[i+half] if i+half < len(ds_names) else ""
        row_tbl = Table(
            [[Paragraph(f"\u2022  {left}", sty["TOCEntry"]),
              Paragraph(f"\u2022  {right}" if right else "", sty["TOCEntry"])]],
            colWidths=[W/2, W/2])
        story.append(row_tbl)

    story.append(Spacer(1, 0.7*cm))
    story.append(Paragraph("Dataset Overview", sty["SectionLabel"]))
    story.append(Spacer(1, 4))
    story.append(make_summary_table(all_datasets, sty, n_ref, n_bench))
    story.append(PageBreak())

    # ── Grouped section: Crop → Disease ──────────────────────────────────────
    grouped = _build_grouped_structure(all_datasets)

    # Pre-collect all image paths for parallel thumbnailing
    all_img_paths = []
    for crop, diseases in grouped.items():
        for disease, entries in diseases.items():
            for ds_name, rp, bp in entries:
                all_img_paths.extend(rp + bp)

    print(f"  [PDF] Generating thumbnails for {len(all_img_paths)} images ...")
    thumb_cache = {}
    def _make_thumb(p):
        return p, thumb(p)
    with ThreadPoolExecutor(max_workers=NUM_WORKERS) as ex:
        for p, buf in ex.map(_make_thumb, all_img_paths):
            thumb_cache[p] = buf

    PDF_COLS = 5
    IW = (W - 0.4*cm) / PDF_COLS
    IH = IW
    LH = 0.7*cm
    CW = IW + 0.1*cm

    def _cached_row(paths):
        # Each cell must be a list of Flowables — never a plain string.
        # Empty padding cells use an empty Paragraph so ReportLab is happy.
        EMPTY = Paragraph("", sty["ClassLabel"])
        cells = []
        for p in (paths or [])[:5]:
            buf   = thumb_cache.get(p)
            fname = os.path.basename(p)[:16] if p else ""
            if buf:
                # Valid image: thumbnail + filename label
                cells.append([RLImage(buf, width=IW, height=IH),
                               Paragraph(fname, sty["ClassLabel"])])
            else:
                # Broken / missing image: placeholder text only
                cells.append([Paragraph("(missing)", sty["ClassLabel"])])
        # Pad to exactly 5 columns with empty flowable cells
        while len(cells) < 5:
            cells.append([EMPTY])
        tbl = Table([cells], colWidths=[CW]*5, rowHeights=[IH + LH])
        tbl.setStyle(TableStyle([
            ("ALIGN",         (0,0),(-1,-1), "CENTER"),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
            ("TOPPADDING",    (0,0),(-1,-1), 2),
            ("BOTTOMPADDING", (0,0),(-1,-1), 2),
            ("LEFTPADDING",   (0,0),(-1,-1), 2),
            ("RIGHTPADDING",  (0,0),(-1,-1), 2),
        ]))
        return tbl

    for crop, diseases in grouped.items():
        # ── Crop header ───────────────────────────────────────────────────────
        story.append(make_banner(crop, W, sty))
        story.append(Spacer(1, 5))

        for disease, entries in diseases.items():
            # ── Disease sub-header ────────────────────────────────────────────
            dis_hdr = Table(
                [[Paragraph(disease.replace("_", " "), sty["DiseaseLabel"])]],
                colWidths=[W], rowHeights=[16])
            dis_hdr.setStyle(TableStyle([
                ("BACKGROUND",    (0,0),(-1,-1), C_ROW_ALT),
                ("LEFTPADDING",   (0,0),(-1,-1), 8),
                ("TOPPADDING",    (0,0),(-1,-1), 2),
                ("BOTTOMPADDING", (0,0),(-1,-1), 2),
                ("LINEBELOW",     (0,0),(-1,-1), 0.5, C_LIGHT_GREY),
            ]))
            story.append(dis_hdr)

            for ds_name, ref_paths, bench_paths in entries:
                if not ref_paths and not bench_paths:
                    continue
                # Source dataset label (small, italic)
                story.append(Paragraph(
                    f"&nbsp;&nbsp;&nbsp;Source: <i>{ds_name}</i>",
                    sty["DatasetDesc"]))
                if ref_paths:
                    story.append(Paragraph(
                        f"    Reference  ({len(ref_paths)} images)",
                        sty["SplitLabel"]))
                    story.append(_cached_row(ref_paths))
                if bench_paths:
                    story.append(Paragraph(
                        f"    Benchmark  ({len(bench_paths)} images)",
                        sty["SplitLabel"]))
                    story.append(_cached_row(bench_paths))

            story.append(Spacer(1, 6))

        story.append(PageBreak())

    doc.build(story, onFirstPage=page_footer, onLaterPages=page_footer)
    print(f"\n  PDF -> {os.path.abspath(output_path)}")


# ═══════════════════════════════════════════════════════════════════════════
#  XLSX EXPORT — grouped by Crop → Disease
# ═══════════════════════════════════════════════════════════════════════════


def generate_xlsx(all_datasets, output_path):
    """
    Sheet 1 — "By Crop & Disease":
        Grouped by Crop (merged cell) → Disease (merged cell) →
        one row per source dataset with Ref and Bench counts.

    Sheet 2 — "Summary":
        One row per unique (Crop, Disease) pair, totals across all sources.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [WARN] openpyxl not installed — skipping xlsx export.")
        return

    # ── Build registry ────────────────────────────────────────────────────────
    from collections import defaultdict, OrderedDict

    def _is_internal(ds_name):
        return (ds_name.endswith("_Diseases") or ds_name.endswith("_Disease")
                or ds_name in ("Alfalfa_Diseases","Corn_Diseases",
                               "Soybean_Diseases","Wheat_Diseases","Rye_Diseases"))

    def _display_source(ds_name):
        return "Internal" if _is_internal(ds_name) else ds_name

    def _normalise_crop(crop):
        for suffix in (" Diseases", " Disease"):
            if crop.endswith(suffix):
                return crop[:-len(suffix)].strip()
        return crop

    # crop -> disease -> [{"dataset", "ref", "bench"}]
    grouped  = defaultdict(lambda: defaultdict(list))
    registry = {}

    for ds_name, ref_df, bench_df, classes, _ in all_datasets:
        if not classes:
            continue

        display_src = _display_source(ds_name)
        src_tag     = safe_name(ds_name)[:12]

        # Count images actually on disk per (crop_dir, disease_dir) for this source
        def _disk_counts(base_dir):
            counts = {}  # (crop_dir, disease_dir) -> int
            if not os.path.isdir(base_dir):
                return counts
            for cr in os.listdir(base_dir):
                cp = os.path.join(base_dir, cr)
                if not os.path.isdir(cp): continue
                for di in os.listdir(cp):
                    dp = os.path.join(cp, di)
                    if not os.path.isdir(dp): continue
                    n = sum(1 for f in os.listdir(dp)
                            if f.lower().endswith(IMAGE_EXT) and src_tag in f)
                    if n > 0:
                        counts[(cr, di)] = n
            return counts

        disk_ref   = _disk_counts(REF_DIR)
        disk_bench = _disk_counts(BENCH_DIR)

        # Build a set of (crop, disease) keys that actually have images on disk
        all_keys = set(disk_ref.keys()) | set(disk_bench.keys())
        if not all_keys:
            # Fall back to DataFrame counts (fresh run before disk write completes)
            ref_counts   = ref_df.groupby(1).size().to_dict()   if hasattr(ref_df,   "__len__") and len(ref_df)   > 0 else {}
            bench_counts = bench_df.groupby(1).size().to_dict() if hasattr(bench_df, "__len__") and len(bench_df) > 0 else {}
            for cls in classes:
                crop, disease = _parse_crop_disease_from_label(ds_name, cls)
                crop = _normalise_crop(crop)
                n_r = ref_counts.get(cls, 0)
                n_b = bench_counts.get(cls, 0)
                if n_r == 0 and n_b == 0:
                    continue
                grouped[crop][disease].append({"dataset": display_src, "ref": n_r, "bench": n_b})
                pk = (crop.lower(), disease.lower())
                if pk not in registry:
                    registry[pk] = {"crop": crop, "disease": disease,
                                     "sources": set(), "total_ref": 0, "total_bench": 0}
                registry[pk]["sources"].add(display_src)
                registry[pk]["total_ref"]   += n_r
                registry[pk]["total_bench"] += n_b
        else:
            # Use disk counts — guaranteed accurate for both fresh and skip runs
            for (crop_dir, disease_dir), n_r in disk_ref.items():
                n_b = disk_bench.get((crop_dir, disease_dir), 0)
                if n_r == 0 and n_b == 0:
                    continue
                _raw_crop = crop_dir.replace("_", " ").title()
                for _sfx in (" Diseases", " Disease"):
                    if _raw_crop.endswith(_sfx): _raw_crop = _raw_crop[:-len(_sfx)].strip(); break
                crop = _raw_crop
                disease = disease_dir.replace("_", " ").title()
                grouped[crop][disease].append({"dataset": display_src, "ref": n_r, "bench": n_b})
                pk = (crop.lower(), disease.lower())
                if pk not in registry:
                    registry[pk] = {"crop": crop, "disease": disease,
                                     "sources": set(), "total_ref": 0, "total_bench": 0}
                registry[pk]["sources"].add(display_src)
                registry[pk]["total_ref"]   += n_r
                registry[pk]["total_bench"] += n_b
            # Also add bench-only entries (no ref images for this source in some disease)
            for (crop_dir, disease_dir), n_b in disk_bench.items():
                if (crop_dir, disease_dir) in disk_ref:
                    continue  # already handled above
                n_r = 0
                if n_b == 0:
                    continue
                _raw_crop = crop_dir.replace("_", " ").title()
                for _sfx in (" Diseases", " Disease"):
                    if _raw_crop.endswith(_sfx): _raw_crop = _raw_crop[:-len(_sfx)].strip(); break
                crop = _raw_crop
                disease = disease_dir.replace("_", " ").title()
                grouped[crop][disease].append({"dataset": display_src, "ref": n_r, "bench": n_b})
                pk = (crop.lower(), disease.lower())
                if pk not in registry:
                    registry[pk] = {"crop": crop, "disease": disease,
                                     "sources": set(), "total_ref": 0, "total_bench": 0}
                registry[pk]["sources"].add(display_src)
                registry[pk]["total_ref"]   += n_r
                registry[pk]["total_bench"] += n_b


    grouped_sorted = {
        crop: dict(sorted(diseases.items()))
        for crop, diseases in sorted(grouped.items())
    }
    sorted_pairs = sorted(registry.items(), key=lambda x: x[0])

    # ── Shared styles ─────────────────────────────────────────────────────────
    TNR = "Times New Roman"

    HDR_FILL    = PatternFill("solid", start_color="1A1A1A")   # near-black
    HDR_FONT    = Font(name=TNR, bold=True, color="FFFFFF", size=11)
    HDR_ALIGN   = Alignment(horizontal="center", vertical="center", wrap_text=True)

    CROP_FILL   = PatternFill("solid", start_color="333333")   # dark grey
    CROP_FONT   = Font(name=TNR, bold=True, color="FFFFFF", size=11)

    DIS_FILL    = PatternFill("solid", start_color="AAAAAA")   # mid grey
    DIS_FONT    = Font(name=TNR, bold=True, color="000000", size=10)

    BODY_FONT   = Font(name=TNR, size=10)
    BODY_ALIGN  = Alignment(vertical="center")
    ALT_FILL_A  = PatternFill("solid", start_color="F5F5F5")
    ALT_FILL_B  = PatternFill("solid", start_color="FFFFFF")

    TOTAL_FILL  = PatternFill("solid", start_color="1A1A1A")
    TOTAL_FONT  = Font(name=TNR, bold=True, color="FFFFFF", size=10)

    thin   = Side(style="thin",   color="CCCCCC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    WRAP_ALIGN  = Alignment(vertical="center", wrap_text=True)
    CTR_ALIGN   = Alignment(horizontal="center", vertical="center")

    def _hdr(ws, headers, col_widths):
        ws.append(headers)
        for col, (cell, w) in enumerate(zip(ws[1], col_widths), start=1):
            cell.font      = HDR_FONT
            cell.fill      = HDR_FILL
            cell.alignment = HDR_ALIGN
            cell.border    = BORDER
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

    def _style(ws, row_idx, n_cols, fill, font, align=None):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill      = fill
            cell.font      = font
            cell.border    = BORDER
            cell.alignment = align or BODY_ALIGN

    wb = Workbook()

    # ════════════════════════════════════════════════════════════════════════
    #  Sheet 1 — By Crop & Disease
    # ════════════════════════════════════════════════════════════════════════
    ws1       = wb.active
    ws1.title = "By Crop & Disease"
    headers1  = ["Crop", "Disease", "Source Dataset", "Ref Images", "Bench Images"]
    widths1   = [22, 30, 28, 14, 15]
    _hdr(ws1, headers1, widths1)

    row_idx  = 2
    alt_body = 0

    for crop, diseases in grouped_sorted.items():
        # Count total rows for this crop (for merging)
        crop_row_start = row_idx
        crop_row_count = sum(len(entries) for entries in diseases.values())

        for disease, entries in diseases.items():
            dis_row_start = row_idx
            for entry in entries:
                alt_body += 1
                fill = ALT_FILL_A if alt_body % 2 == 0 else ALT_FILL_B
                ws1.append(["", "", entry["dataset"], entry["ref"], entry["bench"]])
                _style(ws1, row_idx, 5, fill, BODY_FONT)
                ws1.cell(row_idx, 3).alignment = BODY_ALIGN
                ws1.cell(row_idx, 4).alignment = CTR_ALIGN
                ws1.cell(row_idx, 5).alignment = CTR_ALIGN
                ws1.row_dimensions[row_idx].height = 16
                row_idx += 1

            # Disease merged cell (col 2, rows dis_row_start..row_idx-1)
            dis_end = row_idx - 1
            if dis_row_start == dis_end:
                ws1.cell(dis_row_start, 2).value = disease
            else:
                ws1.merge_cells(start_row=dis_row_start, start_column=2,
                                end_row=dis_end,        end_column=2)
                ws1.cell(dis_row_start, 2).value = disease
            for r in range(dis_row_start, row_idx):
                c = ws1.cell(r, 2)
                c.fill      = DIS_FILL
                c.font      = DIS_FONT
                c.border    = BORDER
                c.alignment = Alignment(horizontal="left", vertical="center",
                                        wrap_text=True)

        # Crop merged cell (col 1, rows crop_row_start..row_idx-1)
        crop_end = row_idx - 1
        if crop_row_start == crop_end:
            ws1.cell(crop_row_start, 1).value = crop
        else:
            ws1.merge_cells(start_row=crop_row_start, start_column=1,
                            end_row=crop_end,         end_column=1)
            ws1.cell(crop_row_start, 1).value = crop
        for r in range(crop_row_start, row_idx):
            c = ws1.cell(r, 1)
            c.fill      = CROP_FILL
            c.font      = CROP_FONT
            c.border    = BORDER
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)

    # Totals row
    n_data = row_idx - 2
    ws1.append(["TOTAL", "", "", f"=SUM(D2:D{row_idx-1})", f"=SUM(E2:E{row_idx-1})"])
    _style(ws1, row_idx, 5, TOTAL_FILL, TOTAL_FONT, CTR_ALIGN)
    ws1.row_dimensions[row_idx].height = 20

    # ════════════════════════════════════════════════════════════════════════
    #  Sheet 2 — Summary (one row per unique Crop × Disease)
    # ════════════════════════════════════════════════════════════════════════
    ws2       = wb.create_sheet("Summary")
    headers2  = ["#", "Crop", "Disease", "# Sources", "Source Datasets",
                 "Total Ref", "Total Bench"]
    widths2   = [5, 20, 30, 10, 44, 12, 13]
    _hdr(ws2, headers2, widths2)

    for i, (pk, info) in enumerate(sorted_pairs, start=1):
        sources_str = ", ".join(sorted(info["sources"]))
        n_src       = len(info["sources"])
        ws2.append([i, info["crop"], info["disease"], n_src,
                    sources_str, info["total_ref"], info["total_bench"]])
        r = i + 1
        fill = ALT_FILL_A if i % 2 == 0 else ALT_FILL_B
        _style(ws2, r, 7, fill, BODY_FONT)
        ws2.cell(r, 5).alignment = WRAP_ALIGN
        ws2.cell(r, 1).alignment = CTR_ALIGN
        ws2.cell(r, 4).alignment = CTR_ALIGN
        ws2.cell(r, 6).alignment = CTR_ALIGN
        ws2.cell(r, 7).alignment = CTR_ALIGN
        ws2.row_dimensions[r].height = 16

    # Totals row
    n2 = len(sorted_pairs)
    ws2.append(["", "TOTAL", f"=COUNTA(C2:C{n2+1})", "", "",
                f"=SUM(F2:F{n2+1})", f"=SUM(G2:G{n2+1})"])
    _style(ws2, n2+2, 7, TOTAL_FILL, TOTAL_FONT, CTR_ALIGN)
    ws2.cell(n2+2, 2).alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[n2+2].height = 20

    wb.save(output_path)
    print(f"  XLSX -> {os.path.abspath(output_path)}")
    print(f"         {len(sorted_pairs)} unique crop-disease pairs")


# ═══════════════════════════════════════════════════════════════════════════
#  PLANT VILLAGE LOADER  (tensorflow_datasets)
# ═══════════════════════════════════════════════════════════════════════════

def load_PlantVillage(n_ref, n_bench):
    """
    Loads PlantVillage via tensorflow_datasets.
    Label format:  "Crop___Disease"  (triple underscore)
    Each label becomes one class; crop and disease are split on "___".

    Requires:  pip install tensorflow tensorflow_datasets
    """
    name = "PlantVillage"
    desc = ("PlantVillage dataset loaded via tensorflow_datasets. "
            "Covers 38 crop-disease classes across 14 crop species including "
            "healthy controls. Images are RGB leaf photos taken under controlled "
            "conditions.")

    if not TFDS_AVAILABLE:
        print("  [SKIP] PlantVillage — tensorflow_datasets not installed. "
              "Run:  pip install tensorflow tensorflow_datasets")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, [])
        # Recover class list from saved folders
        src_tag  = safe_name(name)[:12]
        classes  = []
        for crop_d in (os.listdir(REF_DIR) if os.path.isdir(REF_DIR) else []):
            cp = os.path.join(REF_DIR, crop_d)
            if not os.path.isdir(cp): continue
            for dis_d in os.listdir(cp):
                dp = os.path.join(cp, dis_d)
                if not os.path.isdir(dp): continue
                if any(src_tag in f for f in os.listdir(dp) if f.lower().endswith(IMAGE_EXT)):
                    classes.append(f"{crop_d}__{dis_d}")
        return name, ref, bench, sorted(classes), desc

    dl = os.path.join(DATA_ROOT, "plantvillage-raw")
    if os.path.isdir(dl):
        shutil.rmtree(dl, ignore_errors=True)
    os.makedirs(dl, exist_ok=True)

    print(f"  Loading PlantVillage via tensorflow_datasets -> {dl}")
    try:
        ds, ds_info = tfds.load(
            "plant_village",
            split="train",
            as_supervised=True,
            with_info=True,
        )
    except Exception as e:
        print(f"  [ERROR] Could not load PlantVillage: {e}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    labels      = ds_info.features["label"].names   # e.g. ["Apple___Apple_scab", ...]
    needed      = n_ref + n_bench
    counters    = {}   # cls_key -> count saved

    print(f"  PlantVillage: {len(labels)} classes, streaming images ...")

    for img_tensor, label_idx in ds:
        label_raw = labels[int(label_idx)]          # "Apple___Apple_scab"
        # Normalise: split on "___" -> (crop, disease), safe_name each part
        if "___" in label_raw:
            parts   = label_raw.split("___", 1)
            crop    = safe_name(parts[0].replace("_", " ").strip())
            disease = safe_name(parts[1].replace("_", " ").strip())
        else:
            crop    = safe_name(label_raw)
            disease = "Unknown"

        cls_key = f"{crop}_{disease}"               # single underscore — universal format
        cls_dir = os.path.join(dl, cls_key)
        os.makedirs(cls_dir, exist_ok=True)

        if counters.get(cls_key, 0) >= needed:
            continue                                 # already have enough for this class

        try:
            img_np  = img_tensor.numpy()
            pil_img = Image.fromarray(img_np).convert("RGB")
            count   = counters.get(cls_key, 0)
            dest    = os.path.join(cls_dir, f"{count+1:05d}.jpg")
            pil_img.save(dest, "JPEG", quality=85)
            counters[cls_key] = count + 1
        except Exception:
            continue

        # Stop early once every class has enough images
        if all(v >= needed for v in counters.values()) and len(counters) == len(labels):
            break

    total_saved = sum(counters.values())
    print(f"  PlantVillage: {len(counters)} classes, {total_saved} images staged")

    if total_saved == 0:
        print(f"  [ERROR] No images saved for PlantVillage")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    data    = collect_images_df(dl)
    classes = sorted(data[1].unique().tolist()) if len(data) else []
    ref, bench = split_save_cleanup(data, classes, name, dl, n_ref, n_bench)
    return name, ref, bench, classes, desc


# ═══════════════════════════════════════════════════════════════════════════
#  CDDM LOCAL LOADER
#  Structure: CDDM-images/images/<Crop,Disease>/  (comma-separated folder names)
#  Each folder is both a crop and a disease — no sub-hierarchy needed.
# ═══════════════════════════════════════════════════════════════════════════

def load_CDDM(n_ref, n_bench):
    """
    Loads the CDDM (Comma-Delimited Disease Map) local dataset.
    Folder structure:  <root>/<Crop,Disease>/plant_xxxxx.jpg
    The comma in the folder name separates Crop from Disease, e.g.:
      "Apple,Brown Spot"  ->  crop="Apple", disease="Brown Spot"
    Each folder becomes one class. Classes with fewer than (n_ref + n_bench)
    images are skipped with a warning.
    """
    name = "CDDM"
    desc = ("CDDM local dataset. Classes encoded as 'Crop,Disease' folder names. "
            f"Source: {CDDM_SOURCE_ROOT}")

    if not CDDM_SOURCE_ROOT or not os.path.isdir(CDDM_SOURCE_ROOT):
        print(f"  [SKIP] CDDM — path not found: {CDDM_SOURCE_ROOT}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    # Discover all class folders (any folder directly inside root)
    all_cls_dirs = sorted([
        d for d in os.listdir(CDDM_SOURCE_ROOT)
        if os.path.isdir(os.path.join(CDDM_SOURCE_ROOT, d))
        and not d.startswith('.')
    ])

    if not all_cls_dirs:
        print(f"  [SKIP] CDDM — no class folders found in {CDDM_SOURCE_ROOT}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    # Use safe_name of each folder as the class key stored on disk
    classes = [safe_name(d) for d in all_cls_dirs]

    if already_sampled(name, n_ref, n_bench):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, [])
        src_tag = safe_name(name)[:12]
        saved_classes = []
        for crop_d in (os.listdir(REF_DIR) if os.path.isdir(REF_DIR) else []):
            cp = os.path.join(REF_DIR, crop_d)
            if not os.path.isdir(cp): continue
            for dis_d in os.listdir(cp):
                dp = os.path.join(cp, dis_d)
                if not os.path.isdir(dp): continue
                if any(src_tag in f for f in os.listdir(dp)
                       if f.lower().endswith(IMAGE_EXT)):
                    saved_classes.append(f"{crop_d}__{dis_d}")
        return name, ref, bench, sorted(saved_classes), desc

    total_needed = n_ref + n_bench
    rows, included = [], []

    for raw_dir in all_cls_dirs:
        cls_key  = safe_name(raw_dir)
        src_path = os.path.join(CDDM_SOURCE_ROOT, raw_dir)
        imgs     = collect_images_recursive(src_path)
        if len(imgs) < total_needed:
            print(f"  [SKIP CLASS] {raw_dir}: {len(imgs)} imgs (need {total_needed})")
            continue
        included.append(cls_key)
        for p in imgs:
            rows.append({0: p, 1: cls_key})

    if not rows:
        print(f"  [WARN] No valid classes in CDDM")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    print(f"  [CDDM] {len(included)} classes with >= {total_needed} images")

    data     = pd.DataFrame(rows)
    ref_df   = sample_per_class(data, included, n_ref)
    used     = set(ref_df[0].tolist())
    rem      = data[~data[0].isin(used)]
    bench_df = sample_per_class(rem, included, n_bench)

    ref_df   = save_split(ref_df,   name, "reference")
    bench_df = save_split(bench_df, name, "benchmark")

    n_r = sum(1 for p in ref_df[0]   if p)
    n_b = sum(1 for p in bench_df[0] if p)
    print(f"  [SAVED] {name}: reference={n_r}  benchmark={n_b} "
          f"-> Curated_Dataset/.../CDDM/")

    return name, ref_df, bench_df, included, desc


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    os.makedirs(DATA_ROOT,  exist_ok=True)
    os.makedirs(REF_DIR,    exist_ok=True)
    os.makedirs(BENCH_DIR,  exist_ok=True)

    # ── One-time migration: merge legacy "Crop Diseases" folders -> "Crop" ───
    for base_dir in [REF_DIR, BENCH_DIR]:
        if not os.path.isdir(base_dir):
            continue
        for folder in list(os.listdir(base_dir)):
            folder_path = os.path.join(base_dir, folder)
            if not os.path.isdir(folder_path):
                continue
            normalised = folder
            for suffix in (" Diseases", " Disease"):
                if folder.endswith(suffix):
                    normalised = folder[:-len(suffix)].strip()
                    break
            if normalised == folder:
                continue
            target_path = os.path.join(base_dir, normalised)
            os.makedirs(target_path, exist_ok=True)
            for disease_dir in os.listdir(folder_path):
                src = os.path.join(folder_path, disease_dir)
                dst = os.path.join(target_path, disease_dir)
                if not os.path.isdir(src):
                    continue
                if os.path.exists(dst):
                    for f in os.listdir(src):
                        fsrc = os.path.join(src, f)
                        fdst = os.path.join(dst, f)
                        if not os.path.exists(fdst):
                            shutil.move(fsrc, fdst)
                    shutil.rmtree(src, ignore_errors=True)
                else:
                    shutil.move(src, dst)
            try:
                os.rmdir(folder_path)
                print(f"  [MIGRATE] Merged {folder!r} -> {normalised!r} in {base_dir}")
            except OSError:
                pass

    n_ref, n_bench = ask_sample_counts()

    all_datasets = []

    # ── Online datasets ───────────────────────────────────────────────────────
    print("\n" + "="*60)
    print("  ONLINE DATASETS")
    print("="*60)
    for loader in [load_SBRD, load_MangoLeaf, load_SoybeanPNAS,
                   load_BeanLeaf, load_YellowRust,
                   load_BananaLeaf, load_Cauliflower, load_Lettuce,
                   load_Cucumber, load_DurianLeaf, load_EggplantDisease,
                   load_CottonDisease, load_PumpkinLeaf, load_RoseLeaf,
                   load_CoconutDisease, load_VanillaDisease,
                   load_SugarLeafIDN, load_CucumberZenodo, load_LeafNet, load_PlantVillage,
                   load_NewPlantDiseases, load_PlantDoc,
                   load_PlantDiseasesDataset,
                   # load_FUSARIUM22,  # temporarily disabled
                   ]:
        print(f"\n-- {loader.__name__} --")
        result = loader(n_ref, n_bench)
        all_datasets.append(result)
        nm, ref, bench, cls, _ = result
        print(f"   {nm}: {len(ref)} ref | {len(bench)} bench | {len(cls)} classes")

    # ── InternalData local datasets ───────────────────────────────────────────
    if LOCAL_SOURCE_ROOT:
        if os.path.isdir(LOCAL_SOURCE_ROOT):
            print("\n" + "="*60)
            print("  LOCAL DATASETS  (InternalData)")
            print(f"  Source: {LOCAL_SOURCE_ROOT}")
            print("="*60)
            for cat, cls_map in build_local_class_map(LOCAL_SOURCE_ROOT).items():
                print(f"\n-- Local: {cat} --")
                result = load_local_category(cat, cls_map, n_ref, n_bench, LOCAL_SOURCE_ROOT)
                all_datasets.append(result)
                nm, ref, bench, cls, _ = result
                print(f"   {nm}: {len(ref)} ref | {len(bench)} bench | {len(cls)} classes")
        else:
            print(f"\n[WARN] LOCAL_SOURCE_ROOT not found: {LOCAL_SOURCE_ROOT}")

    # ── CDDM dataset ──────────────────────────────────────────────────────────
    if CDDM_SOURCE_ROOT and os.path.isdir(CDDM_SOURCE_ROOT):
        print("\n" + "="*60)
        print("  LOCAL DATASET  (CDDM)")
        print(f"  Source: {CDDM_SOURCE_ROOT}")
        print("="*60)
        print("\n-- load_CDDM --")
        result = load_CDDM(n_ref, n_bench)
        all_datasets.append(result)
        nm, ref, bench, cls, _ = result
        print(f"   {nm}: {len(ref)} ref | {len(bench)} bench | {len(cls)} classes")

    # ── PDF + XLSX ────────────────────────────────────────────────────────────
    print("\n" + "="*60)
    print("  BUILDING PDF + XLSX")
    print("="*60)
    generate_pdf(all_datasets, OUTPUT_PDF, n_ref, n_bench)
    generate_xlsx(all_datasets, OUTPUT_XLSX)

    # ── Manifest — count actual files on disk, flag shortfalls ───────────────
    def _count_saved(ds_name, split_base, src_tag):
        """Count images on disk tagged with this dataset's source tag."""
        total = 0
        if not os.path.isdir(split_base):
            return total
        for crop_dir in os.listdir(split_base):
            crop_path = os.path.join(split_base, crop_dir)
            if not os.path.isdir(crop_path):
                continue
            for disease_dir in os.listdir(crop_path):
                disease_path = os.path.join(crop_path, disease_dir)
                if not os.path.isdir(disease_path):
                    continue
                total += sum(
                    1 for f in os.listdir(disease_path)
                    if f.lower().endswith(IMAGE_EXT) and src_tag in f
                )
        return total

    def _shortfall_classes(ds_name, n_required):
        """Return list of (crop/disease, ref_count, bench_count) where count < n_required."""
        src_tag = safe_name(ds_name)[:12]
        issues  = []
        for split_base, split_label in [(REF_DIR, "ref"), (BENCH_DIR, "bench")]:
            if not os.path.isdir(split_base):
                continue
            for crop_dir in os.listdir(split_base):
                crop_path = os.path.join(split_base, crop_dir)
                if not os.path.isdir(crop_path):
                    continue
                for disease_dir in os.listdir(crop_path):
                    disease_path = os.path.join(crop_path, disease_dir)
                    if not os.path.isdir(disease_path):
                        continue
                    count = sum(
                        1 for f in os.listdir(disease_path)
                        if f.lower().endswith(IMAGE_EXT) and src_tag in f
                    )
                    if 0 < count < n_required:
                        issues.append(f"{crop_dir}/{disease_dir} [{split_label}]: "
                                      f"{count}/{n_required}")
        return issues

    ds_entries = []
    shortfall_report = []

    for r in all_datasets:
        ds_name = r[0]
        classes = r[3]
        src_tag = safe_name(ds_name)[:12]

        n_ref_saved   = _count_saved(ds_name, REF_DIR,   src_tag)
        n_bench_saved = _count_saved(ds_name, BENCH_DIR, src_tag)
        expected_ref  = n_ref   * len(classes)
        expected_bench= n_bench * len(classes)

        shortfalls = _shortfall_classes(ds_name, n_ref) if classes else []
        if shortfalls:
            shortfall_report.append({"dataset": ds_name, "classes_below_target": shortfalls})

        ds_entries.append({
            "name":               ds_name,
            "classes":            len(classes),
            "reference_images":   n_ref_saved,
            "benchmark_images":   n_bench_saved,
            "expected_reference": expected_ref,
            "expected_benchmark": expected_bench,
            "complete":           (n_ref_saved == expected_ref and
                                   n_bench_saved == expected_bench),
        })

    manifest = {
        "reference_images_per_class": n_ref,
        "benchmark_images_per_class": n_bench,
        "datasets": ds_entries,
    }
    if shortfall_report:
        manifest["shortfalls"] = shortfall_report
        print("\n  [WARN] Some classes had fewer images than requested:")
        for s in shortfall_report:
            print(f"    {s['dataset']}:")
            for c in s["classes_below_target"]:
                print(f"      - {c}")

    mpath = os.path.join(SAMPLES_DIR, "manifest.json")
    with open(mpath, "w") as f:
        json.dump(manifest, f, indent=2)
    print(f"  Manifest -> {os.path.abspath(mpath)}\n")


if __name__ == "__main__":
    main()