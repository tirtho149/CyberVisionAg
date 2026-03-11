"""
Disease Dataset Curator & Report Generator
===========================================
Handles BOTH local datasets (dataloader.py) AND online Kaggle/Zenodo datasets
(data_loader.py) in one unified pipeline.

FEATURES
─────────
• Interactive prompt: asks how many REFERENCE (train) and BENCHMARK (test) images
  per class before doing anything.
• Smart downloading: skips any dataset whose output folder already exists.
• Two saved sets per class:
    samples/<DatasetName>/<ClassName>/reference/   ← train images
    samples/<DatasetName>/<ClassName>/benchmark/   ← test  images
  Files named:  <ClassName>_reference_001.jpg / <ClassName>_benchmark_001.jpg
• Deletes raw download after sampling to save disk space.
• PDF report with two image rows per class (Reference | Benchmark).
• Works with LOCAL folders (set LOCAL_SOURCE_ROOT) and/or ONLINE datasets.

REQUIREMENTS
────────────
  pip install kaggle reportlab Pillow pandas scikit-learn requests tqdm openpyxl
  pip install datasets          # required for LeafNet (HuggingFace)
  pip install kagglehub         # required for Banana Leaf dataset

USAGE
─────
  1.  Set LOCAL_SOURCE_ROOT below (or leave "" to skip local datasets).
  2.  Place ~/.kaggle/kaggle.json for Kaggle datasets.
  3.  python disease_report_generator.py
"""

# ═══════════════════════════════════════════════════════════════════════════
#  IMPORTS
# ═══════════════════════════════════════════════════════════════════════════

import os, sys, zipfile, shutil, difflib, random, json, re, time, hashlib
import multiprocessing
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import requests
from io import BytesIO

import pandas as pd
import numpy as np
from PIL import Image
from sklearn.utils import shuffle as sk_shuffle

# ── HuggingFace / datasets (optional — only needed for LeafNet) ──────────────
try:
    # Patch md5 for environments where usedforsecurity kwarg causes errors
    import hashlib as _hashlib
    _orig_md5 = _hashlib.md5
    def _patched_md5(*a, **kw):
        kw.pop("usedforsecurity", None)
        return _orig_md5(*a, **kw)
    _hashlib.md5 = _patched_md5

    from datasets import load_dataset as hf_load_dataset
    HF_AVAILABLE = True
except ImportError:
    HF_AVAILABLE = False

# ── tqdm shim ───────────────────────────────────────────────────────────────
try:
    from tqdm import tqdm as _tqdm
    def tqdm(iterable=None, **kw): return _tqdm(iterable, **kw)
except ImportError:
    class tqdm:
        def __init__(self, iterable=None, desc="", total=None, **kw):
            self._it = iter(iterable) if iterable is not None else None
            if desc: print(f"  {desc} ...")
        def __iter__(self):  return self._it
        def __next__(self):  return next(self._it)
        def __enter__(self): return self
        def __exit__(self, *a): pass
        def update(self, n=1): pass

# ── Kaggle ───────────────────────────────────────────────────────────────────
try:
    from kaggle.api.kaggle_api_extended import KaggleApi
    KAGGLE_AVAILABLE = True
except ImportError:
    KAGGLE_AVAILABLE = False

# ── kagglehub (used by Banana Leaf dataset) ──────────────────────────────────
try:
    import kagglehub
    KAGGLEHUB_AVAILABLE = True
except ImportError:
    KAGGLEHUB_AVAILABLE = False

# ── ReportLab ────────────────────────────────────────────────────────────────
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Image as RLImage, HRFlowable,
)

# ═══════════════════════════════════════════════════════════════════════════
#  ★  USER CONFIGURATION  ★
# ═══════════════════════════════════════════════════════════════════════════

# Path to your LOCAL disease image folder (set "" to skip local datasets)
LOCAL_SOURCE_ROOT = "InternalData"   # e.g. "/home/user/Desktop/Disease Images for Arti"

DATA_ROOT    = "./data"          # raw download staging area (deleted after use)
SAMPLES_DIR  = "./samples"       # permanent kept samples
OUTPUT_PDF   = "./disease_dataset_report.pdf"
OUTPUT_XLSX  = "./crop_disease_registry.xlsx"
RANDOM_STATE = 42
IMAGE_EXT    = ('.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif', '.webp')

# ── Worker threads: use ALL available logical cores ───────────────────────────
NUM_WORKERS = multiprocessing.cpu_count()
print(f"  [SYSTEM] {NUM_WORKERS} logical cores detected — threads will use all of them")

# ═══════════════════════════════════════════════════════════════════════════
#  INTERACTIVE PROMPT
# ═══════════════════════════════════════════════════════════════════════════

def ask_sample_counts():
    """Ask user how many reference (train) and benchmark (test) images per class."""
    print("\n" + "="*60)
    print("  DISEASE DATASET CURATOR — SETUP")
    print("="*60)
    print("  Two image sets will be saved per class:")
    print("    REFERENCE images  (training / reference samples)")
    print("    BENCHMARK images  (test / evaluation samples)")
    print()

    def ask_int(prompt, default):
        while True:
            raw = input(f"  {prompt} [default: {default}]: ").strip()
            if raw == "":
                return default
            try:
                v = int(raw)
                if v > 0:
                    return v
                print("  Please enter a positive integer.")
            except ValueError:
                print("  Invalid input — enter a whole number.")

    n_ref   = ask_int("How many REFERENCE images per class?", 5)
    n_bench = ask_int("How many BENCHMARK  images per class?", 10)
    print(f"\n  -> Reference: {n_ref}  |  Benchmark: {n_bench}  per class")
    print("="*60 + "\n")
    return n_ref, n_bench

# ═══════════════════════════════════════════════════════════════════════════
#  COLOURS
# ═══════════════════════════════════════════════════════════════════════════

DATASET_COLORS = {
    "SBRD":               "#2e9e5e",
    "Mango Leaf":         "#e8794a",
    "Soybean PNAS":       "#4a8ec2",
    "Bean Leaf":          "#9b6fd4",
    "Yellow Rust":        "#d4a82a",
    "FUSARIUM 22":        "#d45858",
    "Banana Leaf":        "#f5c518",
    "Cauliflower":        "#7cb342",
    "Lettuce":            "#00897b",
    "LeafNet":            "#1B5E20",
    "Alfalfa_Diseases":   "#4CAF50",
    "Corn_Diseases":      "#FF9800",
    "Soybean_Diseases":   "#2196F3",
    "Wheat_Diseases":     "#9C27B0",
    "Mango_Leaf_Disease": "#F44336",
}

def _hex(h):
    h = h.lstrip("#")
    r,g,b = (int(h[i:i+2],16)/255 for i in (0,2,4))
    return colors.Color(r,g,b)

PAGE_W, PAGE_H = A4
MARGIN = 1.8 * cm

# ═══════════════════════════════════════════════════════════════════════════
#  UTILITIES
# ═══════════════════════════════════════════════════════════════════════════

def safe_name(s):
    return re.sub(r'[^A-Za-z0-9_\-]', '_', s.replace(' ', '_'))

def get_closest_match(name, options):
    m = difflib.get_close_matches(name, options, n=1, cutoff=0.2)
    return m[0] if m else None

def rename_folders(base, expected):
    if not os.path.isdir(base): return
    for f in list(os.listdir(base)):
        fp = os.path.join(base, f)
        if not os.path.isdir(fp): continue
        m = get_closest_match(f, expected)
        if m and m != f:
            np_ = os.path.join(base, m)
            if not os.path.exists(np_): os.rename(fp, np_)

def rename_folders_dict(base, rmap):
    if not os.path.isdir(base): return
    for f in list(os.listdir(base)):
        fp = os.path.join(base, f)
        if os.path.isdir(fp) and f in rmap:
            np_ = os.path.join(base, rmap[f])
            if not os.path.exists(np_): os.rename(fp, np_)

def collect_images_flat(folder):
    """All image files directly inside folder (non-recursive)."""
    if not os.path.isdir(folder): return []
    return [os.path.join(folder, f) for f in os.listdir(folder)
            if f.lower().endswith(IMAGE_EXT)
            and os.path.isfile(os.path.join(folder, f))]

def collect_images_df(base_directory):
    """Walk class subdirs -> DataFrame(path=0, label=1)."""
    rows = []
    for sub in os.listdir(base_directory):
        if sub == ".DS_Store": continue
        sp = os.path.join(base_directory, sub)
        if not os.path.isdir(sp): continue
        for f in os.listdir(sp):
            if f.lower().endswith(IMAGE_EXT):
                rows.append({0: os.path.join(sp, f), 1: sub})
    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=[0,1])

# ─── Download helpers ────────────────────────────────────────────────────────

def download_file(url, filename):
    try:
        r = requests.get(url, stream=True, timeout=60)
        total = int(r.headers.get('content-length', 0))
        with open(filename, 'wb') as f, tqdm(desc=os.path.basename(filename),
                total=total, unit='iB', unit_scale=True) as bar:
            for chunk in r.iter_content(1024):
                f.write(chunk); bar.update(len(chunk))
    except Exception as e:
        print(f"  [ERROR] {url}: {e}")

def get_zenodo_urls(record_id):
    try:
        r = requests.get(f"https://zenodo.org/api/records/{record_id}", timeout=30)
        if r.status_code == 200:
            return [(f['links']['self'], f['key']) for f in r.json().get('files',[])]
    except Exception as e:
        print(f"  [ERROR] Zenodo: {e}")
    return []

def extract_zip(zpath, dest):
    try:
        with zipfile.ZipFile(zpath, 'r') as zf: zf.extractall(dest)
    except Exception as e:
        print(f"  [ERROR] Extract {zpath}: {e}")

def kaggle_download(dataset_name, path):
    if not KAGGLE_AVAILABLE:
        print(f"  [SKIP] kaggle not installed — {dataset_name}"); return False
    try:
        api = KaggleApi(); api.authenticate()
        print(f"  Downloading: {dataset_name}")
        api.dataset_download_files(dataset_name, path=path, unzip=True, quiet=False)
        return True
    except Exception as e:
        print(f"  [ERROR] Kaggle: {e}"); return False

# ─── Sample & save ───────────────────────────────────────────────────────────

def sample_per_class(data_df, classes, n):
    """Sample up to n images per class from DataFrame(path=0, label=1)."""
    if data_df is None or len(data_df) == 0:
        return pd.DataFrame(columns=[0,1])
    parts = []
    for cls in classes:
        cd = data_df[data_df[1] == cls]
        if len(cd) == 0:
            print(f"  [WARN] No images for '{cls}'"); continue
        parts.append(cd.sample(n=min(n, len(cd)), random_state=RANDOM_STATE))
    if not parts:
        return pd.DataFrame(columns=[0,1])
    return sk_shuffle(pd.concat(parts, ignore_index=True),
                      random_state=RANDOM_STATE).reset_index(drop=True)

def save_split(sampled_df, dataset_name, split_name):
    """
    Copy images to:
      samples/<dataset_name>/<ClassName>/<split_name>/
    Filename pattern:  <ClassName>_<split_name>_001.jpg

    Returns updated DataFrame with saved paths in column 0.
    Uses all available CPU threads for parallel file I/O.
    """
    if sampled_df is None or len(sampled_df) == 0:
        return sampled_df

    sds = safe_name(dataset_name)

    # Pre-compute (src, dest) pairs per class
    counters = {}
    tasks = []   # (src, dest, row_index)
    for idx, row in sampled_df.iterrows():
        src  = row[0]
        cls  = str(row[1])
        scls = safe_name(cls)
        dest_dir = os.path.join(SAMPLES_DIR, sds, scls, split_name)
        os.makedirs(dest_dir, exist_ok=True)
        counters[scls] = counters.get(scls, 0) + 1
        ext   = os.path.splitext(src)[1].lower() or '.jpg'
        fname = f"{scls}_{split_name}_{counters[scls]:03d}{ext}"
        dest  = os.path.join(dest_dir, fname)
        tasks.append((idx, src, dest))

    # Parallel copy
    results = {}   # row_index -> dest_path or None
    def _copy(args):
        idx, src, dest = args
        try:
            shutil.copy2(src, dest)
            return idx, dest
        except Exception as e:
            print(f"  [COPY ERROR] {src}: {e}")
            return idx, None

    with ThreadPoolExecutor(max_workers=NUM_WORKERS) as ex:
        for idx, dest in ex.map(_copy, tasks):
            results[idx] = dest

    result = sampled_df.copy()
    result[0] = [results.get(i) for i in sampled_df.index]
    return result

def split_save_cleanup(data_df, classes, dataset_name, download_path, n_ref, n_bench):
    """
    1. Sample n_ref  reference images per class (no overlap with benchmark).
    2. Sample n_bench benchmark images per class from remaining pool.
    3. Copy both sets into samples/<dataset_name>/.
    4. Delete download_path to free space.
    Returns (ref_df, bench_df).
    """
    ref_df = sample_per_class(data_df, classes, n_ref)

    # Exclude reference paths from benchmark pool
    used      = set(ref_df[0].tolist())
    remaining = data_df[~data_df[0].isin(used)]
    bench_df  = sample_per_class(remaining, classes, n_bench)

    ref_df   = save_split(ref_df,   dataset_name, "reference")
    bench_df = save_split(bench_df, dataset_name, "benchmark")

    n_r = sum(1 for p in ref_df[0]   if p)
    n_b = sum(1 for p in bench_df[0] if p)
    print(f"  [SAVED] reference={n_r}  benchmark={n_b} "
          f"-> samples/{safe_name(dataset_name)}/")

    if download_path and os.path.exists(download_path):
        shutil.rmtree(download_path, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {download_path}")

    return ref_df, bench_df

# ─── Smart skip: already sampled? ────────────────────────────────────────────

def already_sampled(dataset_name):
    """True if samples/<dataset_name>/ exists and contains at least one image."""
    p = os.path.join(SAMPLES_DIR, safe_name(dataset_name))
    if not os.path.isdir(p): return False
    for root, dirs, files in os.walk(p):
        if any(f.lower().endswith(IMAGE_EXT) for f in files):
            return True
    return False

def load_from_samples(dataset_name, classes):
    """Rebuild ref_df / bench_df from already-saved sample folders."""
    base = os.path.join(SAMPLES_DIR, safe_name(dataset_name))
    ref_rows, bench_rows = [], []
    for cls in classes:
        scls = safe_name(cls)
        for split, rows in [("reference", ref_rows), ("benchmark", bench_rows)]:
            d = os.path.join(base, scls, split)
            if not os.path.isdir(d): continue
            for f in sorted(os.listdir(d)):
                if f.lower().endswith(IMAGE_EXT):
                    rows.append({0: os.path.join(d, f), 1: cls})
    ref_df   = pd.DataFrame(ref_rows)   if ref_rows   else pd.DataFrame(columns=[0,1])
    bench_df = pd.DataFrame(bench_rows) if bench_rows else pd.DataFrame(columns=[0,1])
    return ref_df, bench_df

# ═══════════════════════════════════════════════════════════════════════════
#  ONLINE DATASET LOADERS
# ═══════════════════════════════════════════════════════════════════════════

def load_SBRD(n_ref, n_bench):
    name    = "SBRD"
    classes = ['Healthy','Mild Bacterial Blight','Mild Blast','Mild Brownspot',
               'Mild Tungro','Severe Bacterial Blight','Severe Blast',
               'Severe Brownspot','Severe Tungro']
    desc    = ("Rice leaf disease dataset with severity levels covering Bacterial "
               "Blight, Blast, Brownspot and Tungro at Mild and Severe stages.")
    if already_sampled(name):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc
    dl   = os.path.join(DATA_ROOT, "SBRD")
    base = os.path.join(dl, "Leaf Disease Dataset", "train")
    if not os.path.exists(dl):
        kaggle_download("isaacritharson/severity-based-rice-leaf-diseases-dataset", dl)
    rename_folders(base, classes)
    data = collect_images_df(base)
    ref, bench = split_save_cleanup(data, classes, name, dl, n_ref, n_bench)
    return name, ref, bench, classes, desc


def load_MangoLeaf(n_ref, n_bench):
    name    = "Mango Leaf"
    classes = ['Anthracnose','Bacterial Canker','Cutting Weevil','Die Back',
               'Gall Midge','Healthy','Powdery Mildew','Sooty Mould']
    desc    = ("Mango leaf disease dataset covering 7 diseases plus Healthy. "
               "Includes fungal, bacterial and pest-related leaf conditions.")
    if already_sampled(name):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc
    dl = os.path.join(DATA_ROOT, "mango-leaf-disease-dataset")
    if not os.path.exists(dl):
        kaggle_download("aryashah2k/mango-leaf-disease-dataset", dl)
    rename_folders(dl, classes)
    data = collect_images_df(dl)
    ref, bench = split_save_cleanup(data, classes, name, dl, n_ref, n_bench)
    return name, ref, bench, classes, desc


def load_SoybeanPNAS(n_ref, n_bench):
    name    = "Soybean PNAS"
    classes = ['Bacterial Blight','Bacterial Pustule','Frogeye Leaf Spot',
               'Healthy','Herbicide Injury','Iron Deficiency Chlorosis',
               'Potassium Deficiency','Septoria Brown Spot','Sudden Death Syndrome']
    rmap    = {str(i): c for i, c in enumerate(classes)}
    desc    = ("Soybean stress identification from PNAS. Covers 8 stress/disease "
               "conditions plus Healthy including nutrient deficiencies and "
               "fungal, bacterial and environmental stressors.")
    if already_sampled(name):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc
    dl   = os.path.join(DATA_ROOT, "Soybean-PNAS")
    base = os.path.join(dl, "Training Samples")
    if not os.path.exists(dl):
        os.makedirs(dl, exist_ok=True)
        orig = os.getcwd(); os.chdir(dl)
        for url, fname in get_zenodo_urls("12747481"): download_file(url, fname)
        os.chdir(orig)
        for f in os.listdir(dl):
            if f.lower() == "soybean_stress_identification.zip":
                extract_zip(os.path.join(dl, f), dl)
    rename_folders_dict(base, rmap)
    data = collect_images_df(base)
    ref, bench = split_save_cleanup(data, classes, name, dl, n_ref, n_bench)
    return name, ref, bench, classes, desc


def load_BeanLeaf(n_ref, n_bench):
    name    = "Bean Leaf"
    classes = ['Angular Leaf Spot','Bean Rust','Healthy']
    desc    = ("Bean leaf lesion classification: Angular Leaf Spot (bacterial), "
               "Bean Rust (fungal) and Healthy. Compact well-balanced dataset.")
    if already_sampled(name):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc
    dl   = os.path.join(DATA_ROOT, "bean-leaf-lesions-classification")
    base = os.path.join(dl, "train")
    if not os.path.exists(dl):
        kaggle_download("marquis03/bean-leaf-lesions-classification", dl)
    rename_folders(base, classes)
    data = collect_images_df(base)
    ref, bench = split_save_cleanup(data, classes, name, dl, n_ref, n_bench)
    return name, ref, bench, classes, desc


def load_YellowRust(n_ref, n_bench):
    name    = "Yellow Rust"
    rmap    = {'MR':'Moderately Resistant (MR)','MS':'Moderately Susceptible (MS)',
               'MRMS':'MRMS','0':'No Disease','R':'Resistant (R)','S':'Susceptible (S)'}
    classes = list(rmap.values())
    desc    = ("Yellow Rust 19 wheat disease dataset. Labelled by resistance level: "
               "Resistant -> Moderately Resistant -> MRMS -> Moderately Susceptible "
               "-> Susceptible, plus No Disease.")
    if already_sampled(name):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc
    dl   = os.path.join(DATA_ROOT, "yellowrust19")
    base = os.path.join(dl, "YELLOW-RUST-19", "YELLOW-RUST-19")
    if not os.path.exists(dl):
        kaggle_download("tolgahayit/yellowrust19-yellow-rust-disease-in-wheat", dl)
    rename_folders_dict(base, rmap)
    data = collect_images_df(base)
    ref, bench = split_save_cleanup(data, classes, name, dl, n_ref, n_bench)
    return name, ref, bench, classes, desc


def load_BananaLeaf(n_ref, n_bench):
    """
    Banana Leaf Disease Dataset v1.1 — gimrillozarita/banana-leaf-disease-dataset-v1-1
    4 classes: Cordana, Pestalotiopsis, Sigatoka, Healthy
    Downloaded via kagglehub (falls back to kaggle API if kagglehub unavailable).
    Uses the augmented split (400 imgs/class) for reliable sampling.
    """
    name    = "Banana Leaf"
    classes = ['Cordana', 'Healthy', 'Pestalotiopsis', 'Sigatoka']
    desc    = ("Banana leaf disease dataset (v1.1) covering three prominent fungal "
               "leaf spot diseases: Cordana, Pestalotiopsis, and Sigatoka, plus "
               "Healthy. Collected in Bangladesh; augmented set used for sampling.")

    if already_sampled(name):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc

    base = None

    # ── Download via kagglehub (preferred — as specified) ────────────────────
    if KAGGLEHUB_AVAILABLE:
        try:
            print(f"  [kagglehub] Downloading banana-leaf-disease-dataset-v1-1 ...")
            dl_path = kagglehub.dataset_download(
                "gimrillozarita/banana-leaf-disease-dataset-v1-1")
            print(f"  Path to dataset files: {dl_path}")

            # The dataset has original/ and augmented/ subfolders.
            # Prefer augmented (400 imgs/class) so we always have enough.
            for candidate in [
                os.path.join(dl_path, "augmented"),
                os.path.join(dl_path, "Augmented"),
                os.path.join(dl_path, "augmented set"),
                os.path.join(dl_path, "Augmented Set"),
                dl_path,   # fallback: root itself
            ]:
                if os.path.isdir(candidate):
                    # Check it actually contains class subfolders
                    subs = [d for d in os.listdir(candidate)
                            if os.path.isdir(os.path.join(candidate, d))
                            and not d.startswith('.')]
                    if subs:
                        base = candidate
                        break
        except Exception as e:
            print(f"  [kagglehub ERROR] {e}")

    # ── Fallback: kaggle API ──────────────────────────────────────────────────
    if base is None:
        print(f"  [FALLBACK] Trying kaggle API ...")
        dl = os.path.join(DATA_ROOT, "banana-leaf-disease")
        if not os.path.exists(dl):
            kaggle_download("gimrillozarita/banana-leaf-disease-dataset-v1-1", dl)
        # Search for the augmented folder
        for root, dirs, _ in os.walk(dl):
            for d in dirs:
                if "augment" in d.lower():
                    candidate = os.path.join(root, d)
                    subs = [s for s in os.listdir(candidate)
                            if os.path.isdir(os.path.join(candidate, s))]
                    if subs:
                        base = candidate
                        break
            if base:
                break
        if base is None:
            base = dl   # last resort: root of download

    if base is None or not os.path.isdir(base):
        print(f"  [ERROR] Could not locate Banana Leaf dataset folder.")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    # Normalise folder names to match classes list (case-insensitive)
    rename_folders(base, classes)

    data = collect_images_df(base)
    if len(data) == 0:
        print(f"  [ERROR] No images found in {base}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    # kagglehub downloads to a cache dir — we don't own it, so don't delete it.
    # Pass download_path=None to skip cleanup.
    ref_df  = sample_per_class(data, classes, n_ref)
    used    = set(ref_df[0].tolist())
    rem     = data[~data[0].isin(used)]
    bench_df = sample_per_class(rem, classes, n_bench)
    ref_df   = save_split(ref_df,   name, "reference")
    bench_df = save_split(bench_df, name, "benchmark")

    n_r = sum(1 for p in ref_df[0]   if p)
    n_b = sum(1 for p in bench_df[0] if p)
    print(f"  [SAVED] reference={n_r}  benchmark={n_b} "
          f"-> samples/{safe_name(name)}/")

    return name, ref_df, bench_df, classes, desc


def load_Lettuce(n_ref, n_bench):
    """
    Lettuce Diseases dataset — ashishjstar/lettuce-diseases
    Downloaded via kagglehub (falls back to kaggle API).
    Root: Lettuce_disease_datasets/
    8 classes (auto-discovered from subfolders):
      Bacterial, Downy_mildew_on_lettuce, Healthy,
      Powdery_mildew_on_lettuce, Septoria_blight_on_lettuce,
      Shepherd_purse_weed, Viral, Wilt_and_leaf_blight_or_rot
    Images per class vary — uses whatever is available.
    """
    name = "Lettuce"
    desc = ("Lettuce leaf disease dataset with 8 classes covering bacterial, "
            "fungal and viral conditions: Bacterial, Downy Mildew, Powdery Mildew, "
            "Septoria Blight, Viral, Wilt/Leaf Blight, Shepherd's Purse Weed, "
            "and Healthy. Image counts vary across classes.")

    if already_sampled(name):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        base_s = os.path.join(SAMPLES_DIR, safe_name(name))
        classes = sorted([
            d for d in os.listdir(base_s)
            if os.path.isdir(os.path.join(base_s, d))
        ])
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc

    base = None

    # ── Download via kagglehub (preferred) ───────────────────────────────────
    if KAGGLEHUB_AVAILABLE:
        try:
            print(f"  [kagglehub] Downloading lettuce-diseases ...")
            dl_path = kagglehub.dataset_download("ashishjstar/lettuce-diseases")
            print(f"  Path to dataset files: {dl_path}")
            # Root folder is Lettuce_disease_datasets/ or the dl_path itself
            for candidate in [
                os.path.join(dl_path, "Lettuce_disease_datasets"),
                dl_path,
            ]:
                if os.path.isdir(candidate):
                    subs = [d for d in os.listdir(candidate)
                            if os.path.isdir(os.path.join(candidate, d))
                            and not d.startswith('.')]
                    if subs:
                        base = candidate
                        break
        except Exception as e:
            print(f"  [kagglehub ERROR] {e}")

    # ── Fallback: kaggle API ──────────────────────────────────────────────────
    if base is None:
        print(f"  [FALLBACK] Trying kaggle API ...")
        dl = os.path.join(DATA_ROOT, "lettuce-diseases")
        if not os.path.exists(dl):
            kaggle_download("ashishjstar/lettuce-diseases", dl)
        for candidate in [
            os.path.join(dl, "Lettuce_disease_datasets"),
            dl,
        ]:
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))
                        and not d.startswith('.')]
                if subs:
                    base = candidate
                    break

    if base is None or not os.path.isdir(base):
        print(f"  [ERROR] Could not locate Lettuce dataset folder.")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    # Auto-discover classes from subfolders — handles any truncation/naming
    classes = sorted([
        d for d in os.listdir(base)
        if os.path.isdir(os.path.join(base, d)) and not d.startswith('.')
    ])
    print(f"  [Lettuce] Found {len(classes)} classes: {classes}")

    data = collect_images_df(base)
    if len(data) == 0:
        print(f"  [ERROR] No images found in {base}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    # kagglehub cache — don't delete; pass None to skip cleanup
    ref_df   = sample_per_class(data, classes, n_ref)
    used     = set(ref_df[0].tolist())
    rem      = data[~data[0].isin(used)]
    bench_df = sample_per_class(rem, classes, n_bench)
    ref_df   = save_split(ref_df,   name, "reference")
    bench_df = save_split(bench_df, name, "benchmark")

    n_r = sum(1 for p in ref_df[0]   if p)
    n_b = sum(1 for p in bench_df[0] if p)
    print(f"  [SAVED] reference={n_r}  benchmark={n_b} "
          f"-> samples/{safe_name(name)}/")

    return name, ref_df, bench_df, classes, desc


def load_Cauliflower(n_ref, n_bench):
    """
    Cauliflower Disease Multi Transformation Dataset
    kaggle: shuvokumarbasak2030/cauliflower-disease-multi-transformation-dataset
    Structure: dataset/dataset/train/{Bacterial_Spot_Rot, Black_Rot, Downy_Mildew, No_disease}
    4 classes; uses train split for sampling.
    """
    name    = "Cauliflower"
    classes = ['Bacterial_Spot_Rot', 'Black_Rot', 'Downy_Mildew', 'No_disease']
    desc    = ("Cauliflower leaf disease dataset with multi-transformation augmentation. "
               "Covers Bacterial Spot Rot, Black Rot, Downy Mildew and No Disease "
               "(healthy). Train/val/test splits available; train split used here.")

    if already_sampled(name):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc

    dl   = os.path.join(DATA_ROOT, "cauliflower-disease")
    # Folder structure: dataset/dataset/train/  (as seen in the screenshot)
    base = os.path.join(dl, "dataset", "dataset", "train")

    if not os.path.exists(dl):
        kaggle_download(
            "shuvokumarbasak2030/cauliflower-disease-multi-transformation-dataset", dl)

    if not os.path.isdir(base):
        # Fallback: search for 'train' folder anywhere inside dl
        for root, dirs, _ in os.walk(dl):
            if os.path.basename(root) == "train":
                subs = [d for d in os.listdir(root)
                        if os.path.isdir(os.path.join(root, d))]
                if subs:
                    base = root
                    break

    if not os.path.isdir(base):
        print(f"  [ERROR] Could not locate Cauliflower train folder in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    data = collect_images_df(base)
    if len(data) == 0:
        print(f"  [ERROR] No images found in {base}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    ref, bench = split_save_cleanup(data, classes, name, dl, n_ref, n_bench)
    return name, ref, bench, classes, desc


def load_FUSARIUM22(n_ref, n_bench):
    name    = "FUSARIUM 22"
    rmap    = {'1(HR)':'Highly Resistant','9(HS)':'Highly Susceptible',
               '5(MR)':'Moderately Resistant','3(R)':'Resistant','7(S)':'Susceptible'}
    classes = list(rmap.values())
    desc    = ("Fusarium Wilt disease severity in chickpea. Classes range from Highly "
               "Resistant to Highly Susceptible — ideal for ordinal severity tasks.")
    if already_sampled(name):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc
    dl   = os.path.join(DATA_ROOT, "fusarium22")
    base = os.path.join(dl, "FUSARIUM-22", "dataset_raw")
    if not os.path.exists(dl):
        kaggle_download("tolgahayit/fusarium-wilt-disease-in-chickpea-dataset", dl)
    rename_folders_dict(base, rmap)
    data = collect_images_df(base)
    ref, bench = split_save_cleanup(data, classes, name, dl, n_ref, n_bench)
    return name, ref, bench, classes, desc

# ═══════════════════════════════════════════════════════════════════════════
#  LEAFNET LOADER  (HuggingFace streaming — enalis/LeafNet)
# ═══════════════════════════════════════════════════════════════════════════

def _extract_crop_disease(caption):
    """Parse crop + disease label from a LeafNet caption string."""
    c  = str(caption).strip()
    cl = c.lower()

    # "a image of <crop> healthy leaves"
    m = re.search(r'a image of ([a-z]+) healthy leaves?', cl)
    if m:
        return m.group(1).capitalize(), "Healthy"

    # "a image of <crop> leaves diseased by <Disease> [with ...]"
    m = re.search(r'a image of ([a-z]+) leaves? diseased by ([a-z][a-z\s]*?)(?:\s+with|\s+disease|$)', cl)
    if m:
        return m.group(1).capitalize(), m.group(2).strip().title()

    # "a image of <crop> leaves with <Disease> [with symptoms ...]"
    m = re.search(r'a image of ([a-z]+) leaves? with ([a-z][a-z\s]*?)(?:\s+with|\s+symptoms|$)', cl)
    if m:
        disease = m.group(2).strip().title()
        if len(disease.split()) <= 5:
            return m.group(1).capitalize(), disease

    # Fallback: "<crop> leaves diseased by <Disease>"
    m = re.search(r'([a-z]+) leaves? diseased by ([a-z][a-z\s]*?)(?:\s+with|$)', cl)
    if m:
        return m.group(1).capitalize(), m.group(2).strip().title()

    return None


def _save_pil_image(pil_img, dest_path):
    """Convert a PIL image (any mode) to JPEG and save to dest_path."""
    try:
        img = pil_img.convert("RGB").resize((224, 224), Image.BILINEAR)
        img.save(dest_path, "JPEG", quality=85)
        return True
    except Exception as e:
        print(f"  [IMG SAVE] {e}")
        return False


def load_LeafNet(n_ref, n_bench):
    """
    Stream enalis/LeafNet from HuggingFace.
    - Scans the ENTIRE train split to discover ALL crop-disease classes.
    - Saves n_ref reference + n_bench benchmark images per class to disk
      (samples/LeafNet/<ClassName>/reference|benchmark/).
    - Smart skip: if samples already exist, rebuilds from disk.

    Requires:  pip install datasets
    Optional:  set HUGGINGFACE_HUB_TOKEN env var for private/gated access.
    """
    name = "LeafNet"
    desc = ("LeafNet — HuggingFace dataset (enalis/LeafNet). "
            "Multi-crop leaf disease classification derived from natural-language "
            "captions. Classes are (Crop, Disease) pairs spanning diverse species "
            "and conditions including Healthy controls.")

    if already_sampled(name):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        # Reconstruct class list from saved folder names
        base = os.path.join(SAMPLES_DIR, safe_name(name))
        classes = sorted([
            d for d in os.listdir(base)
            if os.path.isdir(os.path.join(base, d))
        ])
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc

    if not HF_AVAILABLE:
        print("  [SKIP] LeafNet — 'datasets' package not installed. "
              "Run:  pip install datasets")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    print(f"\n[LeafNet] Streaming full dataset — collecting ALL crop-disease classes")
    t0 = time.time()

    # Apply HF token if available
    import os as _os
    if "HUGGINGFACE_HUB_TOKEN" in _os.environ:
        _os.environ["HF_TOKEN"] = _os.environ["HUGGINGFACE_HUB_TOKEN"]

    try:
        ds = hf_load_dataset("enalis/LeafNet", split="train", streaming=True)
    except Exception as e:
        print(f"  [ERROR] Could not load LeafNet: {e}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    # ── Pass 1: stream rows → parse captions + buffer images in parallel ────
    # Producer (main thread) feeds a queue; N worker threads parse captions
    # and deposit results into buffers dict (protected by a lock).
    needed = n_ref + n_bench
    buffers     = {}          # key -> {"label": str, "images": [PIL...]}
    class_order = []
    lock        = threading.Lock()

    def _process_row(row):
        """Called from worker threads: parse caption, return (key, label, pil) or None."""
        result = _extract_crop_disease(row.get("caption", ""))
        if not result:
            return None
        crop, disease = result
        key       = f"{crop}__{disease}"
        cls_label = f"{crop} — {disease}"
        try:
            pil = row["image"]
            if isinstance(pil, np.ndarray):
                pil = Image.fromarray(pil)
        except Exception:
            pil = None
        return key, cls_label, pil

    scanned = 0
    with ThreadPoolExecutor(max_workers=NUM_WORKERS) as ex:
        futures = {}
        batch   = []
        BATCH   = NUM_WORKERS * 4   # submit in batches for low overhead

        for row in ds:
            scanned += 1
            batch.append(row)

            if len(batch) >= BATCH:
                for fut in [ex.submit(_process_row, r) for r in batch]:
                    futures[fut] = True
                batch = []

            # Collect completed futures
            done = [f for f in list(futures) if f.done()]
            for f in done:
                del futures[f]
                res = f.result()
                if res is None:
                    continue
                key, cls_label, pil = res
                with lock:
                    if key not in buffers:
                        buffers[key] = {"label": cls_label, "images": []}
                        class_order.append(key)
                        print(f"  [{scanned:>8,}]  New class {len(buffers):>3}: "
                              f"{cls_label}  ({time.time()-t0:.0f}s)")
                    if pil is not None and len(buffers[key]["images"]) < needed:
                        buffers[key]["images"].append(pil)

            if scanned % 20_000 == 0:
                full = sum(1 for v in buffers.values() if len(v["images"]) >= needed)
                print(f"  [{scanned:>8,}]  {len(buffers)} classes | "
                      f"{full} fully buffered  ({time.time()-t0:.0f}s)")

        # Flush remaining batch
        for fut in [ex.submit(_process_row, r) for r in batch]:
            res = fut.result()
            if res is None:
                continue
            key, cls_label, pil = res
            with lock:
                if key not in buffers:
                    buffers[key] = {"label": cls_label, "images": []}
                    class_order.append(key)
                if pil is not None and len(buffers[key]["images"]) < needed:
                    buffers[key]["images"].append(pil)

    print(f"\n  Scan complete: {len(buffers)} classes in {time.time()-t0:.1f}s")

    # ── Pass 2: save images to disk in parallel ───────────────────────────────
    classes = []
    save_tasks = []   # (pil, dest_path, cls_label, split_name)

    random.seed(RANDOM_STATE)
    for key in class_order:
        entry     = buffers[key]
        cls_label = entry["label"]
        imgs      = entry["images"]

        if len(imgs) < 2:
            print(f"  [SKIP class] {cls_label} — only {len(imgs)} image(s)")
            continue

        random.shuffle(imgs)
        ref_imgs   = imgs[:n_ref]
        bench_imgs = imgs[n_ref: n_ref + n_bench]

        scls = safe_name(cls_label)
        sds  = safe_name(name)
        classes.append(cls_label)

        for split_name, split_imgs in [("reference", ref_imgs), ("benchmark", bench_imgs)]:
            dest_dir = os.path.join(SAMPLES_DIR, sds, scls, split_name)
            os.makedirs(dest_dir, exist_ok=True)
            for idx, pil in enumerate(split_imgs, 1):
                fname = f"{scls}_{split_name}_{idx:03d}.jpg"
                dest  = os.path.join(dest_dir, fname)
                save_tasks.append((pil, dest, cls_label, split_name))

    # Worker: resize + JPEG encode + write
    def _save_task(args):
        pil, dest, cls_label, split_name = args
        ok = _save_pil_image(pil, dest)
        return dest, cls_label, split_name, ok

    print(f"  Saving {len(save_tasks)} images with {NUM_WORKERS} threads ...")
    t1 = time.time()
    ref_rows, bench_rows = [], []
    with ThreadPoolExecutor(max_workers=NUM_WORKERS) as ex:
        for dest, cls_label, split_name, ok in ex.map(_save_task, save_tasks):
            if ok:
                if split_name == "reference":
                    ref_rows.append({0: dest, 1: cls_label})
                else:
                    bench_rows.append({0: dest, 1: cls_label})
    print(f"  Save complete in {time.time()-t1:.1f}s")

    ref_df   = pd.DataFrame(ref_rows)   if ref_rows   else pd.DataFrame(columns=[0,1])
    bench_df = pd.DataFrame(bench_rows) if bench_rows else pd.DataFrame(columns=[0,1])

    print(f"  [SAVED] LeafNet — {len(ref_df)} ref | {len(bench_df)} bench "
          f"| {len(classes)} classes -> samples/{safe_name(name)}/")

    return name, ref_df, bench_df, classes, desc


# ═══════════════════════════════════════════════════════════════════════════
#  LOCAL DATASET LOADER
# ═══════════════════════════════════════════════════════════════════════════

def build_local_class_map(root):
    R = root
    return {
        "Alfalfa_Diseases": {
            "Verticillium_wilt": os.path.join(R, "Alfalfa Diseases", "Verticillium wilt"),
        },
        "Corn_Diseases": {
            "Anthracnose_leaf_spot":    os.path.join(R, "Corn Diseases", "Anthracnose leaf spot and top dieback"),
            "Anthracnose_stalk_rot":    os.path.join(R, "Corn Diseases", "Anthracnose stalk rot"),
            "Bacterial_stalk_rot":      os.path.join(R, "Corn Diseases", "Bacterial stalk rot"),
            "Carbonum_leaf_spot":       os.path.join(R, "Corn Diseases", "Carbonum leaf spot"),
            "Charcoal_stalk_rot":       os.path.join(R, "Corn Diseases", "Charcoal stalk rot"),
            "Common_rust":              os.path.join(R, "Corn Diseases", "Common rust"),
            "Common_smut":              os.path.join(R, "Corn Diseases", "Common smut"),
            "Corn_Stunt_Spiroplasma":   os.path.join(R, "Corn Diseases", "Corn Stunt Spiroplasma"),
            "Crazy_top":                os.path.join(R, "Corn Diseases", "Crazy top"),
            "Diplodia_stalk_rot":       os.path.join(R, "Corn Diseases", "Diplodia stalk rot"),
            "Downy_mildew":             os.path.join(R, "Corn Diseases", "Downy mildew"),
            "Ear_rot_Aspergillus":      os.path.join(R, "Corn Diseases", "Ear rots", "Aspergillus ear rot"),
            "Ear_rot_Diplodia":         os.path.join(R, "Corn Diseases", "Ear rots", "Diplodia ear rot"),
            "Ear_rot_Gibberella":       os.path.join(R, "Corn Diseases", "Ear rots", "Gibberella ear rot"),
            "Ear_rot_Trichoderma":      os.path.join(R, "Corn Diseases", "Ear rots", "Trichoderma ear rot"),
            "Eyespot":                  os.path.join(R, "Corn Diseases", "Eyespot"),
            "Fusarium":                 os.path.join(R, "Corn Diseases", "Fusarium"),
            "Goss_wilt":                os.path.join(R, "Corn Diseases", "Goss's wilt"),
            "Gray_leaf_spot":           os.path.join(R, "Corn Diseases", "Gray leaf spot"),
            "Head_smut":                os.path.join(R, "Corn Diseases", "Head smut"),
            "Northern_corn_leaf_blight":os.path.join(R, "Corn Diseases", "Northern corn leaf blight"),
            "Physoderma_brown_spot":    os.path.join(R, "Corn Diseases", "Physoderma brown spot"),
            "Southern_Corn_Leaf_Blight":os.path.join(R, "Corn Diseases", "Southern Corn Leaf Blight"),
            "Southern_rust":            os.path.join(R, "Corn Diseases", "Southern rust"),
            "Stewarts_disease":         os.path.join(R, "Corn Diseases", "Stewart's disease"),
            "Tar_spot":                 os.path.join(R, "Corn Diseases", "Tar spot"),
        },
        "Soybean_Diseases": {
            "Anthracnose":            os.path.join(R, "Soybean Diseases", "Anthracnose"),
            "Bacterial_Blight":       os.path.join(R, "Soybean Diseases", "Bacterial Blight"),
            "Bacterial_Pustule":      os.path.join(R, "Soybean Diseases", "Bacterial Pustule"),
            "Bean_Pod_Mottle_virus":  os.path.join(R, "Soybean Diseases", "Bean Pod Mottle virus"),
            "Brown_Stem_Rot":         os.path.join(R, "Soybean Diseases", "Brown Stem Rot"),
            "Cercospora":             os.path.join(R, "Soybean Diseases", "Cercospora"),
            "Charcoal_Rot":           os.path.join(R, "Soybean Diseases", "Charcoal Rot"),
            "Diaporthe":              os.path.join(R, "Soybean Diseases", "Diaporthe"),
            "Downy_mildew":           os.path.join(R, "Soybean Diseases", "Downy mildew"),
            "Frogeye_leaf_spot":      os.path.join(R, "Soybean Diseases", "Frogeye leaf spot"),
            "Fusarium":               os.path.join(R, "Soybean Diseases", "Fusarium"),
            "Green_stem":             os.path.join(R, "Soybean Diseases", "Green stem"),
            "Phomopsis":              os.path.join(R, "Soybean Diseases", "Phomopsis"),
            "Phyllosticta_leaf_spot": os.path.join(R, "Soybean Diseases", "Phyllosticta leaf spot"),
            "Phytophthora":           os.path.join(R, "Soybean Diseases", "Phytophthora"),
            "Powdery_Mildew":         os.path.join(R, "Soybean Diseases", "Powdery Mildew"),
            "Purple_Seed_Stain":      os.path.join(R, "Soybean Diseases", "Purple Seed Stain"),
            "Rhizoctonia":            os.path.join(R, "Soybean Diseases", "Rhizoctonia"),
            "Septoria_brown_spot":    os.path.join(R, "Soybean Diseases", "Septoria brown spot"),
            "Soybean_Cyst_Nematode":  os.path.join(R, "Soybean Diseases", "Soybean Cyst Nematode"),
            "Soybean_rust":           os.path.join(R, "Soybean Diseases", "Soybean rust"),
            "Stem_Canker":            os.path.join(R, "Soybean Diseases", "Stem Canker"),
            "Sudden_death_syndrome":  os.path.join(R, "Soybean Diseases", "Sudden death syndrome"),
            "White_Mold":             os.path.join(R, "Soybean Diseases", "White Mold"),
        },
        "Wheat_Diseases": {
            "Rust": os.path.join(R, "Wheat Diseases", "rust"),
        },
        "Mango_Leaf_Disease": {
            "Anthracnose":      os.path.join(R, "mango-leaf-disease-dataset", "Anthracnose"),
            "Bacterial_Canker": os.path.join(R, "mango-leaf-disease-dataset", "Bacterial Canker"),
            "Cutting_Weevil":   os.path.join(R, "mango-leaf-disease-dataset", "Cutting Weevil"),
            "Die_Back":         os.path.join(R, "mango-leaf-disease-dataset", "Die Back"),
            "Gall_Midge":       os.path.join(R, "mango-leaf-disease-dataset", "Gall Midge"),
            "Powdery_Mildew":   os.path.join(R, "mango-leaf-disease-dataset", "Powdery Mildew"),
            "Sooty_Mould":      os.path.join(R, "mango-leaf-disease-dataset", "Sooty Mould"),
        },
    }


def load_local_category(category_name, cls_map, n_ref, n_bench, source_root):
    """Process one local category. Returns (name, ref_df, bench_df, classes, desc)."""
    name = category_name
    desc = (f"Local disease dataset — {category_name.replace('_',' ')}. "
            f"Source: {source_root}")

    if already_sampled(name):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        classes = list(cls_map.keys())
        ref, bench = load_from_samples(name, classes)
        return name, ref, bench, classes, desc

    total_needed = n_ref + n_bench
    rows, included = [], []

    for cls_name, src_folder in cls_map.items():
        imgs = collect_images_flat(src_folder)
        if len(imgs) < total_needed:
            print(f"  [SKIP CLASS] {cls_name}: {len(imgs)} imgs (need {total_needed})")
            continue
        included.append(cls_name)
        for p in imgs:
            rows.append({0: p, 1: cls_name})

    if not rows:
        print(f"  [WARN] No valid classes in {category_name}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    data = pd.DataFrame(rows)
    # Local: no download_path to delete (user's original files stay intact)
    ref_df  = sample_per_class(data, included, n_ref)
    used    = set(ref_df[0].tolist())
    rem     = data[~data[0].isin(used)]
    bench_df = sample_per_class(rem, included, n_bench)

    ref_df   = save_split(ref_df,   name, "reference")
    bench_df = save_split(bench_df, name, "benchmark")

    print(f"  [SAVED] {name}: reference={len(ref_df)}  benchmark={len(bench_df)}")
    return name, ref_df, bench_df, included, desc

# ═══════════════════════════════════════════════════════════════════════════
#  PDF REPORT
# ═══════════════════════════════════════════════════════════════════════════

def build_styles():
    base = getSampleStyleSheet()
    return {
        "ReportTitle":  ParagraphStyle("ReportTitle",  parent=base["Title"],
            fontSize=24, leading=30, spaceAfter=6, alignment=TA_CENTER,
            textColor=colors.HexColor("#3a3a5e"), fontName="Helvetica-Bold"),
        "ReportSubtitle": ParagraphStyle("ReportSubtitle", parent=base["Normal"],
            fontSize=10, leading=14, spaceAfter=4, alignment=TA_CENTER,
            textColor=colors.HexColor("#666688"), fontName="Helvetica"),
        "DatasetTitle": ParagraphStyle("DatasetTitle", parent=base["Heading1"],
            fontSize=14, leading=18, textColor=colors.white,
            fontName="Helvetica-Bold", alignment=TA_LEFT),
        "DatasetDesc":  ParagraphStyle("DatasetDesc",  parent=base["Normal"],
            fontSize=8.5, leading=12, spaceAfter=5,
            textColor=colors.HexColor("#444444"), fontName="Helvetica"),
        "ClassLabel":   ParagraphStyle("ClassLabel",   parent=base["Normal"],
            fontSize=6.5, leading=8, spaceAfter=0, alignment=TA_CENTER,
            textColor=colors.HexColor("#555555"), fontName="Helvetica"),
        "SplitLabel":   ParagraphStyle("SplitLabel",   parent=base["Normal"],
            fontSize=7.5, leading=10, alignment=TA_LEFT,
            textColor=colors.HexColor("#3a3a6a"), fontName="Helvetica-Bold"),
        "SectionLabel": ParagraphStyle("SectionLabel", parent=base["Normal"],
            fontSize=9, leading=11, spaceBefore=3, spaceAfter=2,
            textColor=colors.HexColor("#3a3a6a"), fontName="Helvetica-Bold"),
        "InfoCellBold": ParagraphStyle("InfoCellBold", parent=base["Normal"],
            fontSize=8, leading=10, textColor=colors.HexColor("#222222"),
            fontName="Helvetica-Bold"),
        "InfoCell":     ParagraphStyle("InfoCell",     parent=base["Normal"],
            fontSize=8, leading=10, textColor=colors.HexColor("#444444"),
            fontName="Helvetica"),
        "TOCEntry":     ParagraphStyle("TOCEntry",     parent=base["Normal"],
            fontSize=9, leading=14, leftIndent=10,
            textColor=colors.HexColor("#333355"), fontName="Helvetica"),
        "NoDataMsg":    ParagraphStyle("NoDataMsg",    parent=base["Normal"],
            fontSize=9, leading=12, spaceAfter=8, alignment=TA_CENTER,
            textColor=colors.HexColor("#cc4444"), fontName="Helvetica-Oblique"),
    }

def thumb(img_path, w=130, h=130):
    if not img_path or not os.path.isfile(img_path): return None
    try:
        with Image.open(img_path) as im:
            im = im.convert("RGB")
            im.thumbnail((w, h), Image.LANCZOS)
            buf = BytesIO(); im.save(buf, "JPEG", quality=80); buf.seek(0)
            return buf
    except Exception as e:
        print(f"  [IMG] {e}"); return None

def make_banner(title, hx, width, sty):
    tbl = Table([[Paragraph(title, sty["DatasetTitle"])]],
                colWidths=[width], rowHeights=[24])
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), _hex(hx)),
        ("LEFTPADDING",   (0,0),(-1,-1), 10),
        ("RIGHTPADDING",  (0,0),(-1,-1), 10),
        ("TOPPADDING",    (0,0),(-1,-1), 4),
        ("BOTTOMPADDING", (0,0),(-1,-1), 4),
    ]))
    return tbl

def make_image_row(paths, n_cols, IW, IH, LH, sty, split_label):
    # Always renders exactly 5 columns — caller must pre-slice paths to [:5]
    cells = []
    for p in (paths or [])[:5]:
        buf   = thumb(p)
        fname = os.path.basename(p)[:18] if p else ""
        if buf:
            cells.append([RLImage(buf, width=IW, height=IH),
                          Paragraph(fname, sty["ClassLabel"])])
        else:
            cells.append([Paragraph("(missing)", sty["ClassLabel"])])
    while len(cells) < 5:
        cells.append([""])
    CW = IW + 0.12*cm
    tbl = Table([cells], colWidths=[CW]*5, rowHeights=[IH+LH])
    tbl.setStyle(TableStyle([
        ("ALIGN",         (0,0),(-1,-1),"CENTER"),
        ("VALIGN",        (0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",    (0,0),(-1,-1), 2),
        ("BOTTOMPADDING", (0,0),(-1,-1), 2),
        ("LEFTPADDING",   (0,0),(-1,-1), 3),
        ("RIGHTPADDING",  (0,0),(-1,-1), 3),
    ]))
    return tbl

def build_class_grid(ref_df, bench_df, classes, hx, W, sty, n_ref, n_bench):
    # PDF always shows exactly 5 images per row — hard cap, no exceptions
    PDF_COLS = 5
    IW = (W - 0.4*cm) / PDF_COLS
    IH = IW
    LH = 0.75*cm
    accent = _hex(hx)
    flowables = []

    # Pre-collect all image paths that need thumbnailing
    class_data = []
    for cls in classes:
        ref_paths   = (ref_df[ref_df[1]==cls][0].tolist()
                       if ref_df is not None and len(ref_df) > 0 else [])
        bench_paths = (bench_df[bench_df[1]==cls][0].tolist()
                       if bench_df is not None and len(bench_df) > 0 else [])
        if not ref_paths and not bench_paths:
            continue
        ref_paths   = ref_paths[:PDF_COLS]
        bench_paths = bench_paths[:PDF_COLS]
        class_data.append((cls, ref_paths, bench_paths))

    # Parallel thumbnail generation
    all_paths = []
    for cls, rp, bp in class_data:
        all_paths.extend(rp + bp)

    thumb_cache = {}
    def _make_thumb(p):
        return p, thumb(p)

    with ThreadPoolExecutor(max_workers=NUM_WORKERS) as ex:
        for p, buf in ex.map(_make_thumb, all_paths):
            thumb_cache[p] = buf

    # Build flowables using cached thumbnails
    for cls, ref_paths, bench_paths in class_data:
        # Class header bar
        hdr = Table([[Paragraph(f"<b>{cls.replace('_',' ')}</b>",
                                sty["SectionLabel"])]],
                    colWidths=[W], rowHeights=[17])
        hdr.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1,-1), colors.HexColor(hx+"18")),
            ("LEFTPADDING",   (0,0),(-1,-1), 8),
            ("RIGHTPADDING",  (0,0),(-1,-1), 8),
            ("TOPPADDING",    (0,0),(-1,-1), 2),
            ("BOTTOMPADDING", (0,0),(-1,-1), 2),
            ("LINEBELOW",     (0,0),(-1,-1), 1.0, accent),
        ]))
        flowables.append(hdr)
        flowables.append(Spacer(1, 2))

        def _row(paths, label):
            cells = []
            for p in (paths or [])[:5]:
                buf   = thumb_cache.get(p)
                fname = os.path.basename(p)[:18] if p else ""
                if buf:
                    cells.append([RLImage(buf, width=IW, height=IH),
                                  Paragraph(fname, sty["ClassLabel"])])
                else:
                    cells.append([Paragraph("(missing)", sty["ClassLabel"])])
            while len(cells) < 5:
                cells.append([""])
            CW  = IW + 0.12*cm
            tbl = Table([cells], colWidths=[CW]*5, rowHeights=[IH+LH])
            tbl.setStyle(TableStyle([
                ("ALIGN",         (0,0),(-1,-1),"CENTER"),
                ("VALIGN",        (0,0),(-1,-1),"MIDDLE"),
                ("TOPPADDING",    (0,0),(-1,-1), 2),
                ("BOTTOMPADDING", (0,0),(-1,-1), 2),
                ("LEFTPADDING",   (0,0),(-1,-1), 3),
                ("RIGHTPADDING",  (0,0),(-1,-1), 3),
            ]))
            return tbl

        if ref_paths:
            flowables.append(Paragraph(
                f"  Reference  ({len(ref_paths)} images)", sty["SplitLabel"]))
            flowables.append(_row(ref_paths, "ref"))

        if bench_paths:
            flowables.append(Paragraph(
                f"  Benchmark  ({len(bench_paths)} images)", sty["SplitLabel"]))
            flowables.append(_row(bench_paths, "bench"))

        flowables.append(Spacer(1, 5))

    return flowables

def make_summary_table(all_datasets, sty, n_ref, n_bench):
    TYPE = {
        "SBRD":"Disease + Severity","Mango Leaf":"Disease ID",
        "Soybean PNAS":"Disease ID","Bean Leaf":"Disease ID",
        "Yellow Rust":"Severity / Resistance","FUSARIUM 22":"Severity / Resistance",
        "Banana Leaf":"Disease ID",
        "Cauliflower":"Disease ID",
        "Lettuce":"Disease ID",
        "LeafNet":"Disease ID (HuggingFace)",
        "Alfalfa_Diseases":"Local Disease","Corn_Diseases":"Local Disease",
        "Soybean_Diseases":"Local Disease","Wheat_Diseases":"Local Disease",
        "Mango_Leaf_Disease":"Local Disease",
    }
    rows = [["#","Dataset","Classes","Ref/cls","Bench/cls","Type"]]
    for i,(nm,ref,bench,cls,_) in enumerate(all_datasets,1):
        ok = hasattr(ref,'__len__') and len(ref)>0
        rows.append([str(i), nm, str(len(cls)),
                     str(n_ref) if ok else "N/A",
                     str(n_bench) if ok else "N/A",
                     TYPE.get(nm,"Disease")])
    cw = [0.7*cm,4.0*cm,1.6*cm,1.8*cm,2.2*cm,4.2*cm]
    t  = Table(rows, colWidths=cw, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0),colors.HexColor("#5a5a8a")),
        ("TEXTCOLOR",     (0,0),(-1,0),colors.white),
        ("FONTNAME",      (0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,0), 8),
        ("ALIGN",         (0,0),(-1,0),"CENTER"),
        ("TOPPADDING",    (0,0),(-1,0), 6),
        ("BOTTOMPADDING", (0,0),(-1,0), 6),
        ("FONTNAME",      (0,1),(-1,-1),"Helvetica"),
        ("FONTSIZE",      (0,1),(-1,-1), 7.5),
        ("TOPPADDING",    (0,1),(-1,-1), 4),
        ("BOTTOMPADDING", (0,1),(-1,-1), 4),
        ("ALIGN",         (0,1),(0,-1),"CENTER"),
        ("ALIGN",         (2,1),(4,-1),"CENTER"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),
            [colors.HexColor("#fafafa"),colors.HexColor("#f2f2f9")]),
        ("GRID",          (0,0),(-1,-1),0.4,colors.HexColor("#ccccdd")),
    ]))
    return t

def page_footer(canvas_obj, doc):
    canvas_obj.saveState()
    canvas_obj.setFont("Helvetica", 7)
    canvas_obj.setFillColor(colors.HexColor("#aaaaaa"))
    canvas_obj.drawCentredString(PAGE_W/2, 0.6*cm,
        f"Disease Dataset Report  •  Page {doc.page}")
    canvas_obj.restoreState()

PDF_DISPLAY_MAX = 5   # PDF always shows at most 5 images per row, no exceptions

def generate_pdf(all_datasets, output_path, n_ref, n_bench):
    # Hard cap: regardless of what the user chose, the PDF never shows > 5 per row
    n_ref  = min(n_ref,  PDF_DISPLAY_MAX)
    n_bench = min(n_bench, PDF_DISPLAY_MAX)
    doc = SimpleDocTemplate(output_path, pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=MARGIN,  bottomMargin=1.3*cm,
        title="Disease Dataset Report")
    W   = PAGE_W - 2*MARGIN
    sty = build_styles()
    story = []

    # Cover
    story.append(Spacer(1, 1.8*cm))
    ttbl = Table([[Paragraph(
        "Disease Classification &amp; Severity<br/>Dataset Report", sty["ReportTitle"])]],
        colWidths=[W])
    ttbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1),colors.HexColor("#f8f8fd")),
        ("TOPPADDING",    (0,0),(-1,-1),16),
        ("BOTTOMPADDING", (0,0),(-1,-1),16),
        ("LEFTPADDING",   (0,0),(-1,-1),12),
        ("RIGHTPADDING",  (0,0),(-1,-1),12),
        ("LINEABOVE",     (0,0),(-1,0), 2,colors.HexColor("#4a4a7a")),
        ("LINEBELOW",     (0,0),(-1,-1),2,colors.HexColor("#4a4a7a")),
    ]))
    story.append(ttbl)
    story.append(Spacer(1, 0.35*cm))
    story.append(Paragraph(
        f"Reference: {n_ref} imgs/class  |  Benchmark: {n_bench} imgs/class  "
        f"|  {len(all_datasets)} datasets",
        sty["ReportSubtitle"]))
    story.append(Spacer(1, 0.35*cm))
    story.append(HRFlowable(width=W, thickness=0.5,
                             color=colors.HexColor("#ccccdd")))
    story.append(Spacer(1, 0.6*cm))

    # Legend 2-col
    legend = []
    for nm,_,_,_,_ in all_datasets:
        hx = DATASET_COLORS.get(nm,"#888888")
        dot = Table([[""]], colWidths=[0.32*cm], rowHeights=[0.32*cm])
        dot.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),_hex(hx))]))
        legend.append([dot, Paragraph(nm, sty["TOCEntry"])])
    half = len(legend)//2 + len(legend)%2
    for i in range(half):
        l = legend[i]
        r = legend[i+half] if i+half < len(legend) else ["",""]
        story.append(Table([[l[0],l[1],r[0],r[1]]],
            colWidths=[0.45*cm, W/2-0.45*cm, 0.45*cm, W/2-0.45*cm]))

    story.append(Spacer(1, 0.7*cm))
    story.append(Paragraph("Dataset Overview", sty["SectionLabel"]))
    story.append(Spacer(1, 3))
    story.append(make_summary_table(all_datasets, sty, n_ref, n_bench))
    story.append(PageBreak())

    # Per-dataset pages
    for ds_name, ref_df, bench_df, classes, desc in all_datasets:
        hx = DATASET_COLORS.get(ds_name, "#5a5a8a")

        story.append(make_banner(ds_name, hx, W, sty))
        story.append(Spacer(1, 4))
        story.append(Paragraph(desc, sty["DatasetDesc"]))

        n_r = len(ref_df)   if hasattr(ref_df,  '__len__') else 0
        n_b = len(bench_df) if hasattr(bench_df,'__len__') else 0
        info = Table([[
            Paragraph("<b>Classes</b>",        sty["InfoCellBold"]),
            Paragraph(str(len(classes)),        sty["InfoCell"]),
            Paragraph("<b>Reference imgs</b>",  sty["InfoCellBold"]),
            Paragraph(str(n_r),                 sty["InfoCell"]),
            Paragraph("<b>Benchmark imgs</b>",  sty["InfoCellBold"]),
            Paragraph(str(n_b),                 sty["InfoCell"]),
        ]], colWidths=[2.4*cm,1.1*cm,2.8*cm,1.1*cm,3.0*cm,1.1*cm])
        info.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1,-1),colors.HexColor("#fafafe")),
            ("BOX",           (0,0),(-1,-1),0.5,colors.HexColor("#ccccdd")),
            ("TOPPADDING",    (0,0),(-1,-1),4),
            ("BOTTOMPADDING", (0,0),(-1,-1),4),
            ("LEFTPADDING",   (0,0),(-1,-1),7),
            ("RIGHTPADDING",  (0,0),(-1,-1),7),
            ("VALIGN",        (0,0),(-1,-1),"MIDDLE"),
        ]))
        story.append(info)
        story.append(Spacer(1, 5))

        cls_str = "  |  ".join(c.replace("_"," ") for c in classes)
        story.append(Paragraph(f"<b>Classes:</b>  {cls_str}", sty["DatasetDesc"]))
        story.append(HRFlowable(width=W, thickness=0.6,
                                 color=_hex(hx), spaceAfter=5))

        no_data = (not hasattr(ref_df,'__len__') or len(ref_df)==0)
        if no_data:
            story.append(Paragraph(
                "No data available — check credentials/source path and re-run.",
                sty["NoDataMsg"]))
        else:
            story.extend(build_class_grid(
                ref_df, bench_df, classes, hx, W, sty, n_ref, n_bench))

        story.append(PageBreak())

    doc.build(story, onFirstPage=page_footer, onLaterPages=page_footer)
    print(f"\n  PDF -> {os.path.abspath(output_path)}")

# ═══════════════════════════════════════════════════════════════════════════
#  XLSX EXPORT — Deduplicated Crop × Disease Registry
# ═══════════════════════════════════════════════════════════════════════════

# Known crop per dataset (for datasets where class label = disease only)
DATASET_CROP = {
    "SBRD":               "Rice",
    "Mango Leaf":         "Mango",
    "Soybean PNAS":       "Soybean",
    "Bean Leaf":          "Bean",
    "Yellow Rust":        "Wheat",
    "FUSARIUM 22":        "Chickpea",
    "Banana Leaf":        "Banana",
    "Cauliflower":        "Cauliflower",
    "Lettuce":            "Lettuce",
    # LeafNet encodes crop inside the class label itself ("Crop — Disease")
    # Local datasets use class keys that may embed crop info; treat as-is
}

def _parse_crop_disease_from_label(ds_name, cls_label):
    """
    Return (crop, disease) for a class label given its dataset name.
    - LeafNet labels: "Crop — Disease"
    - All others: crop comes from DATASET_CROP mapping; cls_label is the disease.
    - Local datasets (no crop map): crop = dataset category name, disease = class key.
    """
    if ds_name == "LeafNet" and " — " in cls_label:
        parts = cls_label.split(" — ", 1)
        return parts[0].strip().title(), parts[1].strip().title()

    crop = DATASET_CROP.get(ds_name)
    if crop:
        disease = cls_label.replace("_", " ").strip().title()
        return crop, disease

    # Local / unknown dataset: use dataset name as crop context
    disease = cls_label.replace("_", " ").strip().title()
    return ds_name.replace("_", " ").title(), disease


def generate_xlsx(all_datasets, output_path):
    """
    Build a deduplicated crop × disease registry xlsx.

    Sheet 1 — "Crop-Disease Index":
        Crop | Disease | Total Sources | Source Datasets
        One row per unique (crop, disease) pair. Sorted by Crop then Disease.
        If the same pair appears in multiple datasets → single row, sources merged.

    Sheet 2 — "By Dataset":
        Dataset | Crop | Disease | Ref Images | Bench Images
        One row per (dataset, class). Useful for cross-checking raw counts.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side)
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [WARN] openpyxl not installed — skipping xlsx export.")
        return

    # ── Build deduplicated registry ──────────────────────────────────────────
    # pair_key (crop_lower, disease_lower) -> {crop, disease, sources: set, ref, bench}
    registry  = {}   # dedup index
    raw_rows  = []   # for Sheet 2

    for ds_name, ref_df, bench_df, classes, _ in all_datasets:
        if not classes:
            continue

        ref_counts   = {}
        bench_counts = {}
        if hasattr(ref_df, '__len__') and len(ref_df) > 0:
            ref_counts = ref_df.groupby(1).size().to_dict()
        if hasattr(bench_df, '__len__') and len(bench_df) > 0:
            bench_counts = bench_df.groupby(1).size().to_dict()

        for cls in classes:
            crop, disease = _parse_crop_disease_from_label(ds_name, cls)
            pair_key = (crop.lower(), disease.lower())

            n_ref   = ref_counts.get(cls, 0)
            n_bench = bench_counts.get(cls, 0)

            if pair_key not in registry:
                registry[pair_key] = {
                    "crop":    crop,
                    "disease": disease,
                    "sources": set(),
                    "total_ref":   0,
                    "total_bench": 0,
                }
            registry[pair_key]["sources"].add(ds_name)
            registry[pair_key]["total_ref"]   += n_ref
            registry[pair_key]["total_bench"] += n_bench

            raw_rows.append({
                "dataset": ds_name,
                "crop":    crop,
                "disease": disease,
                "ref":     n_ref,
                "bench":   n_bench,
            })

    # Sort dedup index: crop A→Z, then disease A→Z
    sorted_pairs = sorted(registry.items(), key=lambda x: x[0])

    # ── Styles ───────────────────────────────────────────────────────────────
    HDR_FILL   = PatternFill("solid", start_color="3A3A6E")   # dark indigo
    HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    HDR_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ALT_FILL_A = PatternFill("solid", start_color="EEF0FA")   # very light indigo
    ALT_FILL_B = PatternFill("solid", start_color="FFFFFF")
    BODY_FONT  = Font(name="Arial", size=9)
    BODY_ALIGN = Alignment(vertical="center", wrap_text=False)
    WRAP_ALIGN = Alignment(vertical="center", wrap_text=True)

    MULTI_FILL = PatternFill("solid", start_color="FFF3CD")   # amber — appears in 2+ sources
    MULTI_FONT = Font(name="Arial", size=9, bold=True)

    thin = Side(style="thin", color="C8C8D8")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _apply_header(ws, headers, col_widths):
        ws.append(headers)
        for col, (cell, w) in enumerate(
                zip(ws[1], col_widths), start=1):
            cell.font      = HDR_FONT
            cell.fill      = HDR_FILL
            cell.alignment = HDR_ALIGN
            cell.border    = BORDER
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

    def _style_row(ws, row_idx, n_cols, fill, font, align=None):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill   = fill
            cell.font   = font
            cell.border = BORDER
            cell.alignment = align or BODY_ALIGN

    # ── Workbook ─────────────────────────────────────────────────────────────
    wb = Workbook()

    # ════════════════════════════════════════════════════════════════════════
    #  Sheet 1 — Crop-Disease Index (deduplicated)
    # ════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Crop-Disease Index"

    hdrs1  = ["#", "Crop", "Disease", "# Sources", "Source Datasets",
              "Total Ref Imgs", "Total Bench Imgs"]
    widths1 = [5, 18, 28, 10, 42, 14, 15]
    _apply_header(ws1, hdrs1, widths1)

    for i, (pair_key, info) in enumerate(sorted_pairs, start=1):
        sources_str = ", ".join(sorted(info["sources"]))
        n_sources   = len(info["sources"])
        row_data    = [
            i,
            info["crop"],
            info["disease"],
            n_sources,
            sources_str,
            info["total_ref"],
            info["total_bench"],
        ]
        ws1.append(row_data)
        row_idx = i + 1   # 1-indexed + header

        # Highlight rows that appear in multiple sources
        is_multi = n_sources > 1
        fill  = MULTI_FILL        if is_multi else (ALT_FILL_A if i % 2 == 0 else ALT_FILL_B)
        font  = MULTI_FONT        if is_multi else BODY_FONT
        _style_row(ws1, row_idx, len(hdrs1), fill, font)
        # Sources column wraps
        ws1.cell(row=row_idx, column=5).alignment = WRAP_ALIGN
        ws1.row_dimensions[row_idx].height = 16

    # Totals row
    n_data = len(sorted_pairs)
    ws1.append([
        "", "TOTAL UNIQUE PAIRS", f"=COUNTA(B2:B{n_data+1})",
        "", "", f"=SUM(F2:F{n_data+1})", f"=SUM(G2:G{n_data+1})"
    ])
    tot_row = n_data + 2
    TOTAL_FILL = PatternFill("solid", start_color="3A3A6E")
    TOTAL_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    _style_row(ws1, tot_row, len(hdrs1), TOTAL_FILL, TOTAL_FONT,
               Alignment(horizontal="center", vertical="center"))
    ws1.cell(row=tot_row, column=2).alignment = Alignment(
        horizontal="left", vertical="center")

    # Legend note below totals
    ws1.cell(row=tot_row + 2, column=1).value = (
        "★ Amber rows = same (Crop, Disease) pair found in 2 or more source datasets")
    ws1.cell(row=tot_row + 2, column=1).font = Font(
        name="Arial", size=8, italic=True, color="7A6000")

    # ════════════════════════════════════════════════════════════════════════
    #  Sheet 2 — By Dataset (raw, one row per dataset×class)
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("By Dataset")
    hdrs2  = ["#", "Dataset", "Crop", "Disease", "Ref Images", "Bench Images"]
    widths2 = [5, 22, 18, 32, 12, 13]
    _apply_header(ws2, hdrs2, widths2)

    # Group by dataset for colour banding
    DS_COLOR_MAP = {
        ds: DATASET_COLORS.get(ds, "#888888").lstrip("#")
        for ds, *_ in all_datasets
    }

    prev_ds  = None
    band_idx = 0
    for i, row in enumerate(raw_rows, start=1):
        ds = row["dataset"]
        if ds != prev_ds:
            band_idx += 1
            prev_ds = ds
        fill_hex = DS_COLOR_MAP.get(ds, "AAAAAA")
        # Lighten: blend with white at 85%
        r = int(fill_hex[0:2], 16); g = int(fill_hex[2:4], 16)
        b = int(fill_hex[4:6], 16)
        lr = int(r * 0.18 + 255 * 0.82)
        lg = int(g * 0.18 + 255 * 0.82)
        lb = int(b * 0.18 + 255 * 0.82)
        light_hex = f"{lr:02X}{lg:02X}{lb:02X}"
        row_fill  = PatternFill("solid", start_color=light_hex)

        ws2.append([i, ds, row["crop"], row["disease"], row["ref"], row["bench"]])
        _style_row(ws2, i + 1, len(hdrs2), row_fill, BODY_FONT)

    # Summary footer
    n2 = len(raw_rows)
    ws2.append(["", "TOTAL", "", f"=COUNTA(D2:D{n2+1})",
                f"=SUM(E2:E{n2+1})", f"=SUM(F2:F{n2+1})"])
    _style_row(ws2, n2 + 2, len(hdrs2), TOTAL_FILL, TOTAL_FONT,
               Alignment(horizontal="center", vertical="center"))
    ws2.cell(row=n2 + 2, column=2).alignment = Alignment(
        horizontal="left", vertical="center")

    # ── Save ─────────────────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"  XLSX -> {os.path.abspath(output_path)}")
    print(f"         {len(sorted_pairs)} unique crop-disease pairs  |  "
          f"{len(raw_rows)} total dataset×class rows")
    multi = sum(1 for _, v in sorted_pairs if len(v["sources"]) > 1)
    if multi:
        print(f"         {multi} pair(s) appear in multiple sources (highlighted amber)")


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    os.makedirs(DATA_ROOT,   exist_ok=True)
    os.makedirs(SAMPLES_DIR, exist_ok=True)

    n_ref, n_bench = ask_sample_counts()

    all_datasets = []

    # Online datasets
    print("\n" + "="*60)
    print("  ONLINE DATASETS")
    print("="*60)
    for loader in [load_SBRD, load_MangoLeaf, load_SoybeanPNAS,
                   load_BeanLeaf, load_YellowRust,
                   load_BananaLeaf, load_Cauliflower, load_Lettuce,
                   load_LeafNet,
                   # load_FUSARIUM22,  # temporarily disabled
                   ]:
        print(f"\n-- {loader.__name__} --")
        result = loader(n_ref, n_bench)
        all_datasets.append(result)
        nm, ref, bench, cls, _ = result
        print(f"   {nm}: {len(ref)} ref | {len(bench)} bench | {len(cls)} classes")

    # Local datasets
    if LOCAL_SOURCE_ROOT:
        if os.path.isdir(LOCAL_SOURCE_ROOT):
            print("\n" + "="*60)
            print("  LOCAL DATASETS")
            print(f"  Source: {LOCAL_SOURCE_ROOT}")
            print("="*60)
            for cat, cls_map in build_local_class_map(LOCAL_SOURCE_ROOT).items():
                print(f"\n-- Local: {cat} --")
                result = load_local_category(
                    cat, cls_map, n_ref, n_bench, LOCAL_SOURCE_ROOT)
                all_datasets.append(result)
                nm, ref, bench, cls, _ = result
                print(f"   {nm}: {len(ref)} ref | {len(bench)} bench | {len(cls)} classes")
        else:
            print(f"\n[WARN] LOCAL_SOURCE_ROOT not found: {LOCAL_SOURCE_ROOT}")

    # PDF + XLSX
    print("\n" + "="*60)
    print("  BUILDING PDF + XLSX")
    print("="*60)
    generate_pdf(all_datasets, OUTPUT_PDF, n_ref, n_bench)
    generate_xlsx(all_datasets, OUTPUT_XLSX)

    # Manifest
    manifest = {
        "n_reference_per_class":  n_ref,
        "n_benchmark_per_class":  n_bench,
        "datasets": [
            {"name": r[0], "classes": len(r[3]),
             "reference_images":  len(r[1]) if hasattr(r[1],'__len__') else 0,
             "benchmark_images":  len(r[2]) if hasattr(r[2],'__len__') else 0}
            for r in all_datasets
        ]
    }
    mpath = os.path.join(SAMPLES_DIR, "manifest.json")
    with open(mpath, "w") as f:
        json.dump(manifest, f, indent=2)
    print(f"  Manifest -> {os.path.abspath(mpath)}\n")


if __name__ == "__main__":
    main()